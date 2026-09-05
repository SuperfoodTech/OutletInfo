#!/usr/bin/env python3
"""
export_and_upload_drive.py
==========================
Menggabungkan data outlet per-owner dari multi-aplikator (GoFood dan GrabFood,
serta siap untuk ShopeeFood) ke dalam satu file per-owner:
    YYYY-MM-DD HH_MM [Nama Pemilik].xlsx

Spesifikasi File:
- Menggunakan template standar 37 kolom (sheet 'Listing').
- Memiliki 2 tab (sheet) yang 100% identik:
    Tab 1: 'Listing'
    Tab 2: 'Listing 2'
- Disimpan lokal di folder `output_owners/`.
- Diunggah ke Google Drive ke dalam folder per-owner (membuat folder otomatis jika belum ada)
  melalui Google Apps Script Web App (APP_SCRIPT_URL).
"""

import os
import sys
import glob
import json
import base64
import argparse
import datetime
import urllib.request
import io
import requests
from pathlib import Path
from dotenv import load_dotenv

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ─── Konfigurasi Direktori & URL ───────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
GOFOOD_DIR = BASE_DIR / "GOFOOD"
GRAB_DIR = BASE_DIR / "GRAB"
SHOPEE_DIR = BASE_DIR / "SHOPEE"
CACHE_DIR = BASE_DIR / "cache"
OUTPUT_OWNERS_DIR = BASE_DIR / "output_owners"
TEMPLATE_PATH = BASE_DIR / "YYYY-MM-DD HH_MM Nama Pemilik.xlsx"

GOOGLE_SHEET_VERCEL_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTprbPPf_J5gAVL3PYeHbbdl5ZXQvb17HY2lJGPI2xg13Ly3AGT8eYHLYmU_m1NdtkBVg-qUGv1BoEE/pub?output=csv"

# Muat variabel lingkungan
load_dotenv(BASE_DIR / ".env")
APP_SCRIPT_URL = os.getenv("APP_SCRIPT_URL", "")


def get_template_headers():
    """Mengambil 37 kolom header resmi dari file template."""
    if TEMPLATE_PATH.exists():
        try:
            wb = openpyxl.load_workbook(str(TEMPLATE_PATH), read_only=True)
            sheet = wb["Listing"] if "Listing" in wb.sheetnames else wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            wb.close()
            return headers
        except Exception as e:
            print(f"⚠️ Gagal membaca template ({e}), menggunakan default 37 header.")
            
    # Fallback default 37 kolom jika template tidak terbaca
    return [
        'Nama Pemilik', 'Nama Brand', 'Model', 'Tipe', 'Outlet', 'Nomor HP',
        'Aplikator', 'Nama Portal', 'Group ID', 'Nama Listing', 'Link', 'Store ID',
        'Status Listing', 'Alamat', 'Nama Bank', 'Nama Pemilik Rekening', 'Nomor Rekening',
        'Nama Akses', 'Email FoodMaster1', 'Email FoodMaster2', 'Nama Pengguna',
        'Kata Sandi', 'Nama Portal.1', 'S Nomor HP Akses Pemilik', 'S Username Akses Pemilik',
        'S Kata Sandi Akses Pemilik', 'S Allvbadmin Username Akses Staff',
        'S Allvbadmin Kata Sandi Akses Staff', 'S Bot Username Akses Staff',
        'S Bot Kata Sandi Akses Staff', 'S BD Username Akses Staff',
        'S BD Kata Sandi Akses Staff', 'BD', 'Status Internal', 'Tanggal Live',
        'Tanggal Churn', 'Tarif'
    ]


def load_gofood_data():
    """Memuat data outlet GoFood dari master file dan cache JSON terbaru."""
    records = []
    
    # 1. Dari master file
    gofood_master = GOFOOD_DIR / "master" / "0master.xlsx"
    if not gofood_master.exists():
        masters = sorted(glob.glob(str(GOFOOD_DIR / "master" / "*_master.xlsx")))
        if masters:
            gofood_master = Path(masters[-1])

    if gofood_master.exists():
        try:
            df = pd.read_excel(str(gofood_master), sheet_name="Listing")
            df["Aplikator"] = "GoFood"
            records.append(df)
        except Exception as e:
            print(f"⚠️ Gagal membaca GoFood master: {e}")

    # 2. Dari cache JSON di GOFOOD/cache/
    for jf in glob.glob(str(GOFOOD_DIR / "cache" / "*.json")):
        try:
            with open(jf, "r", encoding="utf-8") as f:
                cdata = json.load(f)
            if isinstance(cdata, dict) and "outlets" in cdata and isinstance(cdata["outlets"], list):
                if cdata["outlets"]:
                    df_c = pd.DataFrame(cdata["outlets"])
                    df_c["Aplikator"] = "GoFood"
                    records.append(df_c)
        except Exception:
            pass

    # 3. Fallback dari GOFOOD/output/*.xlsx
    for f in glob.glob(str(GOFOOD_DIR / "output" / "*.xlsx")):
        try:
            df_part = pd.read_excel(f, sheet_name="Listing")
            df_part["Aplikator"] = "GoFood"
            records.append(df_part)
        except Exception:
            pass

    if records:
        combined = pd.concat(records, ignore_index=True)
        if "Store ID" in combined.columns:
            combined = combined.drop_duplicates(subset=["Store ID"], keep="last")
        return combined
        
    return pd.DataFrame()


def load_grab_data():
    """Memuat data outlet GrabFood dari master file dan cache JSON terbaru."""
    records = []
    
    # 1. Dari master file
    grab_master = GRAB_DIR / "master" / "0master.xlsx"
    if not grab_master.exists():
        masters = sorted(glob.glob(str(GRAB_DIR / "master" / "*_master.xlsx")))
        if masters:
            grab_master = Path(masters[-1])

    if grab_master.exists():
        try:
            df = pd.read_excel(str(grab_master), sheet_name="Listing")
            df["Aplikator"] = "GrabFood"
            records.append(df)
        except Exception as e:
            print(f"⚠️ Gagal membaca Grab master: {e}")

    # 2. Dari cache JSON di GRAB/cache/
    for jf in glob.glob(str(GRAB_DIR / "cache" / "*.json")):
        try:
            with open(jf, "r", encoding="utf-8") as f:
                cdata = json.load(f)
            if isinstance(cdata, dict) and "outlets" in cdata and isinstance(cdata["outlets"], list):
                if cdata["outlets"]:
                    df_c = pd.DataFrame(cdata["outlets"])
                    df_c["Aplikator"] = "GrabFood"
                    records.append(df_c)
        except Exception:
            pass

    # 3. Fallback dari GRAB/output/*.xlsx
    for f in glob.glob(str(GRAB_DIR / "output" / "*.xlsx")):
        try:
            df_part = pd.read_excel(f, sheet_name="Listing")
            df_part["Aplikator"] = "GrabFood"
            records.append(df_part)
        except Exception:
            pass

    if records:
        combined = pd.concat(records, ignore_index=True)
        if "Store ID" in combined.columns:
            combined = combined.drop_duplicates(subset=["Store ID"], keep="last")
        return combined

    return pd.DataFrame()


def load_shopee_data():
    """Memuat data outlet ShopeeFood dari master file atau output terbaru."""
    shopee_master = SHOPEE_DIR / "data" / "0master.xlsx"
    if not shopee_master.exists():
        masters = sorted(glob.glob(str(SHOPEE_DIR / "data" / "Shopee_*.xlsx")), key=os.path.getmtime)
        if masters:
            shopee_master = Path(masters[-1])

    if shopee_master.exists():
        try:
            df = pd.read_excel(str(shopee_master), sheet_name="Listing")
            df["Aplikator"] = "ShopeeFood"
            return df
        except Exception as e:
            print(f"⚠️ Gagal membaca Shopee master: {e}")

    return pd.DataFrame()


def load_vercel_data():
    """
    Memuat data outlet dari Google Sheet Vercel (Live CSV).
    Menyimpan cache lokal di cache/vercel_sheet_cache.csv untuk fallback offline.
    """
    cache_file = CACHE_DIR / "vercel_sheet_cache.csv"
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    csv_text = ""
    # 1. Coba fetch live dari Google Sheet jika memungkinkan
    try:
        req = urllib.request.Request(GOOGLE_SHEET_VERCEL_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            csv_text = resp.read().decode('utf-8', errors='replace')
        if csv_text.strip():
            with open(cache_file, "w", encoding="utf-8") as f:
                f.write(csv_text)
    except Exception as e:
        print(f"⚠️ Fetch live Vercel Sheet melewati batas waktu / offline ({e}). Menggunakan cache lokal.")

    # 2. Fallback ke file cache lokal
    if not csv_text and cache_file.exists():
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                csv_text = f.read()
        except Exception:
            pass

    if not csv_text:
        return pd.DataFrame()

    try:
        df = pd.read_csv(io.StringIO(csv_text))
    except Exception as e:
        print(f"⚠️ Gagal membaca CSV Vercel: {e}")
        return pd.DataFrame()

    # Bersihkan whitespace di nama kolom
    df.columns = [str(c).strip() for c in df.columns]

    # Map nama kolom Vercel ke nama standar internal
    if "Owner" in df.columns:
        df["Nama Pemilik"] = df["Owner"].astype(str).str.strip()
    if "Nama Outlet" in df.columns:
        df["Nama Brand"] = df["Nama Outlet"].astype(str).str.strip()
        df["Nama Listing"] = df["Nama Brand"]

    def clean_app(val):
        s = str(val).strip().lower()
        if "gofood" in s or "go" in s:
            return "GoFood"
        elif "grab" in s:
            return "GrabFood"
        elif "shopee" in s:
            return "ShopeeFood"
        return str(val).strip()

    if "Aplikasi" in df.columns:
        df["Aplikator"] = df["Aplikasi"].apply(clean_app)

    if "Nama Akses" in df.columns:
        df["Nama Portal"] = df["Nama Akses"].astype(str).str.strip()

    # Untuk Shopee, fallback Nama Portal ke Merchant Name
    if "Merchant Name" in df.columns:
        df["Nama Portal"] = df.apply(
            lambda r: r["Merchant Name"] if (pd.isna(r.get("Nama Portal")) or not str(r.get("Nama Portal")).strip()) and pd.notna(r.get("Merchant Name")) else r.get("Nama Portal", ""),
            axis=1
        )

    # Status Listing default LIVE untuk data aktif
    df["Status Listing"] = "LIVE"

    return df


def write_data_to_sheet(ws, df, headers):
    """Menulis data ke sheet yang telah memiliki header, styling, dan lebar kolom dari template."""
    body_font = Font(name='Arial', size=10)
    alignment = Alignment(horizontal='left', vertical='center')

    # Alias mapping jika nama kolom di df sedikit berbeda dengan nama header template
    col_aliases = {
        'Nama Pemilik': ['Owner', 'Nama Pemilik', 'owner', 'pemilik'],
        'Nama Brand': ['Nama Outlet', 'Nama Brand', 'Brand', 'brand'],
        'Aplikator': ['Aplikasi', 'Aplikator', 'app', 'Platform'],
        'Nama Portal': ['Nama Akses', 'Nama Portal', 'Merchant Name', 'portal'],
        'Nama Listing': ['Nama Listing', 'Nama Outlet', 'Nama Brand', 'Listing'],
        'Nomor HP Akses Pemilik': ['S Nomor HP Akses Pemilik', 'Nomor HP Akses Pemilik', 'Nomor HP'],
        'Username Akses Pemilik': ['S Username Akses Pemilik', 'Username Akses Pemilik'],
        'Kata Sandi Akses Pemilik': ['S Kata Sandi Akses Pemilik', 'Kata Sandi Akses Pemilik'],
        'S BD Username Akses Staff': ['S Username Akses Staff', 'S BD Username Akses Staff', 'Username Akses Staff'],
        'S BD Kata Sandi Akses Staff': ['S Kata Sandi Akses Staff', 'S BD Kata Sandi Akses Staff', 'Kata Sandi Akses Staff'],
        'BD': ['BD', 'bd'],
    }

    text_cols = {'Nomor Rekening', 'Nomor HP', 'S Nomor HP Akses Pemilik', 'Store ID', 'Group ID'}

    # Tulis Baris Data mulai baris ke-2
    for idx, (_, r) in enumerate(df.iterrows(), start=2):
        for c_idx, h in enumerate(headers, 1):
            if not h:
                continue
            val = r.get(h)
            if val is None or pd.isna(val) or str(val).strip() == '':
                # Cari via aliases
                aliases = col_aliases.get(h, [])
                for alias in aliases:
                    if alias in r and pd.notna(r.get(alias)) and str(r.get(alias)).strip() != '':
                        val = r.get(alias)
                        break

            val_str = str(val).strip() if pd.notna(val) and val is not None else ''
            if val_str.lower() in ('nan', 'none'):
                val_str = ''

            # Bersihkan suffix .0 pada angka/nomor panjang
            if val_str.endswith(".0") and val_str[:-2].isdigit():
                val_str = val_str[:-2]

            cell = ws.cell(row=idx, column=c_idx, value=val_str)
            cell.font = body_font
            cell.alignment = alignment
            if h in text_cols or c_idx in (6, 9, 12, 17, 24):
                cell.number_format = '@'


def save_owner_workbook(owner_df, file_path, headers):
    """
    Menyimpan data per-owner ke file Excel dengan 2 tab 100% identik:
    - Tab 1: 'Listing'
    - Tab 2: 'Listing 2'
    Dibuat langsung dengan mengkloning template 'YYYY-MM-DD HH_MM Nama Pemilik.xlsx'
    sehingga pewarnaan kolom header (Merah, Pink, Hijau, Oranye), font Arial,
    dan lebar kolom asli terjaga 100% identik.
    """
    if TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
        if "Listing" in wb.sheetnames:
            ws1 = wb["Listing"]
        else:
            ws1 = wb.active
            ws1.title = "Listing"
        
        # Bersihkan baris dummy template di bawah header
        if ws1.max_row > 1:
            ws1.delete_rows(2, ws1.max_row)

        if "Listing 2" in wb.sheetnames:
            del wb["Listing 2"]

        ws2 = wb.copy_worksheet(ws1)
        ws2.title = "Listing 2"
    else:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Listing"
        ws2 = wb.create_sheet(title="Listing 2")

    write_data_to_sheet(ws1, owner_df, headers)
    write_data_to_sheet(ws2, owner_df, headers)
    
    wb.save(file_path)


def upload_file_to_drive(file_path, owner_name, app_script_url):
    """Mengunggah file Excel per-owner ke Google Drive via Google Apps Script Web App."""
    if not app_script_url:
        print("   ⚠️ APP_SCRIPT_URL belum disetel di .env. Lewati upload.")
        return False, {"error": "APP_SCRIPT_URL belum disetel di .env"}

    filename = os.path.basename(file_path)
    print(f"   ☁️ Mengunggah {filename} ke folder Google Drive: [{owner_name}]...")
    
    try:
        with open(file_path, "rb") as f:
            b64_data = base64.b64encode(f.read()).decode("utf-8")

        payload = {
            "fileName": filename,
            "ownerName": owner_name,
            "fileData": b64_data,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }

        response = requests.post(app_script_url, json=payload, timeout=90, allow_redirects=True)
        if response.status_code in (200, 201):
            try:
                res_data = response.json()
                if res_data.get("status") == "success":
                    folder_url = res_data.get("folderUrl") or res_data.get("url", "")
                    file_url = res_data.get("fileUrl", "")
                    print(f"   ✅ Berhasil diunggah! Folder: {folder_url}")
                    return True, {
                        "folderUrl": folder_url,
                        "fileUrl": file_url,
                        "folderName": res_data.get("folderName", owner_name),
                        "fileName": filename
                    }
                else:
                    msg = res_data.get("message", response.text)
                    print(f"   ❌ Gagal: {msg}")
                    return False, {"error": msg}
            except Exception:
                print("   ✅ Berhasil diunggah (respons non-JSON 200).")
                return True, {"folderUrl": "https://drive.google.com/drive/u/0/folders/19VIrypPcBmNNbjDLGS7kxp_yIdBBwXjB", "fileName": filename}
        else:
            print(f"   ❌ HTTP Error {response.status_code}: {response.text}")
            return False, {"error": f"HTTP {response.status_code}"}

    except Exception as e:
        print(f"   ⚠️ Exception saat mengunggah: {e}")
        return False, {"error": str(e)}


def get_owners_with_metadata(source="vercel"):
    """
    Mengambil daftar owner beserta informasi outlet dan timestamp file terbaru.
    source: 'vercel' (default) menggunakan live CSV Vercel Sheet,
            'master' menggunakan master lokal.
    """
    if source == "vercel":
        combined = load_vercel_data()
    else:
        df_go = load_gofood_data()
        df_grab = load_grab_data()
        df_shopee = load_shopee_data()
        dfs = [df for df in [df_go, df_grab, df_shopee] if not df.empty]
        combined = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    
    if combined.empty or "Nama Pemilik" not in combined.columns:
        # Fallback jika Vercel kosong
        df_go = load_gofood_data()
        df_grab = load_grab_data()
        df_shopee = load_shopee_data()
        dfs = [df for df in [df_go, df_grab, df_shopee] if not df.empty]
        combined = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    if combined.empty or "Nama Pemilik" not in combined.columns:
        return []

    combined = combined[combined["Nama Pemilik"].notna()]
    combined = combined[combined["Nama Pemilik"].astype(str).str.strip() != ""]
    combined = combined[combined["Nama Pemilik"].astype(str).str.strip().str.lower() != "nan"]
    
    # Pertahankan urutan unik
    seen = set()
    owners = []
    for o in combined["Nama Pemilik"]:
        o_clean = str(o).strip()
        if o_clean and o_clean.lower() not in seen:
            seen.add(o_clean.lower())
            owners.append(o_clean)

    results = []
    for owner_str in owners:
        o_df = combined[combined["Nama Pemilik"].str.strip().str.lower() == owner_str.lower()]
        go_n = len(o_df[o_df["Aplikator"] == "GoFood"]) if "Aplikator" in o_df.columns else 0
        gr_n = len(o_df[o_df["Aplikator"] == "GrabFood"]) if "Aplikator" in o_df.columns else 0
        sh_n = len(o_df[o_df["Aplikator"] == "ShopeeFood"]) if "Aplikator" in o_df.columns else 0
        
        # Cari timestamp terbaru dari file owner di output_owners/
        clean_name = "".join(c for c in owner_str if c.isalnum() or c in " ._-").strip()
        files = glob.glob(str(OUTPUT_OWNERS_DIR / f"*{clean_name}*.xlsx"))
        files.extend(glob.glob(str(GOFOOD_DIR / "output" / f"*{clean_name}*.xlsx")))
        files.extend(glob.glob(str(GRAB_DIR / "output" / f"*{clean_name}*.xlsx")))
        files.extend(glob.glob(str(SHOPEE_DIR / "data" / f"*{clean_name}*.xlsx")))
        
        latest_mtime = 0
        latest_ts_str = "Belum Ada"
        if files:
            latest_file = max(files, key=os.path.getmtime)
            latest_mtime = os.path.getmtime(latest_file)
            latest_ts_str = datetime.datetime.fromtimestamp(latest_mtime).strftime("%Y-%m-%d %H:%M")

        results.append({
            "owner": owner_str,
            "total": len(o_df),
            "gofood": go_n,
            "grab": gr_n,
            "shopee": sh_n,
            "latest_mtime": latest_mtime,
            "latest_ts": latest_ts_str,
            "is_newest": False
        })

    # Urutkan berdasarkan timestamp modifikasi jika ada file baru
    results.sort(key=lambda x: x["latest_mtime"], reverse=True)
    if results and results[0]["latest_mtime"] > 0:
        results[0]["is_newest"] = True

    return results


def run_live_scraping_for_owner(owner_name, aplikator="all", progress_cb=None):
    """
    Menjalankan live scraping GoFood, Grab, dan Shopee untuk owner spesifik.
    Menggunakan subprocess agar proses terisolasi dan log dapat di-stream real-time.
    """
    import subprocess
    clean_owner = owner_name.strip()
    python_bin = sys.executable

    def send_log(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)
        print(f"[{pct:>3}%] {msg}")

    # 1. LIVE SCRAPING GOFOOD
    if aplikator in ("all", "gofood"):
        send_log(15, f"🚀 [GoFood] Memulai penarikan live untuk '{clean_owner}'...")
        cmd = [
            python_bin,
            str(GOFOOD_DIR / "gofood_scraper.py"),
            "--vercel",
            "--owner", clean_owner,
            "--headless"
        ]
        try:
            p = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                cwd=str(GOFOOD_DIR),
                bufsize=1
            )
            for line in iter(p.stdout.readline, ''):
                line_s = line.strip()
                if line_s:
                    if any(k in line_s for k in ("Store ID", "Berhasil", "Portal", "Login", "Owner", "Restricted", "Memproses", "OTP", "Filter")):
                        send_log(25, f"[GoFood] {line_s[:85]}")
            p.wait()
            if p.returncode == 0:
                send_log(35, f"✅ [GoFood] Selesai memproses '{clean_owner}'.")
            else:
                send_log(35, f"⚠️ [GoFood] Selesai dengan status {p.returncode}.")
        except Exception as e:
            send_log(35, f"⚠️ [GoFood] Gagal menjalankan scraper: {e}")

    # 2. LIVE SCRAPING GRABFOOD
    if aplikator in ("all", "grab"):
        send_log(40, f"🚀 [GrabFood] Memulai penarikan live untuk '{clean_owner}'...")
        cmd = [
            python_bin,
            str(GRAB_DIR / "grab_merchant_scraper.py"),
            "--vercel",
            "--owner", clean_owner
        ]
        try:
            p = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                cwd=str(GRAB_DIR),
                bufsize=1
            )
            for line in iter(p.stdout.readline, ''):
                line_s = line.strip()
                if line_s:
                    if any(k in line_s for k in ("Group ID", "Berhasil", "Target", "Portal", "Login", "Store", "Bank", "Owner", "Filter")):
                        send_log(55, f"[Grab] {line_s[:85]}")
            p.wait()
            if p.returncode == 0:
                send_log(65, f"✅ [GrabFood] Selesai memproses '{clean_owner}'.")
            else:
                send_log(65, f"⚠️ [GrabFood] Selesai dengan status {p.returncode}.")
        except Exception as e:
            send_log(65, f"⚠️ [GrabFood] Gagal menjalankan scraper: {e}")

    # 3. LIVE SCRAPING SHOPEEFOOD
    if aplikator in ("all", "shopee"):
        send_log(68, f"🚀 [ShopeeFood] Memeriksa akun Shopee untuk '{clean_owner}'...")
        try:
            v_df = load_vercel_data()
            o_shopee = v_df[(v_df["Nama Pemilik"].astype(str).str.strip().str.lower() == clean_owner.lower()) & (v_df["Aplikator"] == "ShopeeFood")]
            merchant_name = ""
            if not o_shopee.empty:
                for col in ["Merchant Name", "Nama Portal", "Nama Brand", "Nama Akses"]:
                    if col in o_shopee.columns and pd.notna(o_shopee.iloc[0].get(col)) and str(o_shopee.iloc[0].get(col)).strip():
                        merchant_name = str(o_shopee.iloc[0].get(col)).strip()
                        break
            
            if merchant_name:
                send_log(70, f"[ShopeeFood] Menarik data toko '{merchant_name}'...")
                cmd = [
                    python_bin,
                    str(SHOPEE_DIR / "pull_outlet_info.py"),
                    "--merchant-name", merchant_name
                ]
                p = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    cwd=str(SHOPEE_DIR),
                    bufsize=1
                )
                for line in iter(p.stdout.readline, ''):
                    line_s = line.strip()
                    if line_s and any(k in line_s for k in ("Store", "Berhasil", "Merchant", "Data", "Sukses")):
                        send_log(74, f"[Shopee] {line_s[:85]}")
                p.wait()
                send_log(76, f"✅ [ShopeeFood] Selesai memproses '{merchant_name}'.")
            else:
                send_log(76, f"ℹ️ [ShopeeFood] Tidak ditemukan nama merchant Shopee untuk '{clean_owner}'.")
        except Exception as e:
            send_log(76, f"⚠️ [ShopeeFood] Exception scraper: {e}")


def generate_for_owner_pipeline(owner_name, aplikator="all", upload=True, source="vercel", live_scrape=True, progress_callback=None):
    """
    Pipeline pembuatan file per-owner dan upload Drive dengan progress callback.
    Menggunakan Vercel Sheet sebagai sumber utama dan memperkaya data dengan hasil live scraping.
    
    aplikator: 'all' | 'gofood' | 'grab' | 'shopee'
    source: 'vercel' (default) | 'master'
    live_scrape: True (default) menjalankan scraping live | False (kompilasi cache saja)
    progress_callback: function(percent: int, log_line: str)
    """
    def log(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        print(f"[{pct:>3}%] {msg}")

    log(10, "Inisialisasi direktori dan template 37 kolom...")
    headers = get_template_headers()
    OUTPUT_OWNERS_DIR.mkdir(parents=True, exist_ok=True)

    # 1. LIVE SCRAPING (Jika diaktifkan dan bukan mode __ALL__)
    if live_scrape and owner_name != "__ALL__":
        log(12, f"Menjalankan penarikan live data toko untuk '{owner_name}'...")
        run_live_scraping_for_owner(owner_name, aplikator=aplikator, progress_cb=progress_callback)

    log(78, f"Membaca data sumber ({source.upper()}) untuk aplikator: {aplikator.upper()}...")
    if source == "vercel":
        src_df = load_vercel_data()
    else:
        dfs = []
        if aplikator in ("all", "gofood"):
            df_go = load_gofood_data()
            if not df_go.empty:
                dfs.append(df_go)
        if aplikator in ("all", "grab"):
            df_grab = load_grab_data()
            if not df_grab.empty:
                dfs.append(df_grab)
        if aplikator in ("all", "shopee"):
            df_shopee = load_shopee_data()
            if not df_shopee.empty:
                dfs.append(df_shopee)
        src_df = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    if src_df.empty:
        log(100, "❌ Data sumber kosong!")
        return {"success": False, "error": "Data sumber kosong"}

    log(80, f"Memfilter data untuk Owner: '{owner_name}'...")
    if owner_name == "__ALL__":
        owner_df = src_df.copy()
    else:
        owner_df = src_df[src_df["Nama Pemilik"].astype(str).str.strip().str.lower() == owner_name.strip().lower()].copy()

    if owner_df.empty:
        log(100, f"⚠️ Tidak ditemukan data untuk owner '{owner_name}'.")
        return {"success": False, "error": f"Owner '{owner_name}' tidak ditemukan dalam data"}

    # Filter aplikator
    if aplikator == "gofood":
        owner_df = owner_df[owner_df["Aplikator"] == "GoFood"]
    elif aplikator == "grab":
        owner_df = owner_df[owner_df["Aplikator"] == "GrabFood"]
    elif aplikator == "shopee":
        owner_df = owner_df[owner_df["Aplikator"] == "ShopeeFood"]

    if owner_df.empty:
        log(100, f"⚠️ Tidak ada outlet {aplikator.upper()} untuk owner '{owner_name}'.")
        return {"success": False, "error": f"Tidak ada outlet {aplikator} untuk owner {owner_name}"}

    # Enrich data dari hasil scraping jika tersedia (Bank, Rekening, Store ID, Alamat)
    log(82, "Menggabungkan hasil penarikan data toko (Store ID, Alamat, Bank)...")
    try:
        df_go = load_gofood_data()
        df_gr = load_grab_data()
        df_sh = load_shopee_data()
        scraped_list = [d for d in [df_go, df_gr, df_sh] if not d.empty]
        scraped_all = pd.concat(scraped_list, ignore_index=True) if scraped_list else pd.DataFrame()

        scraped_owner = pd.DataFrame()
        if not scraped_all.empty and "Nama Pemilik" in scraped_all.columns:
            scraped_owner = scraped_all[scraped_all["Nama Pemilik"].astype(str).str.strip().str.lower() == owner_name.strip().lower()]

        final_rows = []
        for app in owner_df["Aplikator"].dropna().unique():
            app_vercel_rows = owner_df[owner_df["Aplikator"] == app]
            app_scraped_rows = scraped_owner[scraped_owner["Aplikator"] == app] if not scraped_owner.empty and "Aplikator" in scraped_owner.columns else pd.DataFrame()

            if not app_scraped_rows.empty:
                v_first = app_vercel_rows.iloc[0]
                for _, s_row in app_scraped_rows.iterrows():
                    row_dict = s_row.to_dict()
                    for col in ["Nama Akses", "Email FoodMaster1", "Email FoodMaster2", "Nama Pengguna", "Kata Sandi",
                                "Nama Portal.1", "S Nomor HP Akses Pemilik", "S Username Akses Pemilik", "S Kata Sandi Akses Pemilik",
                                "S Allvbadmin Username Akses Staff", "S Allvbadmin Kata Sandi Akses Staff",
                                "S Bot Username Akses Staff", "S Bot Kata Sandi Akses Staff",
                                "S BD Username Akses Staff", "S BD Kata Sandi Akses Staff", "BD", "Status Internal"]:
                        if col in v_first and pd.notna(v_first[col]) and str(v_first[col]).strip() != "":
                            if col not in row_dict or pd.isna(row_dict.get(col)) or str(row_dict.get(col)).strip() == "":
                                row_dict[col] = v_first[col]
                    final_rows.append(row_dict)
            else:
                for _, v_row in app_vercel_rows.iterrows():
                    final_rows.append(v_row.to_dict())

        if final_rows:
            owner_df = pd.DataFrame(final_rows)
    except Exception as e:
        print(f"⚠️ Info enrich scraped: {e}")

    # Cek bank_acc.json untuk Shopee jika rekening belum terisi
    bank_json_path = SHOPEE_DIR / "bank_acc.json"
    if bank_json_path.exists():
        try:
            with open(bank_json_path, "r") as f:
                b_acc = json.load(f)
            for idx, row in owner_df.iterrows():
                if row.get("Aplikator") == "ShopeeFood":
                    if pd.isna(row.get("Nama Bank")) or not str(row.get("Nama Bank")).strip():
                        owner_df.at[idx, "Nama Bank"] = b_acc.get("BANK_NAME", "")
                    if pd.isna(row.get("Nama Pemilik Rekening")) or not str(row.get("Nama Pemilik Rekening")).strip():
                        owner_df.at[idx, "Nama Pemilik Rekening"] = b_acc.get("BANK_ACCOUNT_NAME", "")
                    if pd.isna(row.get("Nomor Rekening")) or not str(row.get("Nomor Rekening")).strip():
                        owner_df.at[idx, "Nomor Rekening"] = b_acc.get("BANK_ACCOUNT", "")
        except Exception:
            pass

    go_n = len(owner_df[owner_df["Aplikator"] == "GoFood"]) if "Aplikator" in owner_df.columns else 0
    gr_n = len(owner_df[owner_df["Aplikator"] == "GrabFood"]) if "Aplikator" in owner_df.columns else 0
    sh_n = len(owner_df[owner_df["Aplikator"] == "ShopeeFood"]) if "Aplikator" in owner_df.columns else 0

    log(85, f"Ditemukan {len(owner_df)} outlet (GoFood: {go_n}, Grab: {gr_n}, Shopee: {sh_n}).")

    log(80, "Menyusun file Excel dengan 2 Tab Identik ('Listing' & 'Listing 2')...")
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
    clean_owner = "".join(c for c in owner_name if c.isalnum() or c in " ._-").strip()
    filename = f"{timestamp_str} {clean_owner}.xlsx"
    file_path = OUTPUT_OWNERS_DIR / filename

    save_owner_workbook(owner_df, str(file_path), headers)
    log(85, f"File lokal berhasil dibuat: {filename}")

    drive_res = {}
    if upload:
        log(90, f"Mengunggah ke folder Google Drive: [{clean_owner}]...")
        ok, res = upload_file_to_drive(str(file_path), clean_owner, APP_SCRIPT_URL)
        if ok:
            drive_res = res
            log(100, f"✅ Sukses terunggah ke Google Drive!")
        else:
            log(100, f"⚠️ Gagal upload ke Drive: {res.get('error', 'Unknown')}")
            return {
                "success": False,
                "owner": owner_name,
                "filename": filename,
                "filepath": str(file_path),
                "total": len(owner_df),
                "gofood": go_n,
                "grab": gr_n,
                "shopee": sh_n,
                "error": res.get("error")
            }
    else:
        log(100, "✅ Selesai (Mode Lokal Saja).")

    return {
        "success": True,
        "owner": owner_name,
        "filename": filename,
        "filepath": str(file_path),
        "total": len(owner_df),
        "gofood": go_n,
        "grab": gr_n,
        "shopee": sh_n,
        "folder_url": drive_res.get("folderUrl") or "https://drive.google.com/drive/u/0/folders/19VIrypPcBmNNbjDLGS7kxp_yIdBBwXjB",
        "file_url": drive_res.get("fileUrl", "")
    }


def process_and_combine(owner_filter=None, upload=False):
    """Proses utama penggabungan per-owner dan upload ke Google Drive."""
    print("=" * 65)
    print("🚀 PENGGABUNGAN DATA OUTLET PER-OWNER (GOFOOD, GRAB, SHOPEE)")
    print("=" * 65)

    headers = get_template_headers()
    OUTPUT_OWNERS_DIR.mkdir(parents=True, exist_ok=True)

    print("\n[*] Membaca data master...")
    df_go = load_gofood_data()
    df_grab = load_grab_data()
    df_shopee = load_shopee_data()

    print(f"    - GoFood outlets   : {len(df_go)} baris")
    print(f"    - GrabFood outlets : {len(df_grab)} baris")
    print(f"    - ShopeeFood outlets: {len(df_shopee)} baris")

    dfs = [df for df in [df_go, df_grab, df_shopee] if not df.empty]
    if not dfs:
        print("\n❌ Tidak ada data outlet yang ditemukan.")
        return

    # Gabungkan data seluruh platform
    combined_df = pd.concat(dfs, ignore_index=True)
    
    # Filter baris yang memiliki Nama Pemilik
    if "Nama Pemilik" not in combined_df.columns:
        print("❌ Kolom 'Nama Pemilik' tidak ditemukan dalam data.")
        return

    combined_df = combined_df[combined_df["Nama Pemilik"].notna()]
    combined_df = combined_df[combined_df["Nama Pemilik"].astype(str).str.strip() != ""]

    all_owners = sorted(combined_df["Nama Pemilik"].unique(), key=lambda x: str(x).lower())
    print(f"\n[✓] Ditemukan {len(all_owners)} unique owner lintas platform.")

    # Filter owner jika ditentukan
    if owner_filter:
        owner_filter_lower = owner_filter.strip().lower()
        filtered = [o for o in all_owners if owner_filter_lower in str(o).lower()]
        if not filtered:
            print(f"⚠️ Tidak ditemukan owner dengan kata kunci: '{owner_filter}'")
            print("Daftar owner yang ada:")
            for o in all_owners[:20]:
                print(f"  - {o}")
            if len(all_owners) > 20:
                print(f"  ... dan {len(all_owners) - 20} lainnya.")
            return
        all_owners = filtered
        print(f"[*] Memproses {len(all_owners)} owner yang cocok dengan filter '{owner_filter}'.")

    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
    success_count = 0
    upload_count = 0

    print("\n" + "─" * 65)
    for idx, owner in enumerate(all_owners, 1):
        owner_str = str(owner).strip()
        clean_owner = "".join(c for c in owner_str if c.isalnum() or c in " ._-").strip()
        filename = f"{timestamp_str} {clean_owner}.xlsx"
        file_path = OUTPUT_OWNERS_DIR / filename

        owner_df = combined_df[combined_df["Nama Pemilik"] == owner]
        # Deduplikasi per Store ID jika ada duplikat dalam aplikator yang sama
        if "Store ID" in owner_df.columns and "Aplikator" in owner_df.columns:
            owner_df = owner_df.drop_duplicates(subset=["Aplikator", "Store ID"], keep="first")

        go_count = len(owner_df[owner_df["Aplikator"] == "GoFood"])
        grab_count = len(owner_df[owner_df["Aplikator"] == "GrabFood"])
        shopee_count = len(owner_df[owner_df["Aplikator"] == "ShopeeFood"])

        print(f"\n[{idx}/{len(all_owners)}] 👤 Owner: {owner_str}")
        print(f"     📊 Total Outlet: {len(owner_df)} (GoFood: {go_count}, Grab: {grab_count}, Shopee: {shopee_count})")

        # Simpan file dengan 2 tab 100% identik
        save_owner_workbook(owner_df, str(file_path), headers)
        print(f"     💾 Disimpan: {file_path.name} (Tab 'Listing' & 'Listing 2')")
        success_count += 1

        # Upload ke Google Drive jika diaktifkan
        if upload:
            ok, _ = upload_file_to_drive(str(file_path), owner_str, APP_SCRIPT_URL)
            if ok:
                upload_count += 1

    print("\n" + "=" * 65)
    print("🎉 PROSES SELESAI")
    print(f"   • Total file dibuat lokal: {success_count} file di output_owners/")
    if upload:
        print(f"   • Total berhasil diunggah ke Google Drive: {upload_count}/{success_count}")
    else:
        print("   • Mode Lokal: File belum diunggah ke Google Drive (gunakan flag --upload untuk mengunggah).")
    print("=" * 65)


def list_owners():
    """Tampilkan daftar owner dan jumlah outlet per platform."""
    df_go = load_gofood_data()
    df_grab = load_grab_data()
    df_shopee = load_shopee_data()
    dfs = [df for df in [df_go, df_grab, df_shopee] if not df.empty]
    if not dfs:
        print("Data kosong.")
        return
    combined = pd.concat(dfs, ignore_index=True)
    if "Nama Pemilik" not in combined.columns:
        print("Kolom Nama Pemilik tidak ditemukan.")
        return

    owners = combined[combined["Nama Pemilik"].notna()]["Nama Pemilik"].unique()
    owners = sorted(owners, key=lambda x: str(x).lower())

    print(f"\n{'No.':<4} {'Nama Pemilik':<30} {'GoFood':<8} {'Grab':<8} {'Shopee':<8} {'Total':<8}")
    print("─" * 68)
    for i, o in enumerate(owners, 1):
        o_df = combined[combined["Nama Pemilik"] == o]
        go_n = len(o_df[o_df["Aplikator"] == "GoFood"])
        gr_n = len(o_df[o_df["Aplikator"] == "GrabFood"])
        sh_n = len(o_df[o_df["Aplikator"] == "ShopeeFood"])
        print(f"{i:<4} {str(o):<30} {go_n:<8} {gr_n:<8} {sh_n:<8} {len(o_df):<8}")
    print("─" * 68)
    print(f"Total: {len(owners)} owner terdaftar.\n")


def main():
    parser = argparse.ArgumentParser(
        description="Gabungkan data GoFood & Grab per-owner ke format YYYY-MM-DD HH_MM [Nama Pemilik].xlsx dengan 2 tab identik dan upload ke Google Drive."
    )
    parser.add_argument("--owner", type=str, default=None, help="Filter nama owner tertentu")
    parser.add_argument("--upload", action="store_true", help="Unggah file yang dibuat ke Google Drive")
    parser.add_argument("--local-only", action="store_true", help="Hanya buat file Excel lokal di output_owners/ tanpa upload")
    parser.add_argument("--list-owners", action="store_true", help="Tampilkan daftar semua owner dan jumlah outlet")

    args = parser.parse_args()

    if args.list_owners:
        list_owners()
        return

    do_upload = args.upload and not args.local_only
    process_and_combine(owner_filter=args.owner, upload=do_upload)


if __name__ == "__main__":
    main()

