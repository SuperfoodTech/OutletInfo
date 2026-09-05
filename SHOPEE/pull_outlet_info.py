"""
Shopee Outlet Info Puller (Incremental / Resume Mode)
=====================================================
Menarik data outlet/store dari Shopee Partner API (stores + store detail)
untuk merchant yang belum ditarik / gagal (EXCLUDE: SuperFood, WonderFood, LOKARASA, Gurame Bakar/Do Eat).

Jika sudah ada file Excel hasil sebelumnya, script secara otomatis memuat data existing
dan HANYA melakukan switch browser ke merchant-merchant yang belum ditarik (berdasarkan Group ID / merchantId).

Output: Shopee_{timestamp}.xlsx (tab Listing).

Usage:
    python pull_outlet_info.py
    python pull_outlet_info.py --merchant-id 22380776
"""

import json
import os
import sys
import time
import math
import argparse
from datetime import datetime
from pathlib import Path

# Auto-detect and switch to .venv python if not active
SCRIPT_DIR = Path(__file__).resolve().parent
for venv_candidate in [SCRIPT_DIR / ".venv" / "bin" / "python", SCRIPT_DIR.parent / ".venv" / "bin" / "python", SCRIPT_DIR.parent.parent / ".venv" / "bin" / "python"]:
    if venv_candidate.exists() and sys.executable != str(venv_candidate):
        os.execv(str(venv_candidate), [str(venv_candidate)] + sys.argv)
        break

import requests
import pandas as pd
from openpyxl.styles import Font, Alignment

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ──────────────────────────────────────────────────────────────
# Path Setup
# ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
AUTOMATION_DIR = SCRIPT_DIR / "src" / "shopee-omzet-automation"
CHROME_PROFILE_DIR = SCRIPT_DIR / "data" / "chrome_profile"

# Add automation dir to sys.path for browser module import
if str(AUTOMATION_DIR) not in sys.path:
    sys.path.insert(0, str(AUTOMATION_DIR))

from core import browser

# FORCE the profile directory (same pattern as menu_core/shopee.py)
orig_add_argument = Options.add_argument
def custom_add_argument(self, argument):
    if "--user-data-dir=" in argument:
        argument = f"--user-data-dir={CHROME_PROFILE_DIR}"
        print(f"🔧 [PATCH] Mengalihkan user data dir ke: {argument}")
    orig_add_argument(self, argument)
Options.add_argument = custom_add_argument

# ──────────────────────────────────────────────────────────────
# Constants & Filtering Rules
# ──────────────────────────────────────────────────────────────
SELLER_BASE = "https://foody.shopee.co.id"
TEMPLATE_FILE = SCRIPT_DIR / "YYYY-MM-DD HH_MM Nama Pemilik.xlsx"
SESSION_FILE = SCRIPT_DIR / "data" / "session.json"
CREDS_FILE = SCRIPT_DIR / "credentials.json"
BANK_ACC_FILE = SCRIPT_DIR / "bank_acc.json"
OUTPUT_DIR = SCRIPT_DIR / "data"

# Default credentials fallback
DEFAULT_USERNAME = "allvbadmin"
DEFAULT_PASSWORD = "Master!00!"

# Merchant keywords to EXCLUDE (jangan switch dan lewati merchant ini)
EXCLUDE_KEYWORDS = [
    "nasi padang laura minang",
]

STATUS_MAP = {
    2: "Active",
    1: "Pending",
    0: "Suspended",
    3: "Closed",
}

PORTAL_MAP = {
    "SuperFood": "F",
    "WonderFood": "W",
    "LOKARASA": "L",
    "DoEat": "D",
}


def is_excluded(name: str) -> bool:
    """Check if a merchant or store name should be excluded."""
    if not name:
        return False
    name_lower = name.lower()
    return any(kw in name_lower for kw in EXCLUDE_KEYWORDS)


MERCHANT_INFO_MAP = {}

def get_merchants_to_switch() -> list[dict]:
    """
    Load available merchants from data/merchant_list.json (or response.json / fallback list),
    excluding merchants matching EXCLUDE_KEYWORDS.
    Returns list of dicts with:
      - merchant_id (int / str)
      - merchant_name (str)
      - label (str)
      - occurrence_index (int: 0 for 1st, 1 for 2nd, etc.)
      - total_occurrences (int)
      - is_active (bool)
      - staff_tob_uid (int)
    """
    json_paths = [
        SCRIPT_DIR / "data" / "merchant_list.json",
        AUTOMATION_DIR / "API" / "response.json",
    ]

    merchants = []
    data = None

    for path in json_paths:
        if path.exists():
            try:
                data = json.loads(path.read_text())
                print(f"  ✓ Membaca merchant list dari: {path.name}")
                break
            except Exception as e:
                print(f"  [!] Gagal membaca {path.name}: {e}")

    if data:
        m_list = data.get("data", {}).get("selectMerchant", {}).get("merchantList", [])
        
        # Hitung kemunculan nama merchant untuk menandai duplikat
        name_counts = {}
        for item in m_list:
            name = item.get("merchantName", "").strip()
            if name:
                name_counts[name] = name_counts.get(name, 0) + 1

        name_seen_idx = {}
        for item in m_list:
            m_id = item.get("merchantId")
            m_name = item.get("merchantName", "").strip()
            if not m_name:
                continue

            # Simpan metadata merchant dengan ID sebagai kunci unik
            MERCHANT_INFO_MAP[str(m_id)] = item

            occ_idx = name_seen_idx.get(m_name, 0)
            name_seen_idx[m_name] = occ_idx + 1

            if is_excluded(m_name) or item.get("isActive") is False or item.get("staffStatus") == 2:
                print(f"  [EXCLUDE/INACTIVE] Skipping merchant: '{m_name}' (ID: {m_id})")
                continue

            merchants.append({
                "merchant_id": m_id,
                "merchant_name": m_name,
                "label": m_name,
                "occurrence_index": occ_idx,
                "total_occurrences": name_counts.get(m_name, 1),
                "staff_tob_uid": item.get("staffTobUid"),
                "is_active": item.get("isActive", True),
            })

    if not merchants:
        print("  [!] Menggunakan daftar merchant fallback default...")
        fallback_merchants = [
            "RasaRiang",
            "RasaRiang, Nasi Goreng",
            "Ayam Geprek Suroboyo Amp",
            "HOLANS MRTBK TRANG BULAN",
            "Nasi Goreng Ori",
            "Penyetan Baru Rasa",
            "Bebek dan ayam sambal cumi bumbu kuning",
            "Depotnasipenyetanlonta_",
            "MARTABAK HUWENAKK",
            "Depot Baru Rasa Tandes",
            "Penyetan mbak dina",
            "Salero Minang Raya_",
            "Ayam geprek Putat gede",
            "Warung Bakaran amanah_",
            "Eat Space",
            "Depot Baru Rasa_",
            "Ayam Ori_",
        ]
        for idx, m_name in enumerate(fallback_merchants):
            if not is_excluded(m_name):
                merchants.append({
                    "merchant_id": None,
                    "merchant_name": m_name,
                    "label": m_name,
                    "occurrence_index": 0,
                    "total_occurrences": 1,
                    "staff_tob_uid": None,
                    "is_active": True,
                })

    return merchants


def load_existing_results():
    """Load existing results from the latest Shopee_*.xlsx file if available."""
    excel_files = sorted(OUTPUT_DIR.glob("Shopee_*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not excel_files:
        return [], set(), set()

    latest_file = excel_files[0]
    try:
        try:
            df_exist = pd.read_excel(latest_file, sheet_name="Listing")
        except Exception:
            df_exist = pd.read_excel(latest_file)

        existing_group_ids = set()
        if "Group ID" in df_exist.columns:
            for gid in df_exist["Group ID"].dropna():
                gid_str = str(gid).strip()
                if gid_str and gid_str != "-" and gid_str != "nan":
                    if gid_str.endswith(".0"):
                        gid_str = gid_str[:-2]
                    existing_group_ids.add(gid_str)

        if "Nama Portal" in df_exist.columns:
            existing_merchants = set(df_exist["Nama Portal"].dropna().unique())
        elif "Merchant" in df_exist.columns:
            existing_merchants = set(df_exist["Merchant"].dropna().unique())
        else:
            existing_merchants = set()

        results = df_exist.to_dict("records")
        print(f"  ✓ Memuat {len(results)} outlet dari {len(existing_group_ids)} Group ID / {len(existing_merchants)} merchant existing ({latest_file.name})")
        return results, existing_group_ids, existing_merchants
    except Exception as e:
        print(f"  [!] Gagal membaca file Excel existing {latest_file.name}: {e}")
        return [], set(), set()


def get_driver_user_info(driver) -> dict | None:
    """Fetch current active user info from internal Shopee API via driver."""
    api_js = """
    var done = arguments[arguments.length - 1];
    let token = document.cookie.split('; ').find(row => row.startsWith('shopee_tob_token='))?.split('=')[1];
    fetch('https://api.partner.shopee.co.id/nb/mss/web-api/PartnerAccountServer/GetUserInfo', {
        method: 'POST',
        headers: {
            'Content-Type': 'application/json',
            'x-merchant-token': token || '',
            'x-merchant-language': 'id',
            'x-merchant-login-from': '12'
        },
        body: '{}',
        credentials: 'include'
    })
    .then(r => r.json())
    .then(j => done(j.data || null))
    .catch(() => done(null));
    """
    try:
        driver.set_script_timeout(10)
        return driver.execute_async_script(api_js)
    except Exception:
        return None


CURRENT_OCCURRENCE_INDEX = 0
CURRENT_TARGET_MERCHANT_ID = None

def enhanced_auto_switch_merchant(driver, target_name, is_retry=False):
    """
    Enhanced automated merchant switch that supports duplicate names via CURRENT_OCCURRENCE_INDEX
    and verifies active ID against CURRENT_TARGET_MERCHANT_ID.
    """
    global CURRENT_OCCURRENCE_INDEX, CURRENT_TARGET_MERCHANT_ID
    target_occ_idx = CURRENT_OCCURRENCE_INDEX
    target_id = CURRENT_TARGET_MERCHANT_ID

    print(f"🔄 [MERCHANT] Switching to: '{target_name}' (Target ID: {target_id or '-'}, Occ: {target_occ_idx})...")
    try:
        # Fast Loader Removal
        driver.execute_script("document.querySelectorAll('.ant-spin, [class*=\"loading\"], .shopee-loading').forEach(el => el.remove());")
        
        # PHASE 1: Handle initial merchant selector page right after login
        current_url = driver.current_url
        if "onboarding" in current_url or "merchant-selector" in current_url:
            print(f"  📍 Detected Selector page. Mencoba memilih target langsung...")
            time.sleep(3)
            js_selector_click = """
                var targetRaw = (arguments[0] || "").toLowerCase().trim();
                var targetClean = targetRaw.replace(/_+$/, '').trim();
                var listItems = document.querySelectorAll('.listItem, .merchant-item, li[class*="item"], div[class*="card"]');
                var exact = null;
                var clean = null;
                var partial = null;
                var fallback = null;
                for (var i = 0; i < listItems.length; i++) {
                    var el = listItems[i];
                    var text = (el.innerText || el.textContent || "").toLowerCase().trim();
                    if (!text) continue;
                    if (!fallback) fallback = el;
                    var textClean = text.replace(/_+$/, '').trim();
                    if (text === targetRaw) {
                        exact = el;
                        break;
                    } else if (!clean && (text === targetClean || textClean === targetClean)) {
                        clean = el;
                    } else if (!partial && (text.includes(targetRaw) || (targetClean && text.includes(targetClean)))) {
                        partial = el;
                    }
                }
                var chosen = exact || clean || partial || fallback;
                if (chosen) {
                    if (typeof chosen.scrollIntoView === 'function') chosen.scrollIntoView({block: 'center'});
                    chosen.click();
                    return true;
                }
                return false;
            """
            for _ in range(5):
                if driver.execute_script(js_selector_click, target_name):
                    time.sleep(4)
                    break
                time.sleep(1)

        # PHASE 2: Open profile dropdown and switch
        for switch_attempt in range(3):
            # Ensure we are in dashboard
            if "/food/dashboard" not in driver.current_url:
                driver.get("https://partner.shopee.co.id/food/dashboard")
                time.sleep(3)

            # 1. Hover profile menu
            profile_menu = None
            for sel in [".merchantName", ".user-info", "li.ant-menu-item:last-child", "div[class*=\"header\"] .ant-dropdown-trigger"]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        profile_menu = el
                        break
                except Exception:
                    continue

            if not profile_menu:
                try:
                    profile_menu = WebDriverWait(driver, 10).until(
                        lambda d: d.find_element(By.CSS_SELECTOR, ".merchantName, .user-info, li.ant-menu-item:last-child")
                    )
                except Exception:
                    pass

            if not profile_menu:
                print(f"  [!] Profile menu not found (Attempt {switch_attempt+1}/3)")
                time.sleep(2)
                continue

            try:
                actions = ActionChains(driver)
                actions.move_to_element(profile_menu).click().perform()
            except Exception:
                pass
            time.sleep(1)

            # 2. Click 'Pilih Merchant Lain' / 'Switch Merchant'
            dropdown_opened = False
            try:
                switch_trigger = WebDriverWait(driver, 5).until(
                    lambda d: d.find_element(By.XPATH, "//span[contains(text(), 'Pilih Merchant') or contains(text(), 'Switch Merchant') or contains(text(), 'Ganti Merchant')] | //li[contains(., 'Pilih Merchant')]")
                )
                actions = ActionChains(driver)
                actions.move_to_element(switch_trigger).click().perform()
                dropdown_opened = True
                time.sleep(1.5)
            except Exception:
                # Fallback JS click
                js_found = driver.execute_script("""
                    var spans = document.querySelectorAll('span, p, div, li, a');
                    for (var s of spans) {
                        var text = (s.innerText || '').trim();
                        if (text.includes('Pilih Merchant Lain') || text.includes('Switch Merchant') || text.includes('Ganti Merchant')) {
                            s.click();
                            return true;
                        }
                    }
                    return false;
                """)
                if js_found:
                    dropdown_opened = True
                    time.sleep(1.5)

            if not dropdown_opened:
                print(f"  [!] Dropdown sub-menu tidak terbuka (Attempt {switch_attempt+1}/3)")
                time.sleep(2)
                continue

            # 3. Find target in submenu items (STRICT Exact or Clean match only - NO partial substring leakage)
            js_switch_script = """
                var targetRaw = (arguments[0] || "").toLowerCase().trim();
                var targetClean = targetRaw.replace(/_+$/, '').trim();
                var targetOccIdx = arguments[1] || 0;
                var items = document.querySelectorAll('li.ant-menu-item, li[role="menuitem"], .ant-dropdown-menu-item, [class*="menu-item"]');
                
                var exactMatches = [];
                var cleanMatches = [];

                for (var i = 0; i < items.length; i++) {
                    var el = items[i];
                    if (!el || typeof el.click !== 'function') continue;
                    var text = (el.innerText || el.textContent || "").toLowerCase().trim();
                    if (!text) continue;
                    var textClean = text.replace(/_+$/, '').trim();

                    // Strict matching only
                    if (text === targetRaw) {
                        exactMatches.push(el);
                    } else if (targetClean && (text === targetClean || textClean === targetClean)) {
                        cleanMatches.push(el);
                    }
                }

                var matched = exactMatches.length > 0 ? exactMatches : cleanMatches;
                if (matched.length > 0) {
                    var chosenIdx = Math.min(targetOccIdx, matched.length - 1);
                    var chosen = matched[chosenIdx];
                    if (typeof chosen.scrollIntoView === 'function') chosen.scrollIntoView({block: 'center'});
                    chosen.click();
                    return { ok: true, matchedCount: matched.length, clickedIdx: chosenIdx, matchType: exactMatches.length > 0 ? 'exact' : 'clean' };
                }
                return { ok: false, matchedCount: 0 };
            """

            found_res = None
            for _ in range(12):
                found_res = driver.execute_script(js_switch_script, target_name, target_occ_idx)
                if found_res and found_res.get("ok"):
                    break
                try:
                    driver.execute_script("document.querySelectorAll('.ant-dropdown-menu, ul[role=\"menu\"], .ant-popover-inner-content, div[class*=\"menu\"]').forEach(el => el.scrollTop += 400);")
                except Exception:
                    pass
                time.sleep(0.8)

            if found_res and found_res.get("ok"):
                print(f"  ✅ Clicked '{target_name}' in submenu ({found_res.get('matchType', 'exact')} match, item {found_res.get('clickedIdx', 0)+1}/{found_res.get('matchedCount', 1)}).")
                time.sleep(4)

                # Check onboarding
                if "onboarding" in driver.current_url.lower():
                    browser._handle_onboarding_invitation(driver)
                    time.sleep(3)

                # Verify via GetUserInfo
                user_info = get_driver_user_info(driver)
                if user_info:
                    active_id = str(user_info.get("merchantId") or "")
                    active_nm = user_info.get("merchantName") or ""
                    print(f"  [✓] Active merchant after switch: '{active_nm}' (ID: {active_id})")
                    if target_id and str(target_id) != active_id:
                        print(f"  ⚠️ Warning: Active ID ({active_id}) does not match target ({target_id}). Retrying switch...")
                        continue
                return True
            else:
                print(f"  ⚠️ Outlet '{target_name}' tidak ditemukan di dropdown (Attempt {switch_attempt+1}/3).")
                time.sleep(2)

        return False
    except Exception as e:
        print(f"  [!] Exception in auto_switch_merchant: {e}")
        return False

# Patch auto_switch_merchant in core.browser
browser.auto_switch_merchant = enhanced_auto_switch_merchant


# ──────────────────────────────────────────────────────────────
# Authentication
# ──────────────────────────────────────────────────────────────
def get_auth_session(target_name: str, target_merchant_id: str | int = None, occurrence_index: int = 0) -> tuple:
    """
    Launch browser, login as allvbadmin, switch to target merchant by ID/name, extract tokens, close browser.
    Returns (tob_token, entity_id, extra_cookies) or raises on failure.
    """
    global CURRENT_OCCURRENCE_INDEX, CURRENT_TARGET_MERCHANT_ID
    CURRENT_OCCURRENCE_INDEX = occurrence_index
    CURRENT_TARGET_MERCHANT_ID = target_merchant_id

    browser.set_session_file(SESSION_FILE)

    username = DEFAULT_USERNAME
    password = DEFAULT_PASSWORD
    if CREDS_FILE.exists():
        try:
            creds = json.loads(CREDS_FILE.read_text())
            username = creds.get("username") or creds.get("shopee_username") or username
            password = creds.get("password") or creds.get("shopee_password") or password
        except Exception:
            pass

    print(f"[*] Membuka browser (headless=False) dan memilih merchant: '{target_name}' (ID: {target_merchant_id or '-'}, Occ: {occurrence_index})...")
    
    session_data = browser.get_session(
        username=username,
        password=password,
        headless=False,
        close_browser=False,
        target_name=target_name,
        interactive=False,
    )

    if not session_data or "driver" not in session_data:
        raise RuntimeError(f"Gagal menginisialisasi browser atau login untuk merchant '{target_name}'.")

    driver = session_data["driver"]

    try:
        print("[*] Memperbarui token autentikasi...")
        session = browser.refresh_tokens(driver)
        if not session or "shopee_tob_token" not in session:
            raise RuntimeError("Gagal memperbarui token autentikasi.")

        tob_token = session["shopee_tob_token"]
        entity_id = str(session.get("shopee_tob_entity_id", "") or "")
        extra_cookies = session.get("extra_cookies", {})

        # Validasi ketat jika target_merchant_id ditentukan
        if target_merchant_id and entity_id and entity_id != str(target_merchant_id):
            uinfo = get_driver_user_info(driver)
            actual_id = str(uinfo.get("merchantId") or "") if uinfo else entity_id
            if actual_id != str(target_merchant_id):
                raise RuntimeError(
                    f"❌ GAGAL SWITCH: Merchant aktif di browser adalah ID '{actual_id}', "
                    f"bukan target ID '{target_merchant_id}' ({target_name}). Penarikan data dibatalkan demi integritas data."
                )

        print(f"[✓] Token berhasil didapat. Entity ID: {entity_id}")
        return tob_token, entity_id, extra_cookies
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ──────────────────────────────────────────────────────────────
# Shopee Outlet Client
# ──────────────────────────────────────────────────────────────
class ShopeeOutletClient:
    """Client for Shopee Seller API - Outlet/Store info endpoints."""

    def __init__(self, tob_token: str, entity_id: str, extra_cookies: dict = None):
        self.tob_token = tob_token
        self.entity_id = entity_id
        self.extra_cookies = extra_cookies or {}
        self.session = requests.Session()
        self.user_agent = (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36"
        )

    def _seller_headers(self, override_entity_id: str = None) -> dict:
        """Build headers with cookie string, optionally overriding entity_id."""
        eid = override_entity_id or self.entity_id
        cookies = self.extra_cookies.copy()
        cookies["shopee_tob_token"] = self.tob_token
        cookies["shopee_tob_entity_id"] = str(eid)
        cookie_str = "; ".join(f"{k}={v}" for k, v in cookies.items())

        return {
            "Host": "foody.shopee.co.id",
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "User-Agent": self.user_agent,
            "Cookie": cookie_str,
            "Origin": "https://partner.shopee.co.id",
            "Referer": "https://partner.shopee.co.id/",
        }

    def get_list_stores(self, page_no: int = 1, page_size: int = 50) -> dict:
        """POST /api/seller/stores - List all stores with pagination."""
        url = f"{SELLER_BASE}/api/seller/stores"
        payload = {
            "store_name": "",
            "page_size": page_size,
            "page_no": page_no,
        }

        try:
            resp = self.session.post(
                url,
                json=payload,
                headers=self._seller_headers(),
                timeout=30,
            )
            data = resp.json()
            if data.get("code") == 0:
                return data
            print(f"  [!] get_list_stores failed: code={data.get('code')}, msg={data.get('msg')}")
        except Exception as e:
            print(f"  [!] get_list_stores error: {e}")
        return {"data": {"total": 0, "stores": []}}

    def get_all_stores(self) -> list[dict]:
        """Fetch ALL stores across all pages."""
        all_stores = []
        page_no = 1
        page_size = 50

        # First call to get total
        first_resp = self.get_list_stores(page_no=1, page_size=page_size)
        total = first_resp.get("data", {}).get("total", 0)
        stores = first_resp.get("data", {}).get("stores", [])
        all_stores.extend(stores)

        if total == 0:
            print("[!] Tidak ada store yang ditemukan.")
            return []

        total_pages = math.ceil(total / page_size)
        print(f"[*] Total stores: {total} | Total pages: {total_pages}")
        print(f"  ✓ Page 1/{total_pages}: {len(stores)} stores")

        # Fetch remaining pages
        for page_no in range(2, total_pages + 1):
            time.sleep(0.3)  # Rate limit
            resp = self.get_list_stores(page_no=page_no, page_size=page_size)
            stores = resp.get("data", {}).get("stores", [])
            all_stores.extend(stores)
            print(f"  ✓ Page {page_no}/{total_pages}: {len(stores)} stores")

        print(f"[✓] Total {len(all_stores)} stores fetched.")
        return all_stores

    def get_store_detail(self, store_id: str) -> dict | None:
        """GET /api/seller/store - Get detail for a specific store."""
        url = f"{SELLER_BASE}/api/seller/store"

        try:
            resp = self.session.get(
                url,
                headers=self._seller_headers(override_entity_id=str(store_id)),
                timeout=15,
            )
            data = resp.json()
            if data.get("code") == 0:
                return data.get("data", {}).get("store", {})
            print(f"  [!] get_store_detail({store_id}) failed: code={data.get('code')}, msg={data.get('msg')}")
        except Exception as e:
            print(f"  [!] get_store_detail({store_id}) error: {e}")
        return None


# ──────────────────────────────────────────────────────────────
# Detail Fetching Helper
# ──────────────────────────────────────────────────────────────
def fetch_store_details(
    client: ShopeeOutletClient,
    filtered_stores: list[dict],
) -> tuple[list[dict], int]:
    """Fetch detail for each store in list and format records matching G-N template."""
    results = []
    failed_count = 0

    for idx, fs in enumerate(filtered_stores, 1):
        store = fs["store"]
        merchant_label = fs["merchant"]
        fallback_group_id = fs.get("group_id", "")
        store_id = store.get("id", "")
        store_name = store.get("name", "")
        store_status = store.get("status", 0)

        if idx % 20 == 0 or idx == 1:
            print(f"  [{idx}/{len(filtered_stores)}] Processing: {store_name} ...")

        # Get detail for address
        time.sleep(0.5)  # Rate limit
        detail = client.get_store_detail(str(store_id)) if store_id else None

        if detail:
            location = detail.get("location", {})
            address_parts = []
            addr = location.get("address", "")
            district = location.get("district", "")
            city = location.get("city", "")
            state = location.get("state", "")

            if addr:
                address_parts.append(addr)
            if district:
                address_parts.append(district)
            if city:
                address_parts.append(city)
            if state:
                address_parts.append(state)

            full_address = ", ".join(address_parts)
            group_id = str(detail.get("merchant_id", "") or fallback_group_id or "")
            status_val = detail.get("status", store_status)
        else:
            full_address = ""
            group_id = str(fallback_group_id or "")
            status_val = store_status
            if store_id:
                failed_count += 1

        store_link = f"https://shopee.co.id/now-food/shop/{store_id}" if store_id and str(store_id) != "-" else "-"
        status_listing = STATUS_MAP.get(status_val, str(status_val))

        results.append({
            "Aplikator": "Shopeefood",
            "Nama Portal": merchant_label,
            "Group ID": group_id,
            "Nama Listing": store_name,
            "Link": store_link,
            "Store ID": store_id,
            "Status Listing": status_listing,
            "Alamat": full_address,
        })

    return results, failed_count


def run_pull(
    target_merchants: list[dict] = None,
    target_merchant_id: str = None,
    target_merchant_name: str = None,
    no_resume: bool = False,
    include_excluded: bool = False,
    output_path: Path | str = None,
) -> Path | None:
    """
    Main executor for pulling outlet information.
    Can be called directly from CLI or other modules.
    """
    print("=" * 70)
    print("  SHOPEE OUTLET INFO PULLER")
    if not include_excluded:
        print("  Dilewati (EXCLUDE): SuperFood, WonderFood, LOKARASA, Gurame Bakar, Do Eat")
    print("=" * 70)
    print()

    is_filtered_run = bool(target_merchants is not None or target_merchant_id is not None or target_merchant_name is not None)

    # 1. Load existing results from Excel (only for full batch run without explicit filters)
    if is_filtered_run or no_resume:
        if is_filtered_run:
            print("[*] Mode Filter: Hanya memproses dan mengekspor merchant yang dipilih.")
        else:
            print("[*] Mode --no-resume / fresh: Mengabaikan data existing di Excel.")
        all_results, existing_group_ids, existing_merchants = [], set(), set()
    else:
        all_results, existing_group_ids, existing_merchants = load_existing_results()

    seen_store_ids = set()
    for r in all_results:
        sid = r.get("Store ID")
        if sid and str(sid) != "-" and str(sid) != "nan":
            seen_store_ids.add(str(sid))

    # 2. Determine target merchants
    if target_merchants is not None:
        all_candidates = target_merchants
    else:
        all_candidates = get_merchants_to_switch()

    # Filter targets
    merchants_to_process = []
    for m in all_candidates:
        m_id = m["merchant_id"]
        m_name = m["merchant_name"]

        # Explicit filter
        if target_merchant_id and str(m_id) != str(target_merchant_id):
            continue
        if target_merchant_name and m_name.lower() != target_merchant_name.lower():
            continue

        if not include_excluded and is_excluded(m_name):
            continue

        # Incremental filter (only if no explicit target was supplied and resume enabled)
        if not is_filtered_run and not no_resume:
            if m_id and str(m_id) in existing_group_ids:
                continue
            if m["total_occurrences"] == 1 and m_name in existing_merchants:
                continue

        merchants_to_process.append(m)

    print(f"\n[*] Total {len(merchants_to_process)} merchant yang akan diproses:")
    for idx, m in enumerate(merchants_to_process, 1):
        print(f"    {idx}. {m['merchant_name']} (ID: {m['merchant_id']}, Occ: {m['occurrence_index']})")

    if not merchants_to_process:
        print("\n[✓] Tidak ada merchant yang perlu ditarik (semua sudah ditarik atau tidak cocok filter).")
        return None

    merchant_counts = {}

    # 3. Loop through target merchants
    for switch_idx, m in enumerate(merchants_to_process, 1):
        merchant_name = m["merchant_name"]
        merchant_id = m["merchant_id"]
        occ_idx = m["occurrence_index"]
        fallback_label = m["label"]

        if not include_excluded and is_excluded(merchant_name):
            print(f"\n[EXCLUDE] Melewati switch ke merchant: '{merchant_name}'")
            continue

        print()
        print(f"{'='*70}")
        print(f"  MERCHANT {switch_idx}/{len(merchants_to_process)}: {merchant_name} (ID: {merchant_id})")
        print(f"{'='*70}")

        # Authenticate & switch to this merchant
        print(f"\n[2/5] Authenticating & switching to '{merchant_name}' (ID: {merchant_id})...")
        try:
            tob_token, entity_id, extra_cookies = get_auth_session(
                target_name=merchant_name,
                target_merchant_id=merchant_id,
                occurrence_index=occ_idx,
            )
        except Exception as e:
            print(f"  [!] Gagal auth untuk merchant '{merchant_name}': {e}")
            continue

        # Fetch all stores for this merchant
        print(f"\n[3/5] Fetching all stores for '{merchant_name}' (paginated)...")
        client = ShopeeOutletClient(
            tob_token=tob_token,
            entity_id=entity_id,
            extra_cookies=extra_cookies,
        )
        all_stores = client.get_all_stores()

        if not all_stores:
            print(f"  [!] Tidak ada store ditemukan untuk '{merchant_name}'.")
            continue

        # Filter stores
        print(f"\n[4/5] Collecting stores for '{merchant_name}'...")
        filtered_stores = []
        for store in all_stores:
            store_id = str(store.get("id", ""))
            store_name = store.get("name", "")

            if store_id in seen_store_ids:
                continue  # Skip duplicates

            if not include_excluded and (is_excluded(store_name) or is_excluded(fallback_label)):
                print(f"  [EXCLUDE] Skipping store: {store_name}")
                continue

            filtered_stores.append({
                "store": store,
                "merchant": fallback_label,
                "group_id": str(merchant_id) if merchant_id else "",
            })
            seen_store_ids.add(store_id)

        print(f"  ✓ {len(filtered_stores)} stores from '{merchant_name}' processed")

        for fs in filtered_stores:
            m_lbl = fs["merchant"]
            merchant_counts[m_lbl] = merchant_counts.get(m_lbl, 0) + 1

        # Fetch detail for each filtered store
        if filtered_stores:
            print(f"\n  Fetching detail for {len(filtered_stores)} stores...")
            results, failed = fetch_store_details(client, filtered_stores)
            all_results.extend(results)
            print(f"  ✓ Detail fetched: {len(results)} OK, {failed} failed")

    # 4. Build & Export formatted Excel
    print()
    print("[5/5] Exporting combined data into template format (Columns G-N)...")

    # Insert Inactive record for NASI PADANG LAURA MINANG_ only on full runs
    if not is_filtered_run:
        if not any(r.get("Nama Portal") == "NASI PADANG LAURA MINANG_" or r.get("Merchant") == "NASI PADANG LAURA MINANG_" for r in all_results):
            all_results.append({
                "Aplikator": "Shopeefood",
                "Nama Portal": "NASI PADANG LAURA MINANG_",
                "Group ID": "1506647",
                "Nama Listing": "-",
                "Link": "-",
                "Store ID": "-",
                "Status Listing": "Inactive",
                "Alamat": "-",
            })

    # Standardize records for export (Columns G-N) and deduplicate
    standardized_results = []
    seen_entry_keys = set()

    for r in all_results:
        merchant_name = r.get("Nama Portal") or r.get("Merchant") or ""
        store_name = r.get("Nama Listing") or r.get("Nama") or ""
        store_id = str(r.get("Store ID") or "")
        group_id = str(r.get("Group ID") or "")
        if group_id.endswith(".0"):
            group_id = group_id[:-2]
        if store_id.endswith(".0"):
            store_id = store_id[:-2]

        status_val = r.get("Status Listing") or r.get("Status") or ""
        if status_val == "Suspended":
            status_val = "Active"

        link_val = r.get("Link")
        if not link_val or str(link_val) == "nan":
            link_val = f"https://shopee.co.id/now-food/shop/{store_id}" if store_id and str(store_id) != "-" else "-"

        # Deduplicate
        entry_key = (merchant_name, group_id, store_id) if store_id != "-" else (merchant_name, group_id, store_name)
        if entry_key in seen_entry_keys:
            continue
        seen_entry_keys.add(entry_key)

        standardized_results.append({
            "Aplikator": r.get("Aplikator") or "Shopeefood",
            "Nama Portal": merchant_name,
            "Group ID": group_id,
            "Nama Listing": store_name,
            "Link": link_val,
            "Store ID": store_id,
            "Status Listing": status_val,
            "Alamat": r.get("Alamat") or "",
        })

    # Sort rows by Nama Portal and Nama Listing
    standardized_results.sort(key=lambda x: (str(x.get("Nama Portal", "")), str(x.get("Nama Listing", ""))))

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Output filename
    if output_path:
        excel_path = Path(output_path)
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        excel_path = OUTPUT_DIR / f"Shopee_{timestamp}.xlsx"

    # Use template workbook if available, otherwise build fresh
    import openpyxl
    if TEMPLATE_FILE.exists():
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        if "Listing" in wb.sheetnames:
            ws = wb["Listing"]
        else:
            ws = wb.active
        # Clear existing data rows starting from row 2
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listing"
        default_headers = [
            "Nama Pemilik", "Nama Brand", "Model", "Tipe", "Outlet", "Nomor HP",
            "Aplikator", "Nama Portal", "Group ID", "Nama Listing", "Link", "Store ID", "Status Listing", "Alamat",
            "Nama Bank", "Nama Pemilik Rekening", "Nomor Rekening",
            "Nama Akses", "Email FoodMaster1", "Email FoodMaster2", "Nama Pengguna", "Kata Sandi", "Nama Portal",
            "S Nomor HP Akses Pemilik", "S Username Akses Pemilik", "S Kata Sandi Akses Pemilik",
            "S Allvbadmin Username Akses Staff", "S Allvbadmin Kata Sandi Akses Staff",
            "S Bot Username Akses Staff", "S Bot Kata Sandi Akses Staff",
            "S BD Username Akses Staff", "S BD Kata Sandi Akses Staff",
            "BD", "Status Internal", "Tanggal Live", "Tanggal Churn", "Tarif"
        ]
        ws.append(default_headers)

    font_header = Font(name="Arial", size=10, bold=True)
    font_body = Font(name="Arial", size=10, bold=False)
    align_left = Alignment(horizontal="left", vertical="center")

    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_header
        cell.alignment = align_left

    # Populate rows with mapping to G-N ONLY (Columns 7 to 14)
    for row_idx, r in enumerate(standardized_results, start=2):
        ws.cell(row=row_idx, column=7, value=r.get("Aplikator", "Shopeefood"))            # G: Aplikator
        ws.cell(row=row_idx, column=8, value=r.get("Nama Portal", ""))                   # H: Nama Portal
        ws.cell(row=row_idx, column=9, value=r.get("Group ID", ""))                      # I: Group ID
        ws.cell(row=row_idx, column=10, value=r.get("Nama Listing", ""))                 # J: Nama Listing
        ws.cell(row=row_idx, column=11, value=r.get("Link", ""))                         # K: Link
        ws.cell(row=row_idx, column=12, value=r.get("Store ID", ""))                     # L: Store ID
        ws.cell(row=row_idx, column=13, value=r.get("Status Listing", ""))               # M: Status Listing
        ws.cell(row=row_idx, column=14, value=r.get("Alamat", ""))                       # N: Alamat

        # Format row styling
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = font_body
            cell.alignment = align_left

    wb.save(excel_path)
    print(f"  ✓ Sheet 'Listing': {len(standardized_results)} rows")
    print(f"  ✓ Output: {excel_path}")

    # Final Summary
    print()
    print("=" * 70)
    print("  SELESAI!")
    print(f"  Total outlet akhir: {len(standardized_results)}")
    print(f"  Output: {excel_path}")
    print("=" * 70)
    return excel_path


def main():
    parser = argparse.ArgumentParser(description="Shopee Outlet Info Puller")
    parser.add_argument("--merchant-id", type=str, default=None, help="Tarik hanya merchant dengan ID ini")
    parser.add_argument("--merchant-name", type=str, default=None, help="Tarik hanya merchant dengan nama ini")
    parser.add_argument("--no-resume", action="store_true", help="Abaikan data existing dan tarik ulang dari awal")
    parser.add_argument("--include-excluded", action="store_true", help="Sertakan merchant yang masuk daftar blacklist/exclude")
    parser.add_argument("--output", "-o", type=str, default=None, help="Lokasi/nama file output excel")
    args = parser.parse_args()

    run_pull(
        target_merchant_id=args.merchant_id,
        target_merchant_name=args.merchant_name,
        no_resume=args.no_resume,
        include_excluded=args.include_excluded,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
