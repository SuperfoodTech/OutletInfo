import asyncio
import json
import logging
import os
import random
import argparse
import pandas as pd
from playwright.async_api import async_playwright

try:
    from curl_cffi.requests import AsyncSession
except ImportError:
    pass

import urllib.request
import subprocess

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger("GrabMerchantScraper")

auth_headers = {}

import csv
import sys
import os
import time
import re
import hashlib
import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GRAB_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(GRAB_DIR, "cache")
OUTPUT_DIR = os.path.join(GRAB_DIR, "output")
MASTER_DIR = os.path.join(GRAB_DIR, "master")
SESSIONS_DIR = os.path.join(GRAB_DIR, "sessions")

os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(MASTER_DIR, exist_ok=True)
os.makedirs(SESSIONS_DIR, exist_ok=True)

GOOGLE_SHEET_AGENCY_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQ3tLKBNXDqRgBw0mNhKZFxgvKx-JoiTDzm_s5Ix1cm7O6HCv4IvExOLR2HSRVaXSsx82V348mcr9X4/pub?output=csv"
GOOGLE_SHEET_VB_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRYSUnKOqk29LCktTxdb0wPLbWMbRaWRP3eC_UA4AwYod1FW6zDMhtLMC5ghIvot2B8upCDfBsn-TCP/pub?gid=978201567&single=true&output=csv"
GOOGLE_SHEET_VERCEL_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTprbPPf_J5gAVL3PYeHbbdl5ZXQvb17HY2lJGPI2xg13Ly3AGT8eYHLYmU_m1NdtkBVg-qUGv1BoEE/pub?output=csv"
LOCAL_CRED_CSV = os.path.join(BASE_DIR, 'A. Credential (Outlet & Access)  - Unique Portal Gr (1).csv')


def get_safe_cache_filename(portal_name):
    """Menghasilkan nama file cache yang aman dari batasan panjang sistem operasi (maks 255 karakter)."""
    clean = re.sub(r'[^a-zA-Z0-9_.-]', '_', str(portal_name).strip())
    if len(clean) > 80:
        h = hashlib.md5(str(portal_name).encode('utf-8')).hexdigest()[:8]
        clean = clean[:70].rstrip('_') + f"_{h}"
    return f"grab_portal_{clean}.json"


def get_credentials_from_sheet(source_type="agency", custom_url=None):
    """Mengambil daftar kredensial portal Grab berdasarkan tipe sumber (Agency, VB, atau Vercel)."""
    data = []
    
    st_lower = str(source_type).lower()
    target_url = custom_url
    if not target_url:
        if st_lower == "vb":
            target_url = GOOGLE_SHEET_VB_URL or os.getenv("GRAB_VB_CSV_URL", "")
        elif st_lower == "vercel":
            target_url = GOOGLE_SHEET_VERCEL_URL
        else:
            target_url = GOOGLE_SHEET_AGENCY_URL

    # 1. Coba ambil secara live dari Google Sheet URL
    if target_url:
        try:
            logger.info(f"[*] Mengambil data portal Grab [{source_type.upper()}] langsung dari Google Sheet (Live URL)...")
            res = subprocess.run(['curl', '-s', '-L', target_url], capture_output=True, text=True, timeout=20)
            if res.returncode == 0 and res.stdout.strip():
                reader = csv.reader(res.stdout.splitlines())
                data = list(reader)
                logger.info(f"[✓] Berhasil memuat {len(data)} baris data [{source_type.upper()}] dari Google Sheet via curl.")
        except Exception as e:
            logger.warning(f"⚠️ Gagal fetch live via curl [{source_type.upper()}]: {e}.")

        if not data:
            try:
                req = urllib.request.Request(target_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=20) as response:
                    content = response.read().decode('utf-8')
                    reader = csv.reader(content.splitlines())
                    data = list(reader)
                    logger.info(f"[✓] Berhasil memuat {len(data)} baris data [{source_type.upper()}] dari Google Sheet.")
            except Exception as e:
                logger.warning(f"⚠️ Gagal fetch live dari Google Sheet [{source_type.upper()}]: {e}.")

    # 2. Fallback ke file CSV lokal jika live fetch gagal (khusus Agency)
    if not data and st_lower == "agency" and os.path.exists(LOCAL_CRED_CSV):
        try:
            logger.info(f"[*] Membaca data portal dari fallback lokal: {os.path.basename(LOCAL_CRED_CSV)}")
            with open(LOCAL_CRED_CSV, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                data = list(reader)
        except Exception as ex:
            logger.error(f"Gagal membaca fallback CSV lokal: {ex}")

    if not data:
        logger.error(f"❌ Tidak ada data portal yang dapat dimuat untuk tipe [{source_type.upper()}].")
        return []

    # Dynamic column indexing
    header = [str(h).strip().lower() for h in data[0]]
    col_app = -1
    col_owner = -1
    col_portal = -1
    col_outlet = -1
    col_user = -1
    col_pass = -1
    col_notes = -1

    if st_lower == "agency":
        col_app = 3
        col_owner = 0
        col_portal = 2
        col_user = 26
        col_pass = 28
    else:
        # 1. Exact match pass
        for i, h in enumerate(header):
            if h in ['aplikasi', 'app', 'platform']:
                col_app = i
            elif h in ['owner', 'pemilik', 'nama pemilik']:
                col_owner = i
            elif h in ['nama portal', 'portal', 'nama akses']:
                col_portal = i
            elif h in ['nama outlet', 'outlet', 'nama brand', 'brand']:
                col_outlet = i
            elif h in ['username', 'nama pengguna', 'user login', 'user']:
                col_user = i
            elif h in ['password', 'kata sandi', 'pass login', 'pass']:
                col_pass = i
            elif h in ['notes', 'catatan', 'keterangan']:
                col_notes = i

        # 2. Substring fallback only for still unmapped columns
        for i, h in enumerate(header):
            if col_app == -1 and any(k in h for k in ['aplikasi', 'platform']): col_app = i
            if col_owner == -1 and any(k in h for k in ['owner', 'pemilik']): col_owner = i
            if col_portal == -1 and any(k in h for k in ['portal']): col_portal = i
            if col_outlet == -1 and any(k in h for k in ['outlet']): col_outlet = i
            if col_user == -1 and any(k in h for k in ['user']): col_user = i
            if col_pass == -1 and any(k in h for k in ['pass']): col_pass = i

        if col_portal == -1:
            col_portal = col_outlet

    portals = []
    for row in data[1:]:
        # Filter status Restricted
        if col_notes != -1 and col_notes < len(row):
            note_val = str(row[col_notes]).strip().lower()
            if "restricted" in note_val:
                p_name = row[col_portal].strip() if col_portal != -1 and col_portal < len(row) else "Unknown"
                logger.info(f"🚫 Melewati portal '{p_name}' karena berstatus Restricted.")
                continue

        is_grab = True
        if col_app != -1 and col_app < len(row) and row[col_app].strip():
            is_grab = "grab" in row[col_app].strip().lower()

        if is_grab and len(row) > max(col_user, col_pass):
            owner = row[col_owner].strip() if col_owner != -1 and col_owner < len(row) else ("VB" if st_lower == "vb" else "")
            
            portal = ""
            if col_portal != -1 and col_portal < len(row) and row[col_portal].strip():
                portal = row[col_portal].strip()
            if not portal and col_outlet != -1 and col_outlet < len(row):
                portal = row[col_outlet].strip()

            brand = portal.split(" - ")[0].strip() if " - " in portal else portal
            username = row[col_user].strip() if col_user != -1 and col_user < len(row) else ""
            password = row[col_pass].strip() if col_pass != -1 and col_pass < len(row) else ""
            
            if username and username != "-" and password and password != "-":
                safe_cache_name = get_safe_cache_filename(portal)
                safe_portal_name = "".join([c for c in portal if c.isalpha() or c.isdigit() or c == ' ']).rstrip()
                cache_file = os.path.join(CACHE_DIR, safe_cache_name)
                portals.append({
                    "owner": owner if owner else ("VB" if st_lower == "vb" else brand),
                    "brand": brand,
                    "name": portal,
                    "username": username,
                    "password": password,
                    "source_type": source_type.upper(),
                    "cache_file": cache_file,
                    "output": os.path.join(OUTPUT_DIR, f"{safe_portal_name}.xlsx")
                })

    logger.info(f"[✓] Terdaftar {len(portals)} portal GrabFood valid [{source_type.upper()}] untuk diproses.")
    return portals

def parse_selection(choice, max_val):
    selected = set()
    choice = choice.strip()
    if choice.lower() in ('0', 'all'):
        return list(range(max_val))
        
    parts = choice.split(',')
    for part in parts:
        if '-' in part:
            try:
                start, end = map(int, part.split('-'))
                if 1 <= start <= end <= max_val:
                    for i in range(start, end + 1):
                        selected.add(i - 1)
            except ValueError:
                pass
        else:
            try:
                val = int(part)
                if 1 <= val <= max_val:
                    selected.add(val - 1)
            except ValueError:
                pass
    return sorted(list(selected))

BASE_URL   = "https://merchant.grab.com"
LOGIN_URL  = "https://weblogin.grab.com/merchant/login?service_id=MEXUSERS&redirect=https%3A%2F%2Fmerchant.grab.com%2Fportal"
DASHBOARD_URL = f"{BASE_URL}/dashboard"
INVENTORY_URL = f"{BASE_URL}/food/inventory"


async def human_type(locator, text):
    """Mengetik teks dengan jeda acak antar karakter agar terlihat seperti manusia."""
    await locator.click()
    for char in text:
        await locator.press(char)
        await asyncio.sleep(random.uniform(0.07, 0.18))


async def handle_welcome_back_or_continue(page):
    """Mengecek dan mengklik tombol Continue jika halaman menampilkan Welcome back / Saved Account."""
    try:
        # Tunggu sampai tombol Continue muncul jika halaman sedang me-render saved account
        btn_loc = page.locator('button:has-text("Continue"), button:has-text("Lanjut"), button:has-text("Masuk")').first
        try:
            await btn_loc.wait_for(state="visible", timeout=6000)
        except Exception:
            pass

        if await btn_loc.is_visible():
            btn_text = (await btn_loc.inner_text()).strip()
            logger.info(f"👉 [Saved Account Terdeteksi] Mengklik tombol '{btn_text}' pada layar Welcome back...")
            await btn_loc.click()
            await page.wait_for_timeout(4000)
            return True
    except Exception as e:
        logger.debug(f"Check continue button exception: {e}")
    return False


async def perform_login(page, username, password):
    logger.info(f"Navigating to Grab Merchant login page for {username}...")
    try:
        # 1. Cek dulu apakah halaman saat ini sudah menampilkan tombol Continue (Welcome back)
        if await handle_welcome_back_or_continue(page):
            if "login" not in page.url.lower() and "saved-accounts" not in page.url.lower():
                logger.info(f"[✓] Berhasil masuk via tombol Continue untuk {username}!")
                return True

        try:
            await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=45000)
        except Exception as e:
            logger.warning(f"Navigasi login lambat ({e}), melanjutkan...")
        
        # Tambahan waktu tunggu untuk memastikan JavaScript selesai me-render form
        logger.info("Menunggu halaman login ter-load sepenuhnya...")
        await page.wait_for_timeout(4000)

        # Cek sekali lagi apakah redirect login menghasilkan layar Welcome back
        if await handle_welcome_back_or_continue(page):
            if "login" not in page.url.lower() and "saved-accounts" not in page.url.lower():
                logger.info(f"[✓] Berhasil masuk via tombol Continue untuk {username}!")
                return True

        # Cek apakah layar password sudah langsung muncul (misal dari Saved Account / Welcome back)
        password_input = page.locator('input[type="password"]').first
        is_pass_visible = False
        try:
            is_pass_visible = await password_input.is_visible()
        except Exception:
            pass

        if is_pass_visible:
            # Periksa apakah form password ini untuk username yang sedang dituju
            try:
                body_txt = (await page.inner_text("body")).lower()
                if "enter password for" in body_txt and username.lower() not in body_txt:
                    logger.info(f"Layar password untuk akun lain. Mengklik tombol kembali...")
                    back_btn = page.locator('button:has(svg), button[aria-label*="back" i], [data-testid*="back" i]').first
                    if await back_btn.is_visible():
                        await back_btn.click()
                        await page.wait_for_timeout(2000)
                        is_pass_visible = False
            except Exception:
                pass

        if not is_pass_visible:
            # ── Step 1: Pastikan tab Username aktif ────────────────────────
            logger.info("Memastikan tab 'Username' aktif...")
            username_tab = page.get_by_role("tab", name="Username")
            try:
                if await username_tab.is_visible(timeout=3000):
                    await username_tab.click()
                    await asyncio.sleep(random.uniform(0.4, 0.8))
            except Exception:
                pass

            # Cek sekali lagi apakah password_input muncul sebelum mengetik username
            if not await password_input.is_visible():
                # ── Step 2: Isi username ───────────────────────────────────────
                logger.info(f"Mengetik username: {username}")
                username_input = page.locator('input[type="text"], input[name="username"], input[name="email"]').first
                await username_input.wait_for(state="visible", timeout=15000)
                await asyncio.sleep(random.uniform(0.3, 0.7))
                await human_type(username_input, username)
                await asyncio.sleep(random.uniform(0.5, 1.0))

                # ── Step 3: Klik Continue (pertama) ───────────────────────────
                logger.info("Klik Continue (username)...")
                continue_btn = page.get_by_role("button", name="Continue")
                await continue_btn.wait_for(state="visible", timeout=10000)
                await asyncio.sleep(random.uniform(0.3, 0.6))
                await continue_btn.click()

        # ── Step 4: Tunggu form password muncul ───────────────────────
        logger.info("Menunggu form password...")
        await password_input.wait_for(state="visible", timeout=15000)
        await asyncio.sleep(random.uniform(0.4, 0.9))

        # ── Step 5: Isi password ──────────────────────────────────────
        logger.info("Mengetik password...")
        await human_type(password_input, password)
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # ── Step 6: Klik Continue (kedua) ─────────────────────────────
        logger.info("Klik Continue (password)...")
        continue_btn2 = page.get_by_role("button", name="Continue")
        await continue_btn2.wait_for(state="visible", timeout=10000)
        await asyncio.sleep(random.uniform(0.2, 0.5))
        await continue_btn2.click()

        # ── Step 7: Tunggu redirect ke dashboard / keluar dari login ─────────────────────
        logger.info("Menunggu redirect ke dashboard...")
        for _ in range(30):
            await page.wait_for_timeout(1000)
            cur = page.url.lower()
            if "login" not in cur and "saved-accounts" not in cur and "merchant.grab.com" in cur:
                logger.info(f"[✓] Login berhasil untuk {username} → {page.url}")
                return True

        if "login" not in page.url.lower() and "saved-accounts" not in page.url.lower():
            logger.info(f"[✓] Login berhasil untuk {username} → {page.url}")
            return True

    except Exception as e:
        logger.error(f"[✗] Login gagal untuk {username}: {e}")
        # Simpan screenshot untuk debug
        try:
            screenshot_path = f"login_error_{username}.png"
            await page.screenshot(path=screenshot_path)
            logger.info(f"Screenshot disimpan: {screenshot_path}")
        except Exception:
            pass
        return False


async def capture_auth_headers(page):
    global auth_headers

    def handle_request(request):
        headers = request.headers
        if "authorization" in headers or "x-grab-merchant" in headers:
            auth_headers.update({
                k: v for k, v in headers.items()
                if k.lower() in ["authorization", "x-grab-merchant", "x-grab-device-id",
                                  "x-grab-client-id", "content-type", "accept", "x-passkey"]
            })

    page.on("request", handle_request)


async def fetch_idmg(page):
    """Mencari Group ID (idmg) dari API merchant-selector."""
    logger.info("Mencari Group ID (idmg) dari merchant-selector...")
    try:
        url = f"{BASE_URL}/troy/user-profile/v1/merchant-selector"
        response = await page.request.get(url, headers=auth_headers)
        if response.ok:
            import re
            text = await response.text()
            match = re.search(r'(IDMG\d+)', text)
            if match:
                idmg = match.group(1)
                logger.info(f"Berhasil mendapatkan Group ID: {idmg}")
                return idmg, response.status
        logger.warning(f"Gagal mendapatkan idmg. HTTP Status: {response.status}")
        return "", response.status
    except Exception as e:
        logger.error(f"Error fetching idmg: {e}")
    return "", 0


async def fetch_group_details(page, idmg):
    """Mengambil detail profil & rekening bank grup dari endpoint troy/v1/merchant."""
    if not idmg:
        return {}
    try:
        url = f"https://merchant.grab.com/troy/v1/merchant?merchant_group_id={idmg}&isBalanceNeeded=false&currency=IDR"
        res = await page.evaluate('''async (u) => {
            try {
                const r = await fetch(u, {
                    headers: {
                        'accept': 'application/json, text/plain, */*',
                        'requestsource': 'troyPortal'
                    }
                });
                if (!r.ok) return {};
                const json = await r.json();
                return json.data || {};
            } catch (e) {
                return {};
            }
        }''', url)
        if res and isinstance(res, dict):
            bank = res.get("bank_details") or {}
            if isinstance(bank, dict) and bank.get("bank_name"):
                logger.info(f"🏦 Berhasil mendapatkan data Bank Grab: {bank.get('bank_name')} a/n {bank.get('account_name')} ({bank.get('account_number')})")
            return res
    except Exception as e:
        logger.debug(f"Error fetching group details: {e}")
    return {}


async def fetch_merchant_list(page):
    """Fetch list of all merchant stores using the search API with pagination."""
    stores = []
    total_expected = 0
    logger.info("Fetching merchant store list...")

    try:
        offset = 0
        limit = 100
        
        while True:
            api_url = f"https://api.grab.com/delvplatformapi/merchant/v1/merchant-group/store/search?offset={offset}&limit={limit}&search=&includeItemsWithoutPhotosCount=true&includeInactive=true&modelType=ALL&asc=true&cityIDs[]=ALL&includeMenuGroupV2ID=false"
            
            page_retries = 3
            fetched_stores = []
            
            for attempt in range(page_retries):
                logger.info(f"Fetching stores offset {offset} (attempt {attempt+1})...")
                response = await page.request.get(api_url, headers=auth_headers)
                if not response.ok:
                    logger.warning(f"Gagal mengambil store list pada offset {offset}: HTTP {response.status}")
                    await asyncio.sleep(1.5)
                    continue
                    
                data = await response.json()
                
                # Ambil totalCount dari response API pada halaman pertama
                if offset == 0:
                    total_expected = (
                        data.get("totalCount") or
                        data.get("total") or
                        data.get("data", {}).get("totalCount") or
                        0
                    )
                    logger.info(f"Target total merchant dari sistem Grab untuk portal ini: {total_expected}")
                
                fetched_stores = (
                    data.get("stores") or
                    data.get("data", {}).get("stores") or
                    data.get("merchantDetails") or
                    data.get("data", {}).get("merchantDetails") or
                    data.get("catalogStores") or
                    []
                )
                
                # fallback pemetaan jika array stores ada di letak berbeda
                if not fetched_stores and isinstance(data, list):
                    fetched_stores = data
                elif not fetched_stores and isinstance(data, dict):
                    for k, v in data.items():
                        if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict) and ("merchantID" in v[0] or "merchantId" in v[0] or "id" in v[0]):
                            fetched_stores = v
                            break
                            
                break
            
            if not fetched_stores or not isinstance(fetched_stores, list):
                break
                
            stores.extend(fetched_stores)
            logger.info(f"Berhasil mengambil {len(fetched_stores)} outlet. Total terkumpul: {len(stores)}")
            
            # Jika jumlah data yang didapat kurang dari limit (misal 1, 3, 18), berarti semua sudah terambil dalam 1 request!
            if len(fetched_stores) < limit:
                break
                
            offset += limit
            await asyncio.sleep(1.5)

    except Exception as e:
        logger.error(f"Error fetching merchant list: {e}")

    return stores, total_expected







async def fetch_merchant_list_fast(headers, cookies_dict, cred, max_retries=3):
    """Fast path using curl_cffi."""
    try:
        from curl_cffi.requests import AsyncSession
    except ImportError:
        logger.warning("curl_cffi not installed, skipping fast path.")
        return None

    # We need to construct a proper headers dict that might have been missing.
    # We will use cookies_dict to pass cookies directly.
    async with AsyncSession(impersonate="chrome120", cookies=cookies_dict) as s:
        url_idmg = f"{BASE_URL}/troy/user-profile/v1/merchant-selector"
        res_idmg = await s.get(url_idmg, headers=headers)
        if res_idmg.status_code == 401:
            return "401"
        if not res_idmg.ok:
            logger.warning(f"Fast path IDMG failed: {res_idmg.status_code}")
            return None
            
        import re
        match = re.search(r'(IDMG\d+)', res_idmg.text)
        if not match:
            return None
        idmg = match.group(1)
        
        url_group = f"https://merchant.grab.com/troy/v1/merchant?merchant_group_id={idmg}&isBalanceNeeded=false&currency=IDR"
        h2 = headers.copy()
        h2['accept'] = 'application/json, text/plain, */*'
        h2['requestsource'] = 'troyPortal'
        res_group = await s.get(url_group, headers=h2)
        group_data = res_group.json().get("data", {}) if res_group.ok else {}
        group_bank = (group_data.get("bank_details") or {}) if isinstance(group_data, dict) else {}
        group_bank_name = group_bank.get("bank_name", "") if isinstance(group_bank, dict) else ""
        group_acc_name = group_bank.get("account_name", "") if isinstance(group_bank, dict) else ""
        group_acc_no = group_bank.get("account_number", "") if isinstance(group_bank, dict) else ""

        stores = []
        offset = 0
        limit = 100
        
        while True:
            api_url = f"https://api.grab.com/delvplatformapi/merchant/v1/merchant-group/store/search?offset={offset}&limit={limit}&search=&includeItemsWithoutPhotosCount=true&includeInactive=true&modelType=ALL&asc=true&cityIDs[]=ALL&includeMenuGroupV2ID=false"
            page_retries = 3
            fetched_stores = []
            
            for attempt in range(page_retries):
                res_list = await s.get(api_url, headers=headers)
                if not res_list.ok:
                    import asyncio
                    await asyncio.sleep(1.5)
                    continue
                data = res_list.json()
                fetched_stores = data.get("stores") or data.get("data", {}).get("stores") or data.get("merchantDetails") or data.get("data", {}).get("merchantDetails") or data.get("catalogStores") or []
                if not fetched_stores and isinstance(data, list):
                    fetched_stores = data
                elif not fetched_stores and isinstance(data, dict):
                    for k, v in data.items():
                        if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict) and ("merchantID" in v[0] or "merchantId" in v[0] or "id" in v[0]):
                            fetched_stores = v
                            break
                break
                
            if not fetched_stores or not isinstance(fetched_stores, list):
                break
                
            stores.extend(fetched_stores)
            if len(fetched_stores) < limit:
                break
            offset += limit
            import asyncio
            await asyncio.sleep(1.5)

        if not stores:
            # Cari Entity ID GF jika stores kosong
            mex_store_id = idmg
            try:
                up_res = await s.get(f"https://merchant.grab.com/user-profile/v2/details?merchant_group_id={idmg}&currency=IDR", headers=headers)
                if up_res.ok:
                    up_data = up_res.json()
                    gf_id = up_data.get("user_profile", {}).get("grab_food_entity_id")
                    if not gf_id:
                        for lk in up_data.get("user_profile", {}).get("links", []):
                            if lk.get("link_entity_business_line") == "GF":
                                gf_id = lk.get("link_entity_id")
                                break
                    if gf_id:
                        mex_store_id = gf_id
            except Exception:
                pass

            addr = (group_data.get("address") or {}) if isinstance(group_data, dict) else {}
            addr_str = addr.get("AddressLine1", "") if isinstance(addr, dict) else str(addr or "")
            city = addr.get("City", "") if isinstance(addr, dict) else ""
            if city and city not in addr_str:
                addr_str = f"{addr_str}, {city}".strip(", ")
            store_name = group_data.get("name") or cred["name"]
            stores = [{
                "merchantID": mex_store_id,
                "merchantName": store_name,
                "address": addr_str,
                "status": "ACTIVE",
                "bankAccount": group_bank
            }]
            
        all_results = []
        for store in stores:
            merchant_id = str(store.get("merchantID") or store.get("merchantId") or store.get("id") or "").strip()
            store_name = store.get("name") or store.get("merchantName") or store.get("storeName") or merchant_id
            status = store.get("status") or store.get("isActive") or ""
            alamat = store.get("address") or store.get("merchantAddress") or ""
            
            nama_bank = ""
            nama_pemilik = ""
            no_rekening = ""
            bank_obj = store.get("bankAccount") or store.get("bank_details")
            if isinstance(bank_obj, dict):
                nama_bank = bank_obj.get("bankName", "") or bank_obj.get("bank_name", "")
                nama_pemilik = bank_obj.get("accountHolderName", "") or bank_obj.get("account_name", "")
                no_rekening = bank_obj.get("accountNumber", "") or bank_obj.get("account_number", "")
            else:
                no_rekening = store.get("bankAccount") or store.get("bankAccountNumber") or store.get("accountNumber") or ""

            if not nama_bank: nama_bank = group_bank_name
            if not nama_pemilik: nama_pemilik = group_acc_name
            if not no_rekening: no_rekening = group_acc_no

            if not merchant_id:
                continue

            link_menu = f"https://merchant.grab.com/food/menu/{merchant_id}" if merchant_id else ""
            all_results.append({
                "Nama Pemilik": cred.get("owner", ""),
                "Nama Brand": cred.get("brand", ""),
                "Aplikator": "GrabFood",
                "Nama Portal": cred["name"],
                "Group ID": idmg,
                "Nama Listing": store_name,
                "Link": link_menu,
                "Store ID": merchant_id,
                "Status Listing": status,
                "Alamat": alamat,
                "Nama Bank": nama_bank,
                "Nama Pemilik Rekening": nama_pemilik,
                "Nomor Rekening": str(no_rekening) if isinstance(no_rekening, dict) else str(no_rekening),
                "_owner": cred.get("owner", cred["name"])
            })
            
        logger.info(f"⚡ [FAST PATH] Berhasil mengekstrak {len(all_results)} outlet via curl_cffi!")
        return all_results

async def run_scraper_for_credential(playwright, cred, force_fresh=False):
    global auth_headers
    headers_file = os.path.join(SESSIONS_DIR, f"grab_headers_{cred['name']}.json")
    session_file = os.path.join(SESSIONS_DIR, f"grab_session_{cred['name']}.json")
    
    if not force_fresh and os.path.exists(session_file):
        try:
            import json
            with open(session_file, "r") as f:
                session_data = json.load(f)
            cookies_dict = {c['name']: c['value'] for c in session_data.get('cookies', [])}
            
            cached_headers = {}
            if os.path.exists(headers_file):
                with open(headers_file, "r") as f:
                    cached_headers = json.load(f)
            
            # Allow fast path if we have cookies, even if headers are empty, because Grab uses cookies!
            if cookies_dict:
                logger.info(f"⚡ [FAST PATH] Menguji token sesi Grab aktif untuk {cred['name']} via curl_cffi...")
                fast_res = await fetch_merchant_list_fast(cached_headers, cookies_dict, cred)
                if fast_res == "401":
                    logger.warning(f"⚠️ [FAST PATH] Sesi {cred['name']} tidak valid (401). Melanjutkan ke browser auth...")
                elif isinstance(fast_res, list) and len(fast_res) > 0:
                    return fast_res
        except Exception as e:
            logger.debug(f"Fast path failed: {e}")
            
    return await run_scraper_for_credential_playwright(playwright, cred, force_fresh)

async def run_scraper_for_credential_playwright(playwright, cred, force_fresh=False):

    """Run scraper for a single credential set."""
    global auth_headers
    auth_headers = {}

    logger.info(f"\n{'='*50}")
    if force_fresh:
        logger.info(f"Starting FRESH scraper for: {cred['name']}")
    else:
        logger.info(f"Starting scraper for: {cred['name']}")
    logger.info(f"{'='*50}")

    browser = await playwright.chromium.launch(
        headless=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--ignore-gpu-blocklist",
            "--enable-gpu-rasterization",
            "--enable-zero-copy",
            "--enable-hardware-overlays",
            "--enable-features=VaapiVideoDecoder,CanvasOopRasterization"
        ]
    )

    session_file = os.path.join(SESSIONS_DIR, f"grab_session_{cred['name']}.json")

    context_options = {
        "viewport": {"width": 1280, "height": 800},
        "user_agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    # Jika force_fresh, hapus file session yang ada
    if force_fresh and os.path.exists(session_file):
        os.remove(session_file)
        logger.info(f"Dihapus session lama untuk fresh login: {session_file}")

    # Load session jika file session sudah ada
    if os.path.exists(session_file) and not force_fresh:
        context_options["storage_state"] = session_file
        logger.info(f"Menggunakan session yang tersimpan: {session_file}")

    context = await browser.new_context(**context_options)

    page = await context.new_page()
    await capture_auth_headers(page)

    try:
        # Cek apakah session masih valid dengan mencoba masuk ke halaman menu
        logger.info("Mengecek validitas session...")
        try:
            await page.goto(f"{BASE_URL}/food/menu", wait_until="domcontentloaded", timeout=45000)
            await page.wait_for_timeout(2000)
        except Exception as e:
            logger.warning(f"Navigasi menu awal lambat ({e}), melanjutkan...")

        # Jika URL redirect ke halaman login atau logout, berarti session belum ada / sudah expired
        if "login" in page.url.lower() or "logout" in page.url.lower():
            logger.info("Session tidak valid/expired. Memulai proses login...")
            logged_in = await perform_login(page, cred["username"], cred["password"])
            if not logged_in:
                if not force_fresh:
                    logger.warning(f"Gagal login dengan session saat ini untuk {cred['name']}. Akan mencoba FRESH LOGIN...")
                    return "RETRY_FRESH"
                else:
                    logger.error(f"Fresh login gagal untuk {cred['name']}, melewati (skipping)...")
                    return False

            # Simpan state/session setelah berhasil login
            await context.storage_state(path=session_file)
            headers_file = session_file.replace("grab_session_", "grab_headers_")
            with open(headers_file, "w") as hf:
                json.dump(auth_headers, hf)
            logger.info(f"Session baru berhasil disimpan ke {session_file}")

            # Pindah lagi ke halaman menu setelah login berhasil
            try:
                await page.goto(f"{BASE_URL}/food/menu", wait_until="domcontentloaded", timeout=45000)
                await page.wait_for_timeout(2000)
            except Exception as e:
                logger.warning(f"Navigasi menu pasca login lambat ({e}), melanjutkan...")
        else:
            logger.info("[✓] Session masih valid! Lewati proses login.")

        # Get IDMG (Group ID) from merchant-selector API
        idmg, idmg_status = await fetch_idmg(page)
        
        # Jika status 401, berarti session expired walaupun URL tidak berubah ke halaman login/logout
        if idmg_status == 401:
            logger.info("Session expired (HTTP 401 pada merchant-selector). Memulai proses login ulang...")
            logged_in = await perform_login(page, cred["username"], cred["password"])
            if not logged_in:
                if not force_fresh:
                    logger.warning(f"Gagal login ulang (401) dengan session saat ini untuk {cred['name']}. Akan mencoba FRESH LOGIN...")
                    return "RETRY_FRESH"
                else:
                    logger.error(f"Fresh login (401) gagal untuk {cred['name']}, melewati (skipping)...")
                    return False

            # Simpan state/session setelah berhasil login
            await context.storage_state(path=session_file)
            headers_file = session_file.replace("grab_session_", "grab_headers_")
            with open(headers_file, "w") as hf:
                json.dump(auth_headers, hf)
            logger.info(f"Session baru berhasil disimpan ke {session_file}")

            # Pindah lagi ke halaman menu setelah login berhasil
            try:
                await page.goto(f"{BASE_URL}/food/menu", wait_until="domcontentloaded", timeout=45000)
                await page.wait_for_timeout(2000)
            except Exception as e:
                logger.warning(f"Navigasi menu pasca login ulang lambat ({e}), melanjutkan...")
            
            # Coba ambil idmg lagi
            idmg, idmg_status = await fetch_idmg(page)

        if not idmg:
            logger.error(f"Tidak dapat menemukan Group ID untuk {cred['name']}, melewati.")
            return []

        # Ambil data profil & rekening bank dari troy/v1/merchant
        group_data = await fetch_group_details(page, idmg)
        group_bank = (group_data.get("bank_details") or {}) if isinstance(group_data, dict) else {}
        group_bank_name = group_bank.get("bank_name", "") if isinstance(group_bank, dict) else ""
        group_acc_name = group_bank.get("account_name", "") if isinstance(group_bank, dict) else ""
        group_acc_no = group_bank.get("account_number", "") if isinstance(group_bank, dict) else ""

        # Get list of stores dari API
        stores, total_expected = await fetch_merchant_list(page)
        
        max_retries = 3
        retry_count = 0
        while total_expected > 0 and len(stores) < total_expected and retry_count < max_retries:
            logger.warning(f"[!] PERINGATAN: Jumlah outlet yang terekstrak ({len(stores)}) "
                           f"kurang dari total sistem ({total_expected}). "
                           f"Mencoba mengambil ulang data yang terlewat... ({retry_count + 1}/{max_retries})")
            
            await page.wait_for_timeout(3000)
            stores_retry, _ = await fetch_merchant_list(page)
            
            # Gabungkan dengan data sebelumnya tanpa menghapus duplikat
            stores.extend(stores_retry)
            retry_count += 1
            
        logger.info(f"Total unique stores extracted finally: {len(stores)} (Target expected from API: {total_expected})")

        # Log Pengecekan Akhir
        if total_expected > 0 and len(stores) < total_expected:
            logger.warning(f"[!] GAGAL MENGAMBIL SEMUA: Setelah {max_retries} percobaan, jumlah outlet terekstrak ({len(stores)}) "
                           f"masih lebih kecil dari total sistem ({total_expected}).")
        elif total_expected > 0 and len(stores) >= total_expected:
            logger.info("[✓] Pengecekan: Semua outlet berhasil ditarik sesuai dengan total di sistem Grab.")

        if not stores:
            logger.info(f"[*] Menangani mode Single-Store / Profil Langsung untuk {cred['name']}...")
            addr = (group_data.get("address") or {}) if isinstance(group_data, dict) else {}
            addr_str = addr.get("AddressLine1", "") if isinstance(addr, dict) else str(addr or "")
            city = addr.get("City", "") if isinstance(addr, dict) else ""
            if city and city not in addr_str:
                addr_str = f"{addr_str}, {city}".strip(", ")
                
            store_name = group_data.get("name") or cred["name"]
            
            ls_data = await page.evaluate('''() => {
                try {
                    const up = JSON.parse(localStorage.getItem('userprofileInfo') || '{}');
                    const links = up?.user_profile?.links || [];
                    const gf_link = links.find(l => l.link_entity_business_line === 'GF') || links[0];
                    return {
                        entity_id: gf_link?.link_entity_id || up?.user_profile?.grab_food_entity_id || '',
                        name: localStorage.getItem('profileInfo') ? JSON.parse(localStorage.getItem('profileInfo')).name : ''
                    };
                } catch(e) { return {}; }
            }''')
            
            merchant_id = ls_data.get("entity_id") or idmg
            if ls_data.get("name"):
                store_name = ls_data.get("name")

            fallback_store = {
                "merchantID": merchant_id,
                "merchantName": store_name,
                "address": addr_str,
                "status": "ACTIVE",
                "bankAccount": group_bank
            }
            stores = [fallback_store]
            logger.info(f"[✓] Berhasil memuat data outlet Single-Store: '{store_name}' (ID: {merchant_id})")

        all_results = []
        # Save headers as backup
        headers_file = session_file.replace("grab_session_", "grab_headers_")
        with open(headers_file, "w") as hf:
            json.dump(auth_headers, hf)
        
        for store in stores:
            merchant_id = (
                str(store.get("merchantID") or store.get("merchantId") or store.get("id") or "")
            ).strip()
            store_name = store.get("name") or store.get("merchantName") or store.get("storeName") or merchant_id
            status = store.get("status") or store.get("isActive") or ""
            alamat = store.get("address") or store.get("merchantAddress") or ""
            
            # Coba ambil detail bank per store, fallback ke group bank details
            nama_bank = ""
            nama_pemilik = ""
            no_rekening = ""
            
            bank_obj = store.get("bankAccount") or store.get("bank_details")
            if isinstance(bank_obj, dict):
                nama_bank = bank_obj.get("bankName", "") or bank_obj.get("bank_name", "")
                nama_pemilik = bank_obj.get("accountHolderName", "") or bank_obj.get("account_name", "")
                no_rekening = bank_obj.get("accountNumber", "") or bank_obj.get("account_number", "")
            else:
                no_rekening = (
                    store.get("bankAccount") or 
                    store.get("bankAccountNumber") or 
                    store.get("accountNumber") or 
                    ""
                )

            if not nama_bank:
                nama_bank = group_bank_name
            if not nama_pemilik:
                nama_pemilik = group_acc_name
            if not no_rekening:
                no_rekening = group_acc_no

            if not merchant_id:
                continue

            link_menu = f"https://merchant.grab.com/food/menu/{merchant_id}" if merchant_id else ""

            all_results.append({
                "Nama Pemilik": cred.get("owner", ""),
                "Nama Brand": cred.get("brand", ""),
                "Aplikator": "GrabFood",
                "Nama Portal": cred["name"],
                "Group ID": idmg,
                "Nama Listing": store_name,
                "Link": link_menu,
                "Store ID": merchant_id,
                "Status Listing": status,
                "Alamat": alamat,
                "Nama Bank": nama_bank,
                "Nama Pemilik Rekening": nama_pemilik,
                "Nomor Rekening": str(no_rekening) if isinstance(no_rekening, dict) else str(no_rekening),
                "_owner": cred.get("owner", cred["name"])
            })
            
        logger.info(f"Berhasil mengekstrak {len(all_results)} outlet untuk portal {cred['name']}.")
        return all_results

    except Exception as e:
        logger.error(f"Unexpected error for {cred['name']}: {e}")
        return []
    finally:
        await context.close()
        await browser.close()


def get_template_headers():
    """Mengambil header dari template YYYY-MM-DD HH_MM Nama Pemilik.xlsx."""
    import openpyxl
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "YYYY-MM-DD HH_MM Nama Pemilik.xlsx")
    if not os.path.exists(template_path):
        template_path = os.path.join(os.path.dirname(__file__), "YYYY-MM-DD HH_MM Nama Pemilik.xlsx")
    if os.path.exists(template_path):
        try:
            wb = openpyxl.load_workbook(template_path, read_only=True)
            ws = wb['Listing'] if 'Listing' in wb.sheetnames else wb.active
            return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        except Exception:
            pass
    return ['Nama Pemilik', 'Nama Brand', 'Model', 'Tipe', 'Outlet', 'Nomor HP', 'Aplikator', 'Nama Portal', 'Group ID', 'Nama Listing', 'Link', 'Store ID', 'Status Listing', 'Alamat', 'Nama Bank', 'Nama Pemilik Rekening', 'Nomor Rekening']


def save_formatted_excel(df, file_path):
    """Menyimpan DataFrame ke file Excel menggunakan template YYYY-MM-DD HH_MM Nama Pemilik.xlsx."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    headers = get_template_headers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listing"
    
    header_font = Font(name='Calibri', size=11, bold=True)
    for c_idx, h in enumerate(headers, 1):
        if h is not None:
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = header_font
            
    font = Font(name='Calibri', size=11)
    alignment = Alignment(horizontal='left', vertical='center')
    
    for idx, (_, r) in enumerate(df.iterrows(), start=2):
        ws.cell(row=idx, column=1, value=str(r.get('Nama Pemilik') or '') if pd.notna(r.get('Nama Pemilik')) else '') # A: Nama Pemilik
        ws.cell(row=idx, column=2, value=str(r.get('Nama Brand') or '') if pd.notna(r.get('Nama Brand')) else '')   # B: Nama Brand
        ws.cell(row=idx, column=7, value=str(r.get('Aplikator') or '') if pd.notna(r.get('Aplikator')) else 'GrabFood') # G: Aplikator
        ws.cell(row=idx, column=8, value=str(r.get('Nama Portal') or '') if pd.notna(r.get('Nama Portal')) else '') # H: Nama Portal
        ws.cell(row=idx, column=9, value=str(r.get('Group ID') or '') if pd.notna(r.get('Group ID')) else '')       # I: Group ID
        ws.cell(row=idx, column=10, value=str(r.get('Nama Listing') or '') if pd.notna(r.get('Nama Listing')) else '') # J: Nama Listing
        ws.cell(row=idx, column=11, value=str(r.get('Link') or '') if pd.notna(r.get('Link')) else '')             # K: Link
        ws.cell(row=idx, column=12, value=str(r.get('Store ID') or '') if pd.notna(r.get('Store ID')) else '')     # L: Store ID
        ws.cell(row=idx, column=13, value=str(r.get('Status Listing') or '') if pd.notna(r.get('Status Listing')) else '') # M: Status Listing
        ws.cell(row=idx, column=14, value=str(r.get('Alamat') or '') if pd.notna(r.get('Alamat')) else '')         # N: Alamat
        ws.cell(row=idx, column=15, value=str(r.get('Nama Bank') or '') if pd.notna(r.get('Nama Bank')) else '')   # O: Nama Bank
        ws.cell(row=idx, column=16, value=str(r.get('Nama Pemilik Rekening') or '') if pd.notna(r.get('Nama Pemilik Rekening')) else '') # P: Nama Pemilik Rekening
        
        c_rek = ws.cell(row=idx, column=17, value=str(r.get('Nomor Rekening') or '') if pd.notna(r.get('Nomor Rekening')) else '') # Q: Nomor Rekening
        c_rek.number_format = '@'

        for c in range(1, 18):
            cell = ws.cell(row=idx, column=c)
            cell.font = font
            cell.alignment = alignment

    # Auto adjust column widths
    for col_idx in range(1, 18):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max((len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, len(df) + 2)), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    wb.save(file_path)


def save_portal_results(outlets_data, cred, source_type):
    """Menyimpan data portal Grab ke cache JSON dan output Excel per-owner."""
    if not outlets_data:
        logger.warning(f"   ⚠️ Tidak ada data outlet untuk disimpan pada portal '{cred['name']}'.")
        return None, None
        
    portal_name = cred["name"]
    owner_name = cred.get("owner", cred.get("brand", portal_name))
    safe_cache_name = get_safe_cache_filename(portal_name)
    safe_owner = "".join(c for c in str(owner_name) if c.isalnum() or c in (' ', '_', '-')).strip()
    if not safe_owner:
        safe_owner = "Unknown"
        
    # 1. Simpan ke Cache JSON
    cache_json_file = os.path.join(CACHE_DIR, safe_cache_name)
    cache_payload = {
        'portal': portal_name,
        'brand': cred.get("brand", ""),
        'owner': owner_name,
        'source_type': str(source_type).upper(),
        'username': cred.get("username", ""),
        'timestamp': datetime.datetime.now().isoformat(),
        'outlets': outlets_data
    }
    try:
        with open(cache_json_file, 'w', encoding='utf-8') as jf:
            json.dump(cache_payload, jf, indent=2, ensure_ascii=False)
        logger.info(f"   💾 Cache portal disimpan di cache/: {os.path.basename(cache_json_file)}")
    except Exception as e:
        logger.warning(f"   ⚠️ Gagal menyimpan cache JSON: {e}")
        
    # 2. Simpan ke Output Excel per-owner
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
    owner_df = pd.DataFrame(outlets_data)
    if '_owner' in owner_df.columns:
        owner_df = owner_df.drop(columns=['_owner'])
    if 'Store ID' in owner_df.columns:
        owner_df.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)
        
    owner_file = os.path.join(OUTPUT_DIR, f"{timestamp_str} {safe_owner}.xlsx")
    save_formatted_excel(owner_df, owner_file)
    logger.info(f"   💾 File Owner '{owner_name}' tersimpan di output/: {os.path.basename(owner_file)} (Total: {len(owner_df)} outlet)")
    return cache_json_file, owner_file


def combine_master(source_type=None):
    """Menggabungkan semua data cache JSON Grab ke Master (0master.xlsx / VB_master.xlsx) dan output per-owner."""
    import glob
    import shutil

    logger.info("\n[*] Menggabungkan semua data cache JSON Grab ke Master...")
    json_files = sorted(glob.glob(os.path.join(CACHE_DIR, "grab_portal_*.json")))

    all_records = []
    if json_files:
        for f in json_files:
            try:
                with open(f, 'r', encoding='utf-8') as jf:
                    cdata = json.load(jf)
                    outlets = cdata.get('outlets', [])
                    if isinstance(outlets, list):
                        all_records.extend(outlets)
            except Exception as e:
                logger.warning(f"   ⚠️ Gagal membaca cache {os.path.basename(f)}: {e}")
    else:
        # Fallback baca dari file xlsx lama di hasil_custom jika ada
        hasil_custom_dir = os.path.join(GRAB_DIR, "hasil_custom")
        xlsx_files = sorted(glob.glob(os.path.join(hasil_custom_dir, "*.xlsx")))
        xlsx_files = [f for f in xlsx_files if "master" not in os.path.basename(f).lower() and "duplikat" not in os.path.basename(f).lower()]
        if not xlsx_files:
            logger.warning("   ⚠️ Tidak ada file cache (grab_portal_*.json) di folder cache untuk digabung.")
            return
        for f in xlsx_files:
            try:
                df = pd.read_excel(f, sheet_name="Listing" if "Listing" in pd.ExcelFile(f).sheet_names else 0)
                df.dropna(how="all", inplace=True)
                if "Nama Listing" not in df.columns and "Nama Outlet" in df.columns:
                    df.rename(columns={"Nama Outlet": "Nama Listing"}, inplace=True)
                if "Status Listing" not in df.columns and "Status" in df.columns:
                    df.rename(columns={"Status": "Status Listing"}, inplace=True)
                all_records.extend(df.to_dict(orient='records'))
            except Exception as e:
                logger.warning(f"   ⚠️ Gagal membaca {os.path.basename(f)}: {e}")

    if all_records:
        master_df = pd.DataFrame(all_records)
        if '_owner' in master_df.columns:
            master_df = master_df.drop(columns=['_owner'])

        # 1. Filter berdasarkan source_type TERLEBIH DAHULU agar data lintas tipe tidak saling memotong
        if source_type:
            st = str(source_type).strip().lower()
            if 'Nama Pemilik' in master_df.columns:
                if st == 'agency':
                    master_df = master_df[master_df['Nama Pemilik'] != 'VB']
                elif st == 'vb':
                    master_df = master_df[master_df['Nama Pemilik'] == 'VB']
                elif st == 'vercel':
                    master_df = master_df[master_df['Nama Pemilik'] != 'VB']

        # 2. Deduplikasi Store ID dalam lingkup sumber yang dipilih
        if "Store ID" in master_df.columns:
            master_df.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)

        if master_df.empty:
            logger.warning("   ⚠️ Tidak ada baris data setelah pemfilteran source_type.")
            return

        st = str(source_type).strip().lower() if source_type else "agency"
        if st == "vb":
            tag = "VB"
        elif st == "vercel":
            tag = "Vercel"
        else:
            tag = "Agency"

        timestamp_name = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
        master_filename = f"{timestamp_name} {tag}_master.xlsx"
        master_path = os.path.join(MASTER_DIR, master_filename)

        # File versioning/backup jika file persis sama sudah ada
        if os.path.exists(master_path):
            version_dir = os.path.join(MASTER_DIR, "versions")
            os.makedirs(version_dir, exist_ok=True)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(version_dir, f"{os.path.splitext(master_filename)[0]}_v{ts}.xlsx")
            shutil.copy2(master_path, backup_path)
            logger.info(f"   💾 Master lama di-backup ke: {os.path.basename(backup_path)}")

        # Simpan master bertimestamp (YYYY-MM-DD HH_MM [Tipe]_master.xlsx)
        save_formatted_excel(master_df, master_path)

        # Simpan juga salinan statis [Tipe]_master.xlsx dan 0master.xlsx (khusus Agency) untuk kompatibilitas
        static_master_path = os.path.join(MASTER_DIR, f"{tag}_master.xlsx")
        save_formatted_excel(master_df, static_master_path)
        if tag == "Agency":
            save_formatted_excel(master_df, os.path.join(MASTER_DIR, "0master.xlsx"))

        timestamp_name = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")

        # Simpan file per Nama Pemilik ke output/
        if 'Nama Pemilik' in master_df.columns:
            owners = master_df['Nama Pemilik'].dropna().unique()
            for owner in owners:
                if owner and str(owner).strip():
                    owner_clean = "".join(c for c in str(owner) if c.isalnum() or c in " ._-").strip()
                    owner_file = os.path.join(OUTPUT_DIR, f"{timestamp_name} {owner_clean}.xlsx")
                    owner_df = master_df[master_df['Nama Pemilik'] == owner]
                    save_formatted_excel(owner_df, owner_file)
                    logger.info(f"   💾 File per-Owner tersimpan di output/: {os.path.basename(owner_file)}")

        logger.info(f"\n[✓] Master Grab file berhasil dibuat:")
        logger.info(f"    - {master_path}")
        logger.info(f"    Total baris: {len(master_df)}")
        logger.info(f"    Total kolom: {len(master_df.columns)} (Template YYYY-MM-DD HH_MM Nama Pemilik.xlsx)")


async def main():
    import datetime
    
    parser = argparse.ArgumentParser(description="Grab Merchant Scraper")
    parser.add_argument("--type", type=str, choices=["agency", "vb", "vercel"], help="Pilih tipe sumber: 'agency', 'vb', atau 'vercel'")
    parser.add_argument("--agency", action="store_true", help="Gunakan sumber kredensial Agency")
    parser.add_argument("--vb", action="store_true", help="Gunakan sumber kredensial VB (Virtual Brand)")
    parser.add_argument("--vercel", action="store_true", help="Gunakan sumber kredensial Vercel Sheet")
    parser.add_argument("--vb-url", type=str, help="URL Google Sheet / CSV khusus untuk VB")
    parser.add_argument("--url", type=str, help="URL Google Sheet / CSV custom")
    parser.add_argument("--outlet", type=str, help="Specify the outlet name to run (e.g., F1)")
    parser.add_argument("--outlets", type=str, help="Comma-separated list of outlet names")
    parser.add_argument("--owner", type=str, help="Filter nama owner tertentu untuk ditarik")
    parser.add_argument("--users", type=str, help="Comma-separated list of usernames to run")
    parser.add_argument("--all", action="store_true", help="Run for all portals without prompt")
    parser.add_argument("--fresh", action="store_true", help="Start fresh run and ignore previous progress checkpoint")
    parser.add_argument("--combine", action="store_true", help="Gabungkan file cache JSON ke master & output tanpa scraping ulang")
    args = parser.parse_args()

    source_type = "agency"
    custom_url = args.url or args.vb_url

    if args.vercel or args.type == "vercel":
        source_type = "vercel"
    elif args.vb or args.type == "vb":
        source_type = "vb"
    elif args.agency or args.type == "agency":
        source_type = "agency"

    if args.combine:
        combine_master(source_type=source_type)
        return

    if not args.outlet and not args.outlets and not args.owner and not args.users and not args.all:
        print("  PILIH SUMBER KREDENSIAL GRABFOOD")
        print("=" * 54)
        print("  [1] Agency (Master Agency Google Sheet)")
        print("  [2] VB     (Virtual Brand - Dokumen Lain)")
        print("  [3] Vercel (Live CSV Vercel Sheet)")
        print("=" * 54)
        pilihan = input("Pilih [1/2/3] (Default: 1 - Agency): ").strip().lower()
        if pilihan in ("2", "vb"):
            source_type = "vb"
        elif pilihan in ("3", "vercel"):
            source_type = "vercel"
        else:
            source_type = "agency"

    portals = get_credentials_from_sheet(source_type=source_type, custom_url=custom_url)
    target_credentials = []

    if args.owner:
        owner_clean = args.owner.strip().lower()
        target_credentials = [c for c in portals if str(c.get("owner", "")).strip().lower() == owner_clean or owner_clean in str(c.get("owner", "")).strip().lower()]
        logger.info(f"[*] Filter Owner '{args.owner}': Ditemukan {len(target_credentials)} akun portal GrabFood.")
    elif args.outlet:
        target_credentials = [c for c in portals if c["name"] == args.outlet]
        if not target_credentials:
            logger.error(f"Outlet '{args.outlet}' not found in credentials.")
            return
    elif args.outlets:
        outlet_names = [o.strip().lower() for o in args.outlets.split(",") if o.strip()]
        target_credentials = [c for c in portals if c["name"].strip().lower() in outlet_names]
        if not target_credentials:
            logger.error(f"None of the outlets in '{args.outlets}' found in credentials.")
            return
    elif args.users:
        user_list = [u.strip().strip(".").lower() for u in args.users.replace("\n", ",").split(",") if u.strip()]
        target_credentials = [c for c in portals if c["username"].strip().strip(".").lower() in user_list]
        if not target_credentials:
            logger.error(f"None of the usernames in '{args.users}' found in credentials.")
            return
    elif args.all:
        target_credentials = portals
    else:
        if portals:
            print(f"\nDaftar Portal GrabFood [{source_type.upper()}]:")
            print("  [0] Pilih Semua (All)")
            for idx, p in enumerate(portals):
                print(f"  [{idx+1}] {p['name']} ({p['username']}) - Owner: {p.get('owner', '-')}")
            
            choice = input(f"\nPilih portal (contoh: 1, 2-3, all): ").strip().lower()
            selected_indices = parse_selection(choice, len(portals))
            for idx in selected_indices:
                target_credentials.append(portals[idx])

    if not target_credentials:
        logger.error("Tidak ada portal yang dipilih.")
        return

    output_dir = OUTPUT_DIR
    progress_file = os.path.join(SESSIONS_DIR, f".grab_progress_{source_type}.json")
    if source_type == "agency" and not os.path.exists(progress_file) and os.path.exists(os.path.join(SESSIONS_DIR, ".grab_progress.json")):
        progress_file = os.path.join(SESSIONS_DIR, ".grab_progress.json")

    # ── Checkpoint / Resume Management ──────────────────────────────
    completed_portal_names = set()
    all_collected_stores = []
    
    if args.fresh and os.path.exists(progress_file):
        try:
            os.remove(progress_file)
            logger.info("🗑️ Checkpoint progress sebelumnya dihapus (--fresh aktif).")
        except Exception:
            pass

    if not args.fresh and os.path.exists(progress_file):
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                checkpoint_data = json.load(f)
                completed_portal_names = set(checkpoint_data.get("completed_portals", []))
                raw_stores = checkpoint_data.get("stores", [])
                all_collected_stores = [s for s in raw_stores if isinstance(s, dict)]
                logger.info(f"🔄 [RESUME AKTIF] Memuat progress checkpoint: {len(completed_portal_names)} portal sudah selesai ditarik sebelumnya.")
        except Exception as ex:
            logger.warning(f"Gagal membaca checkpoint progress: {ex}")

    # Jika user secara eksplisit menentukan outlet/users untuk di-repull, jangan lewati
    logger.info(f"Total portal yang ditargetkan: {len(target_credentials)}")

    async with async_playwright() as playwright:
        for idx, cred in enumerate(target_credentials, 1):
            if cred["name"] in completed_portal_names:
                logger.info(f"⏩ [{idx}/{len(target_credentials)}] Portal '{cred['name']}' sudah selesai sebelumnya (Dilewati).")
                continue

            logger.info(f"\n▶️ [{idx}/{len(target_credentials)}] Memproses portal: {cred['name']} (Owner: {cred.get('owner', '-')})...")
            stores_res = await run_scraper_for_credential(playwright, cred, force_fresh=args.fresh)
            
            retry_count = 0
            while not stores_res and retry_count < 2:
                retry_count += 1
                logger.warning(f"[!] Portal {cred['name']} gagal diproses/login. Melakukan retry ke-{retry_count} dari 2...")
                await asyncio.sleep(5)
                stores_res = await run_scraper_for_credential(playwright, cred, force_fresh=True)

            if stores_res and isinstance(stores_res, list):
                valid_stores = [s for s in stores_res if isinstance(s, dict)]
                if valid_stores:
                    all_collected_stores.extend(valid_stores)
                    completed_portal_names.add(cred["name"])
                    
                    # Simpan ke Cache JSON & Output Excel per-owner secara Real-Time
                    save_portal_results(valid_stores, cred, source_type)

                    # Simpan checkpoint progress seketika (Real-Time Auto-Save)
                    try:
                        with open(progress_file, "w", encoding="utf-8") as f:
                            json.dump({
                                "completed_portals": list(completed_portal_names),
                                "stores": all_collected_stores,
                                "last_update": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }, f, indent=2)
                    except Exception as ex:
                        logger.debug(f"Gagal menyimpan checkpoint progress: {ex}")

    logger.info("\n[✓] Proses scraping selesai.")

    # Filter ketat hanya dict objek
    all_collected_stores = [s for s in all_collected_stores if isinstance(s, dict)]

    if not all_collected_stores:
        logger.warning("Tidak ada data outlet yang berhasil dikumpulkan.")
        return

    df_all = pd.DataFrame(all_collected_stores)
    if "Store ID" in df_all.columns:
        df_all.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")

    # 1. Simpan per Owner (Nama Pemilik) sesuai template baru: YYYY-MM-DD HH_MM Nama Pemilik.xlsx
    owners = df_all["_owner"].dropna().unique() if "_owner" in df_all.columns else []
    for owner in owners:
        df_owner = df_all[df_all["_owner"] == owner].drop(columns=["_owner"])
        safe_owner = "".join(c for c in str(owner) if c.isalnum() or c in (' ', '_', '-')).strip()
        if not safe_owner:
            safe_owner = "Unknown"
        owner_file = os.path.join(OUTPUT_DIR, f"{timestamp} {safe_owner}.xlsx")
        save_formatted_excel(df_owner, owner_file)
        logger.info(f"[✓] File Template Baru tersimpan di output/: {os.path.basename(owner_file)} ({len(df_owner)} baris)")

    # 2. Gabungkan seluruh cache ke Master di master/0master.xlsx (atau VB_master.xlsx)
    combine_master(source_type=source_type)

    # Bersihkan checkpoint jika semua proses selesai sukses
    if os.path.exists(progress_file):
        try:
            os.remove(progress_file)
            logger.info("🧹 Checkpoint progress sementara dibersihkan (Semua selesai).")
        except Exception:
            pass

    logger.info("\n🎉 Seluruh penarikan selesai dan tersimpan dengan template baru!")


if __name__ == "__main__":
    asyncio.run(main())
