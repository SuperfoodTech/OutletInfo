import pandas as pd
import pandas.io.formats.excel
pandas.io.formats.excel.ExcelFormatter.header_style = None

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border

def apply_formatting_and_sheets(input_file, output_file, create_tabs=False):
    try:
        # Read the input excel
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"Error reading {input_file}: {e}")
        return

    # Define the desired tab and sorting order
    desired_order = ["F1", "F2S", "W1", "L1", "L2", "DE1S", "JF1", "JF1S"]

    # Check what the portal column is named. It could be "Sumber" or "Portal".
    portal_col = None
    for col in ["Portal", "Sumber"]:
        if col in df.columns:
            portal_col = col
            break

    # Sort data in Main Tab if portal_col exists
    if portal_col:
        portal_rank = {p: i for i, p in enumerate(desired_order)}
        df['_sort_rank'] = df[portal_col].map(lambda x: portal_rank.get(x, 999))
        df = df.sort_values(['_sort_rank', portal_col]).drop(columns=['_sort_rank']).reset_index(drop=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Drop the portal column if it exists before writing
        df_to_write = df.drop(columns=[portal_col]) if portal_col else df
        
        # Write "Main Tab" sheet (formerly "All")
        df_to_write.to_excel(writer, sheet_name='Main Tab', index=False)
        
        # Write individual sheets for each portal in the specified order if requested
        if create_tabs and portal_col:
            for portal in desired_order:
                portal_df = df[df[portal_col] == portal].drop(columns=[portal_col])
                sheet_name = str(portal)[:31]
                portal_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
            # Create tabs for any other unexpected portals that might exist
            other_portals = [p for p in df[portal_col].dropna().unique() if p not in desired_order]
            for portal in sorted(other_portals):
                portal_df = df[df[portal_col] == portal].drop(columns=[portal_col])
                sheet_name = str(portal)[:31]
                portal_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Now apply formatting to all created sheets
    try:
        wb = load_workbook(output_file)
        font = Font(name='Arial', size=10, bold=False)
        alignment = Alignment(horizontal='left', vertical='center')
        no_border = Border()
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = no_border
                    
            # Optionally adjust column widths slightly for better readability
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                # Cap the width so it doesn't get ridiculously wide
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
    except Exception as e:
        print(f"Error applying formatting to {output_file}: {e}")
