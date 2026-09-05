#!/usr/bin/env python3
import os
import re
import sys
import json
import time
import csv
import random
import requests
import shutil
import datetime
from pathlib import Path
import urllib.request
from urllib.request import urlopen
from urllib.parse import urlparse, urlencode, urlunparse, parse_qsl
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright

# Muat file .env dari folder proyek
BASE_DIR = Path(__file__).parent.parent
ENV_PATH = BASE_DIR / ".env"
load_dotenv(ENV_PATH)

# --- CONFIG ---
APP_SCRIPT_URL = os.getenv("APP_SCRIPT_URL", "")
OTP_ENDPOINT_URL = os.getenv("OTP_ENDPOINT_URL", "https://script.google.com/macros/s/AKfycbwRViqfGkDtQGmUDD0PycfSGyEBPgx2uaxelHdKIr__4rZ5aq41j1En5Wb96CgEmRvM/exec")
GMAIL_OTP_LABEL = os.getenv("GMAIL_OTP_LABEL", "OTP-GO")

SESSION_DIR = Path(__file__).parent / "session"
SESSION_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR = Path(__file__).parent / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
MASTER_DIR = Path(__file__).parent / "master"
MASTER_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)


def normalisasi_nomor_hp(nomor_hp):
    if not nomor_hp:
        return ""
    nomor_hp = str(nomor_hp).strip()
    if "@" in nomor_hp:
        return nomor_hp
    nomor_hp = re.sub(r'\D', '', nomor_hp)
    if nomor_hp.startswith("62"):
        return nomor_hp[2:]
    if nomor_hp.startswith("0"):
        return nomor_hp[1:]
    return nomor_hp


def load_gofood_session(identifier):
    if not identifier:
        return None
    ident_str = str(identifier).strip().lower()
    sanitized = re.sub(r'[^a-zA-Z0-9_.-]', '_', ident_str)
    json_file = SESSION_DIR / f"session_gofood_{sanitized}.json"
    if not json_file.exists():
        phone_norm = normalisasi_nomor_hp(identifier)
        if phone_norm:
            json_file_old = SESSION_DIR / f"session_{phone_norm}.json"
            if json_file_old.exists():
                json_file = json_file_old
    if not json_file.exists():
        return None
    try:
        with open(json_file, 'r') as f:
            return json.load(f)
    except Exception:
        return None


def save_gofood_session(identifier, cookies, access_token=None):
    if not identifier:
        return
    ident_str = str(identifier).strip().lower()
    if ident_str in ("", "-", "nan", "none", "null"):
        return
        
    if isinstance(cookies, dict):
        session_data = cookies
    else:
        session_data = {
            "email": identifier,
            "access_token": access_token,
            "cookies": cookies,
            "timestamp": datetime.datetime.now().isoformat()
        }

    sanitized = re.sub(r'[^a-zA-Z0-9_.-]', '_', ident_str)
    json_file = SESSION_DIR / f"session_gofood_{sanitized}.json"
    try:
        os.makedirs(os.path.dirname(json_file), exist_ok=True)
        with open(json_file, 'w') as f:
            json.dump(session_data, f, indent=4)
        phone_norm = normalisasi_nomor_hp(identifier)
        if phone_norm and phone_norm != sanitized:
            json_file_old = SESSION_DIR / f"session_{phone_norm}.json"
            with open(json_file_old, 'w') as f:
                json.dump(session_data, f, indent=4)
        print(f"   💾 Sesi login berhasil disimpan untuk {identifier}.")
    except Exception as e:
        print(f"   ⚠️ Gagal menyimpan berkas sesi untuk {identifier}: {e}")


def ambil_otp_dari_endpoint(url_dasar, action="getOtpEmail", label_email=None):
    if not url_dasar:
        return ""

    if "docs.google.com/spreadsheets" in url_dasar:
        try:
            with urlopen(url_dasar, timeout=30) as response:
                content = response.read().decode("utf-8").strip()
                lines = content.splitlines()
                if not lines or len(lines) < 2:
                    return ""
                reader = csv.reader(lines)
                rows = list(reader)
                headers = [h.strip().lower() for h in rows[0]]
                
                otp_idx = -1
                for idx, h in enumerate(headers):
                    if "otp" in h:
                        otp_idx = idx
                        break
                
                if otp_idx == -1:
                    otp_idx = 1 if len(rows[0]) > 1 else 0
                    
                last_row = rows[-1]
                if len(last_row) > otp_idx:
                    return last_row[otp_idx].strip()
                return ""
        except Exception as e:
            print(f"⚠️ Gagal membaca OTP dari Sheets: {e}")
            return ""

    parsed = urlparse(url_dasar)
    query_params = dict(parse_qsl(parsed.query))
    query_params["action"] = action
    if label_email:
        query_params["label"] = label_email
    url_final = urlunparse(parsed._replace(query=urlencode(query_params)))

    try:
        with urlopen(url_final, timeout=30) as response:
            return response.read().decode("utf-8").strip()
    except Exception as e:
        print(f"   ⚠️ Gagal membaca OTP dari endpoint: {e}")
        return ""


def tunggu_otp_terbaru(url_dasar, action="getOtpEmail", label_email=None, interval_detik=3, otp_awal_override=None, timeout_detik=15, page=None):
    """
    Menunggu OTP terbaru yang berbeda dari nilai awal agar tidak memakai OTP sebelumnya.
    Jika parameter page diberikan, memantau halaman web secara live dan langsung membatalkan
    tunggu jika terdeteksi banner/pesan error pemblokiran.
    """
    if otp_awal_override is not None:
        otp_awal = otp_awal_override
    else:
        try:
            otp_awal = ambil_otp_dari_endpoint(url_dasar, action=action, label_email=label_email)
        except Exception:
            otp_awal = ""
    
    batas_waktu = time.time() + timeout_detik
    print(f"   🤖 Menunggu OTP baru masuk ke inbox label [{label_email}] (maksimal {timeout_detik} detik)...")

    while time.time() < batas_waktu:
        if page and not page.is_closed():
            try:
                is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                if is_blocked:
                    print(f"   🚫 [TERBLOKIR / LIMIT] Terdeteksi pesan di halaman web: '{block_msg}'. Membatalkan penantian OTP.")
                    return ""
            except Exception:
                pass

        time.sleep(interval_detik)
        try:
            otp_baru = ambil_otp_dari_endpoint(url_dasar, action=action, label_email=label_email)
            if otp_baru and otp_baru != otp_awal:
                return otp_baru
        except Exception:
            pass

    return ""
    

def get_gmail_otp(email=None, req_time=None, max_wait_sec=90):
    """Alias helper untuk menarik OTP dari endpoint Google Apps Script."""
    return tunggu_otp_terbaru(OTP_ENDPOINT_URL, action="getOtpEmail", label_email=GMAIL_OTP_LABEL, timeout_detik=max_wait_sec)


def trigger_submit_otp(page):
    """Memicu pengiriman form OTP (Enter key & klik tombol Verifikasi/Masuk)."""
    if not page or page.is_closed():
        return
    time.sleep(0.5)
    # 1. Coba Enter via keyboard
    try:
        page.keyboard.press("Enter")
    except Exception:
        pass
    time.sleep(0.5)

    # 2. Cari dan klik tombol submit / verifikasi jika ada
    submit_selectors = [
        "button:has-text('Verifikasi')",
        "button:has-text('Konfirmasi')",
        "button:has-text('Lanjutkan')",
        "button:has-text('Lanjut')",
        "button:has-text('Masuk')",
        "button:has-text('Verify')",
        "button:has-text('Submit')",
        "button[type='submit']"
    ]
    for sel in submit_selectors:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0 and btn.is_visible():
                print(f"   👉 Mengklik tombol submit OTP: '{btn.inner_text().strip()}'")
                btn.click(timeout=3000)
                time.sleep(1)
                break
        except Exception:
            pass


def isi_kode_otp(page, otp_code):
    """
    Memasukkan kode OTP 4-6 digit dengan multi-metode yang andal
    dan langsung memicu submit form.
    """
    otp_code = str(otp_code).strip()
    if not otp_code or not page or page.is_closed():
        return False

    print(f"   🤖 Memasukkan kode OTP ({len(otp_code)} digit): {otp_code}...")

    # Cari semua input form yang relevan untuk OTP
    selectors = [
        'input[autocomplete="one-time-code"]',
        'input[aria-label*="digit" i]',
        'input[aria-label*="kode" i]',
        'input[aria-label*="otp" i]',
        'div[class*="otp" i] input',
        'div[class*="pin" i] input',
        'div[class*="code" i] input',
        'input[maxlength="1"]',
        'input[type="tel"]',
        'input[type="number"]',
        'input[type="text"]'
    ]

    all_inputs = []
    for sel in selectors:
        try:
            matched = page.locator(sel).all()
            visible = [inp for inp in matched if inp.is_visible() and inp.is_enabled()]
            if len(visible) >= len(otp_code):
                all_inputs = visible
                break
            elif len(visible) > 0 and not all_inputs:
                all_inputs = visible
        except Exception:
            pass

    success = False

    # METODE 1: Jika ada kotak digit terpisah (4 atau 6 kotak)
    if len(all_inputs) >= len(otp_code):
        try:
            print(f"   🤖 Mengisi ke {len(all_inputs)} kotak digit secara berurutan...")
            for idx, char in enumerate(otp_code):
                inp = all_inputs[idx]
                inp.click(timeout=2000)
                time.sleep(0.05)
                inp.fill(char)
                time.sleep(0.05)
                page.keyboard.press(char)
                time.sleep(0.05)
            print("   ✅ Berhasil mengisi seluruh kotak digit OTP.")
            success = True
        except Exception as e:
            print(f"   ⚠️ Percobaan metode multi-box gagal: {e}")

    # METODE 2: Klik input pertama dan gunakan keyboard sequential typing
    if not success and all_inputs:
        try:
            print("   🤖 Mengisi via fokus input pertama & keyboard typing...")
            first_input = all_inputs[0]
            first_input.click(timeout=2000)
            time.sleep(0.1)
            first_input.focus()
            time.sleep(0.1)
            
            # Jika single input box, coba fill langsung juga
            if len(all_inputs) == 1:
                try:
                    first_input.fill(otp_code)
                    time.sleep(0.2)
                except Exception:
                    pass
            
            # Ketik per digit via page.keyboard
            for char in otp_code:
                page.keyboard.press(char)
                time.sleep(0.1)
            print("   ✅ Berhasil mengetik OTP via keyboard global.")
            success = True
        except Exception as e:
            print(f"   ⚠️ Percobaan keyboard typing gagal: {e}")

    # METODE 3: Javascript direct DOM injection
    if not success:
        try:
            print("   🤖 Mengisi via Javascript DOM event injection...")
            page.evaluate("""(code) => {
                const inputs = Array.from(document.querySelectorAll('input:not([type="hidden"]):not([type="checkbox"]):not([type="radio"])')).filter(el => el.offsetParent !== null);
                if (inputs.length >= code.length) {
                    for (let i = 0; i < code.length; i++) {
                        inputs[i].focus();
                        inputs[i].value = code[i];
                        inputs[i].dispatchEvent(new Event('input', { bubbles: true }));
                        inputs[i].dispatchEvent(new Event('change', { bubbles: true }));
                    }
                } else if (inputs.length > 0) {
                    inputs[0].focus();
                    inputs[0].value = code;
                    inputs[0].dispatchEvent(new Event('input', { bubbles: true }));
                    inputs[0].dispatchEvent(new Event('change', { bubbles: true }));
                }
            }""", otp_code)
            print("   ✅ Injeksi Javascript OTP selesai.")
            success = True
        except Exception as e:
            print(f"   ⚠️ Gagal injeksi Javascript OTP: {e}")

    # Memicu submit tombol / Enter setelah OTP dimasukkan
    trigger_submit_otp(page)
    return success


def deteksi_pesan_blokir_atau_error(page):
    """
    Mendeteksi secara instan apakah halaman login GoBiz menampilkan notifikasi error,
    rate limit, akun terblokir, atau limit OTP, sehingga tidak membuang waktu menunggu 90 detik.
    Mengembalikan (is_blocked: bool, pesan_error: str).
    """
    try:
        error_selectors = [
            'div[role="alert"]',
            'div[class*="alert" i]',
            'div[class*="error" i]',
            'div[class*="toast" i]',
            'div[class*="notification" i]',
            'div[class*="snackbar" i]',
            'div[class*="banner" i]',
            'p[class*="error" i]',
            'span[class*="error" i]',
            'div[class*="helper-text" i]',
            'div[class*="message" i]'
        ]

        keywords_blokir = [
            "terlalu banyak", "too many", "terblokir", "diblokir", "blocked",
            "dibatasi", "limit", "coba lagi nanti", "coba lagi dalam", "try again",
            "tidak terdaftar", "belum terdaftar", "tidak valid", "invalid",
            "mencapai batas", "kuota habis", "terjadi gangguan", "kesalahan"
        ]

        for sel in error_selectors:
            try:
                elements = page.locator(sel).all()
                for el in elements:
                    if el.is_visible():
                        txt = el.inner_text().strip()
                        if txt and any(kw in txt.lower() for kw in keywords_blokir):
                            return True, txt
            except Exception:
                pass

        try:
            body_text = page.locator("body").inner_text().lower()
            for kw in [
                "diblok sementara", "terlalu banyak kesalahan", "coba lagi dalam",
                "terlalu banyak percobaan", "akun anda terblokir", "akun diblokir", 
                "terlalu banyak permintaan", "too many attempts", "temporarily blocked",
                "coba lagi dalam beberapa saat", "silakan coba lagi nanti"
            ]:
                if kw in body_text:
                    return True, kw
        except Exception:
            pass

    except Exception:
        pass
        
    return False, ""


def tutup_semua_popup(page):
    """
    Dismisses common popups, cookie consent banners, and onboarding overlays
    """
    # 1. Cookie consent & Help overlay
    cookie_selectors = [
        'text="Terima Semua Cookie"',
        'text="Accept All Cookies"',
        'button:has-text("Terima Semua Cookie")',
        'button:has-text("Terima Semua")',
        'button:has-text("Terima")',
        'button:has-text("Accept")',
        'button#onetrust-accept-btn-handler',
        'button:has-text("Lewati")',
        'button:has-text("Lewati Tutorial")',
        'button:has-text("Selesai")',
        'button:has-text("Nanti Saja")',
        'button:has-text("Tutup")',
        '[aria-label="close"]',
        '[aria-label="Close"]',
        'button:has-text("×")',
        'button:has-text("✕")',
        'div[class*="help"] button',
        'div[class*="bantuan"] button'
    ]
    for sel in cookie_selectors:
        try:
            loc = page.locator(sel)
            for i in range(min(loc.count(), 3)):
                candidate = loc.nth(i)
                if candidate.is_visible():
                    candidate.click(timeout=1000, force=True)
                    time.sleep(0.2)
        except Exception:
            pass

    # 2. Fallback Eksekusi Javascript
    js_click_script = """
    () => {
        const texts = ["terima semua cookie", "accept all cookies", "terima semua", "terima", "accept", "lewati", "lewati tutorial", "selesai", "tutup", "nanti saja", "×", "✕"];
        const elements = document.querySelectorAll('button, a, span, div, p');
        for (const el of elements) {
            const text = (el.innerText || el.textContent || "").toLowerCase().trim();
            if (texts.includes(text) && el.offsetWidth > 0 && el.offsetHeight > 0) {
                try { el.click(); } catch(e) {}
            }
        }
        const widgets = document.querySelectorAll('[class*="help"], [class*="bantuan"], [id*="help"]');
        for (const w of widgets) {
            if (w.innerText && w.innerText.includes("bantuan")) {
                w.style.display = "none";
            }
        }
    }
    """
    try:
        frames = [page] + page.frames
        for frame in frames:
            try:
                frame.evaluate(js_click_script)
            except Exception:
                pass
    except Exception:
        pass


GOOGLE_SHEET_AGENCY_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQ3tLKBNXDqRgBw0mNhKZFxgvKx-JoiTDzm_s5Ix1cm7O6HCv4IvExOLR2HSRVaXSsx82V348mcr9X4/pub?output=csv"
GOOGLE_SHEET_VB_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRYSUnKOqk29LCktTxdb0wPLbWMbRaWRP3eC_UA4AwYod1FW6zDMhtLMC5ghIvot2B8upCDfBsn-TCP/pub?gid=369960309&single=true&output=csv"
GOOGLE_SHEET_VERCEL_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTprbPPf_J5gAVL3PYeHbbdl5ZXQvb17HY2lJGPI2xg13Ly3AGT8eYHLYmU_m1NdtkBVg-qUGv1BoEE/pub?output=csv"


def get_template_headers():
    """Mengambil header dari template YYYY-MM-DD HH_MM Nama Pemilik.xlsx."""
    import openpyxl
    template_path = BASE_DIR / "YYYY-MM-DD HH_MM Nama Pemilik.xlsx"
    if template_path.exists():
        try:
            wb = openpyxl.load_workbook(str(template_path), read_only=True)
            ws = wb['Listing'] if 'Listing' in wb.sheetnames else wb.active
            return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        except Exception:
            pass
    return ['Nama Pemilik', 'Nama Brand', 'Model', 'Tipe', 'Outlet', 'Nomor HP', 'Aplikator', 'Nama Portal', 'Group ID', 'Nama Listing', 'Link', 'Store ID', 'Status Listing', 'Alamat', 'Nama Bank', 'Nama Pemilik Rekening', 'Nomor Rekening']


def save_formatted_excel(df, file_path):
    """Menyimpan DataFrame ke file Excel menggunakan template YYYY-MM-DD HH_MM Nama Pemilik.xlsx."""
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Font, Alignment
    
    headers = get_template_headers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listing"
    
    header_font = Font(name='Calibri', size=11, bold=True)
    for c_idx, h in enumerate(headers, 1):
        if h is not None:
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = header_font
            
    font = Font(name='Calibri', size=11)
    alignment = Alignment(horizontal='left', vertical='center')
    
    for idx, (_, r) in enumerate(df.iterrows(), start=2):
        ws.cell(row=idx, column=1, value=str(r.get('Nama Pemilik') or '') if pd.notna(r.get('Nama Pemilik')) else '') # A: Nama Pemilik
        ws.cell(row=idx, column=2, value=str(r.get('Nama Brand') or '') if pd.notna(r.get('Nama Brand')) else '')   # B: Nama Brand
        ws.cell(row=idx, column=7, value=str(r.get('Aplikator') or '') if pd.notna(r.get('Aplikator')) else 'GoFood') # G: Aplikator
        ws.cell(row=idx, column=8, value=str(r.get('Nama Portal') or '') if pd.notna(r.get('Nama Portal')) else '') # H: Nama Portal
        ws.cell(row=idx, column=9, value=str(r.get('Group ID') or '') if pd.notna(r.get('Group ID')) else '')       # I: Group ID
        ws.cell(row=idx, column=10, value=str(r.get('Nama Listing') or '') if pd.notna(r.get('Nama Listing')) else '') # J: Nama Listing
        ws.cell(row=idx, column=11, value=str(r.get('Link') or '') if pd.notna(r.get('Link')) else '')             # K: Link
        ws.cell(row=idx, column=12, value=str(r.get('Store ID') or '') if pd.notna(r.get('Store ID')) else '')     # L: Store ID
        ws.cell(row=idx, column=13, value=str(r.get('Status Listing') or '') if pd.notna(r.get('Status Listing')) else '') # M: Status Listing
        ws.cell(row=idx, column=14, value=str(r.get('Alamat') or '') if pd.notna(r.get('Alamat')) else '')         # N: Alamat
        ws.cell(row=idx, column=15, value=str(r.get('Nama Bank') or '') if pd.notna(r.get('Nama Bank')) else '')   # O: Nama Bank
        ws.cell(row=idx, column=16, value=str(r.get('Nama Pemilik Rekening') or '') if pd.notna(r.get('Nama Pemilik Rekening')) else '') # P: Nama Pemilik Rekening
        
        c_rek = ws.cell(row=idx, column=17, value=str(r.get('Nomor Rekening') or '') if pd.notna(r.get('Nomor Rekening')) else '') # Q: Nomor Rekening
        c_rek.number_format = '@'

        for c in range(1, 18):
            cell = ws.cell(row=idx, column=c)
            cell.font = font
            cell.alignment = alignment

    # Auto adjust column widths
    for col_idx in range(1, 18):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max((len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, len(df) + 2)), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    wb.save(file_path)


def get_safe_cache_filename(portal_name):
    """Menghasilkan nama file cache yang aman dari batasan panjang sistem operasi (maks 255 karakter)."""
    import hashlib
    clean = re.sub(r'[^a-zA-Z0-9_.-]', '_', str(portal_name).strip())
    if len(clean) > 80:
        h = hashlib.md5(str(portal_name).encode('utf-8')).hexdigest()[:8]
        clean = clean[:70].rstrip('_') + f"_{h}"
    return f"gofood_portal_{clean}.json"


def get_credentials_from_sheet(source_type="agency", custom_url=None):
    """Mengambil daftar kredensial portal GoFood berdasarkan tipe sumber (Agency, VB, atau Vercel)."""
    import subprocess
    data = []
    target_url = custom_url
    st_lower = str(source_type).lower()
    
    if not target_url:
        if st_lower == "vb":
            target_url = GOOGLE_SHEET_VB_URL or os.getenv("GOFOOD_VB_CSV_URL", "")
        elif st_lower == "vercel":
            target_url = GOOGLE_SHEET_VERCEL_URL
        else:
            target_url = GOOGLE_SHEET_AGENCY_URL

    # 1. Fetch live dari Google Sheet URL
    if target_url:
        try:
            print(f"[*] Mengambil data portal GoFood [{source_type.upper()}] langsung dari Google Sheet (Live URL)...")
            res = subprocess.run(['curl', '-s', '-L', target_url], capture_output=True, text=True, timeout=20)
            if res.returncode == 0 and res.stdout.strip():
                reader = csv.reader(res.stdout.splitlines())
                data = list(reader)
                print(f"[✓] Berhasil memuat {len(data)} baris data [{source_type.upper()}] dari Google Sheet.")
        except Exception as e:
            print(f"⚠️ Gagal fetch live dari Google Sheet [{source_type.upper()}]: {e}.")

        if not data:
            try:
                req = urllib.request.Request(target_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=20) as response:
                    content = response.read().decode('utf-8')
                    reader = csv.reader(content.splitlines())
                    data = list(reader)
                    print(f"[✓] Berhasil memuat {len(data)} baris data [{source_type.upper()}] dari Google Sheet (fallback).")
            except Exception as e:
                print(f"⚠️ Gagal fetch fallback dari Google Sheet [{source_type.upper()}]: {e}.")

    # 2. Fallback CSV lokal untuk Agency
    if not data and st_lower == "agency":
        csv_file = BASE_DIR / "A. Credential (Outlet & Access)  - Unique Portal Go.csv"
        if os.path.exists(csv_file):
            try:
                print(f"[*] Membaca data portal dari fallback file CSV: {os.path.basename(csv_file)}")
                with open(csv_file, "r", encoding="utf-8") as f:
                    data = list(csv.reader(f))
            except Exception as e:
                print(f"⚠️ Gagal membaca fallback CSV lokal: {e}")

    if not data:
        print(f"❌ Tidak ada data portal yang dapat dimuat untuk tipe [{source_type.upper()}].")
        return []

    header = [str(h).strip().lower() for h in data[0]]
    col_app = -1
    col_owner = -1
    col_portal = -1
    col_email1 = -1
    col_email2 = -1
    col_pass = -1
    col_notes = -1

    if st_lower == "agency":
        col_app = 3
        col_owner = 0
        col_portal = 2
        col_email1 = 24
        col_email2 = 25
        col_pass = 28
    else:
        for i, h in enumerate(header):
            if h in ['aplikasi', 'app', 'platform']:
                col_app = i
            elif h in ['owner', 'nama pemilik', 'pemilik']:
                col_owner = i
            elif h in ['nama akses', 'nama portal', 'portal', 'nama outlet', 'brand']:
                if col_portal == -1 or h in ['nama portal', 'portal', 'nama akses']:
                    col_portal = i
            elif h in ['email', 'email foodmaster1', 'email login', 'email 1', 'email1']:
                col_email1 = i
            elif h in ['email foodmaster2', 'email 2', 'email2']:
                col_email2 = i
            elif 'email' in h:
                if col_email1 == -1:
                    col_email1 = i
                elif col_email2 == -1:
                    col_email2 = i
            elif h in ['password', 'kata sandi', 'pass login', 'pass']:
                col_pass = i
            elif any(k in h for k in ['notes', 'catatan', 'keterangan']):
                col_notes = i

    unique_portals = {}
    for row in data[1:]:
        # Filter Restricted
        if col_notes != -1 and col_notes < len(row):
            note_val = str(row[col_notes]).strip().lower()
            if "restricted" in note_val:
                p_name = row[col_portal].strip() if col_portal != -1 and col_portal < len(row) else "Unknown"
                print(f"🚫 Melewati portal '{p_name}' karena berstatus Restricted.")
                continue

        is_gofood = True
        if col_app != -1 and col_app < len(row) and row[col_app].strip():
            is_gofood = "gofood" in row[col_app].strip().lower()

        if is_gofood:
            owner = row[col_owner].strip() if col_owner != -1 and col_owner < len(row) else ("VB" if source_type == "vb" else "")
            portal = row[col_portal].strip() if col_portal != -1 and col_portal < len(row) else ""
            if not portal:
                if len(row) > 1 and row[1].strip():
                    portal = row[1].strip()
                elif len(row) > 0 and row[0].strip():
                    portal = f"Portal {row[0].strip()}"
                else:
                    portal = "Portal Unknown"

            brand = portal.split(" - ")[0].strip() if " - " in portal else portal
            email1 = row[col_email1].strip() if col_email1 != -1 and col_email1 < len(row) else ""
            email2 = row[col_email2].strip() if col_email2 != -1 and col_email2 < len(row) else ""
            password = row[col_pass].strip() if col_pass != -1 and col_pass < len(row) else ""
            if password == "-":
                password = ""
            
            emails = []
            if email1 and email1 != "-":
                emails.append(email1)
            if email2 and email2 != "-" and email2 != email1:
                emails.append(email2)
                
            if emails:
                primary_email = emails[0].strip().lower()
                owner_val = owner or ('VB' if source_type == 'vb' else brand)
                dedup_key = (owner_val.strip().lower(), primary_email)
                
                if dedup_key not in unique_portals:
                    safe_cache_name = get_safe_cache_filename(portal)
                    cache_file = CACHE_DIR / safe_cache_name
                    unique_portals[dedup_key] = {
                        'owner': owner_val,
                        'brand': brand,
                        'portal': portal,
                        'email': emails[0],
                        'emails': emails,
                        'password': password,
                        'source_type': source_type.upper(),
                        'cache_file': str(cache_file),
                        'output': str(cache_file)
                    }
                else:
                    # Gabungkan nama portal / brand jika terdapat perbedaan
                    if portal not in unique_portals[dedup_key]['portal']:
                        unique_portals[dedup_key]['portal'] += f", {portal}"
                    for em in emails:
                        if em not in unique_portals[dedup_key]['emails']:
                            unique_portals[dedup_key]['emails'].append(em)
                    if password and not unique_portals[dedup_key]['password']:
                        unique_portals[dedup_key]['password'] = password

    portals = list(unique_portals.values())
    print(f"[✓] Terdaftar {len(portals)} akun login GoFood unik [{source_type.upper()}] untuk diproses.")
    return portals


def parse_selection(choice, max_val):
    selected = set()
    choice = choice.strip()
    if choice.lower() in ('0', 'all'):
        return list(range(max_val))
        
    parts = choice.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            try:
                start, end = map(int, part.split('-'))
                if 1 <= start <= end <= max_val:
                    for i in range(start, end + 1):
                        selected.add(i - 1)
            except ValueError:
                pass
        else:
            try:
                val = int(part)
                if 1 <= val <= max_val:
                    selected.add(val - 1)
            except ValueError:
                pass
    return sorted(list(selected))


def upload_to_drive(file_path):
    import base64
    if not APP_SCRIPT_URL:
        return
    print(f"\n[*] Mengunggah {os.path.basename(file_path)} ke Google Drive...")
    try:
        with open(file_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("utf-8")
            
        payload = {
            "fileName": os.path.basename(file_path),
            "fileBase64": encoded,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        response = requests.post(APP_SCRIPT_URL, json=payload, allow_redirects=True)
        if response.status_code in (200, 201):
            print("   ✅ Berhasil diunggah ke Google Drive!")
    except Exception as e:
        print(f"   ⚠️ Terjadi kesalahan saat mengunggah: {e}")


def combine_master(cache_dir=None, master_dir=None, output_dir=None, source_type=None):
    import pandas as pd
    import glob
    import json

    cache_dir = Path(cache_dir) if cache_dir else CACHE_DIR
    master_dir = Path(master_dir) if master_dir else MASTER_DIR
    output_dir = Path(output_dir) if output_dir else OUTPUT_DIR

    master_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n[*] Menggabungkan semua data cache JSON ke Master...")
    json_files = sorted(glob.glob(str(cache_dir / "gofood_portal_*.json")))
    
    all_records = []
    if json_files:
        for f in json_files:
            try:
                with open(f, 'r', encoding='utf-8') as jf:
                    cdata = json.load(jf)
                    outlets = cdata.get('outlets', [])
                    if isinstance(outlets, list):
                        all_records.extend(outlets)
            except Exception as e:
                print(f"   ⚠️ Gagal membaca cache {f}: {e}")
    else:
        # Fallback baca dari file xlsx lama jika ada
        xlsx_files = sorted(glob.glob(str(DATA_DIR / "GOFOOD_outlets_*.xlsx")))
        if not xlsx_files:
            print("   ⚠️ Tidak ada file cache (gofood_portal_*.json) di folder cache untuk digabung.")
            return
        for f in xlsx_files:
            try:
                df = pd.read_excel(f, sheet_name="Listing" if "Listing" in pd.ExcelFile(f).sheet_names else 0)
                df.dropna(how="all", inplace=True)
                all_records.extend(df.to_dict(orient='records'))
            except Exception as e:
                print(f"   ⚠️ Gagal membaca {f}: {e}")
            
    if all_records:
        master_df = pd.DataFrame(all_records)
            
        # 1. Filter berdasarkan source_type TERLEBIH DAHULU agar data lintas tipe tidak saling memotong
        if source_type:
            st = str(source_type).strip().lower()
            if 'Nama Pemilik' in master_df.columns:
                if st == 'agency':
                    master_df = master_df[master_df['Nama Pemilik'] != 'VB']
                elif st == 'vb':
                    master_df = master_df[master_df['Nama Pemilik'] == 'VB']
                elif st == 'vercel':
                    master_df = master_df[master_df['Nama Pemilik'] != 'VB']

        # 2. Deduplikasi Store ID dalam lingkup sumber yang dipilih
        if "Store ID" in master_df.columns:
            master_df.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)

        if master_df.empty:
            print("   ⚠️ Tidak ada baris data setelah pemfilteran source_type.")
            return

        st = str(source_type).strip().lower() if source_type else "agency"
        if st == "vb":
            tag = "VB"
        elif st == "vercel":
            tag = "Vercel"
        else:
            tag = "Agency"

        timestamp_name = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
        master_filename = f"{timestamp_name} {tag}_master.xlsx"
        master_path = master_dir / master_filename
        
        # File versioning jika file dengan nama sama sudah ada
        if master_path.exists():
            version_dir = master_dir / "versions"
            version_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = version_dir / f"{master_path.stem}_v{timestamp}.xlsx"
            shutil.copy2(str(master_path), str(backup_path))
            print(f"   💾 Master lama di-backup ke: {backup_path.name}")
            
        # Simpan master bertimestamp (YYYY-MM-DD HH_MM [Tipe]_master.xlsx)
        save_formatted_excel(master_df, str(master_path))

        # Simpan juga salinan statis [Tipe]_master.xlsx dan 0master.xlsx (khusus Agency) untuk kompatibilitas
        static_master_path = master_dir / f"{tag}_master.xlsx"
        save_formatted_excel(master_df, str(static_master_path))
        if tag == "Agency":
            save_formatted_excel(master_df, str(master_dir / "0master.xlsx"))
        
        timestamp_name = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
        
        # Simpan file per Nama Pemilik ke output/
        if 'Nama Pemilik' in master_df.columns:
            owners = master_df['Nama Pemilik'].dropna().unique()
            for owner in owners:
                if owner and str(owner).strip():
                    owner_clean = "".join(c for c in str(owner) if c.isalnum() or c in " ._-").strip()
                    owner_file = output_dir / f"{timestamp_name} {owner_clean}.xlsx"
                    owner_df = master_df[master_df['Nama Pemilik'] == owner]
                    save_formatted_excel(owner_df, str(owner_file))
                    print(f"   💾 File per-Owner tersimpan di output/: {owner_file.name}")
        
        print("\n[✓] Master file berhasil dibuat:")
        print(f"    - {master_path}")
        print(f"    Total baris: {len(master_df)}")
        print(f"    Total kolom: 37 (Template YYYY-MM-DD HH_MM Nama Pemilik.xlsx)")
        
        if APP_SCRIPT_URL:
            upload_to_drive(str(master_path))


def extract_gofood_rest_id(src, store_id):
    """
    Ekstraksi rest_id (GoFood Restaurant UUID / ID) dari data GoBiz.
    Memastikan rest_id BUKAN store_id (seperti G025124092).
    """
    apps = src.get('applications', {}) if isinstance(src.get('applications'), dict) else {}
    goresto = apps.get('goresto', {}) if isinstance(apps.get('goresto'), dict) else {}
    gofood = apps.get('gofood', {}) if isinstance(apps.get('gofood'), dict) else {}
    external_ids = src.get('external_ids', {}) if isinstance(src.get('external_ids'), dict) else {}
    metadata = src.get('metadata', {}) if isinstance(src.get('metadata'), dict) else {}
    
    candidates = [
        goresto.get('goresto_id'),
        goresto.get('restaurant_id'),
        goresto.get('external_id'),
        gofood.get('restaurant_id'),
        gofood.get('external_id'),
        (external_ids.get('restaurant', [None])[0] if isinstance(external_ids.get('restaurant'), list) and external_ids.get('restaurant') else None),
        metadata.get('restaurant_id'),
        metadata.get('gofood_restaurant_id'),
        metadata.get('restaurant_uuid'),
        src.get('restaurant_id'),
        src.get('resto_id'),
        goresto.get('id'),
        gofood.get('id'),
    ]
    
    for cand in candidates:
        if cand and isinstance(cand, str):
            c_clean = cand.strip()
            # Validasi: rest_id valid tidak boleh kosong dan tidak boleh sama dengan store_id
            if c_clean and c_clean != store_id:
                # Pastikan bukan ID merchant GoBiz berawalan 'G' + digit (contoh: G025124092)
                if not (c_clean.startswith('G') and len(c_clean) >= 8 and c_clean[1:].isdigit()):
                    return c_clean
            # Jika berupa UUID (mengandung tanda '-')
            if c_clean and '-' in c_clean and len(c_clean) >= 32:
                return c_clean

    return ""


def fetch_gobiz_merchants_fast(access_token, cookies=None, page_size=1000):
    """
    Menarik data outlet GoBiz langsung via REST API menggunakan curl_cffi
    dengan impersonate='chrome120' untuk proteksi TLS/WAF dan kecepatan maksimal (tanpa browser).
    """
    if not access_token:
        return False, 401, [], "No access token provided"

    is_curl_cffi = False
    try:
        from curl_cffi import requests as cffi_requests
        is_curl_cffi = True
    except ImportError:
        import requests as cffi_requests

    if not str(access_token).startswith("Bearer "):
        token_header = f"Bearer {access_token}"
    else:
        token_header = str(access_token)

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Authentication-Type': 'go-id',
        'Authorization': token_header,
        'Content-Type': 'application/json',
        'Origin': 'https://portal.gofoodmerchant.co.id',
        'Referer': 'https://portal.gofoodmerchant.co.id/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }

    cookie_dict = {}
    if cookies:
        if isinstance(cookies, list):
            for c in cookies:
                if isinstance(c, dict) and 'name' in c and 'value' in c:
                    cookie_dict[c['name']] = c['value']
        elif isinstance(cookies, dict):
            cookie_dict = cookies

    all_hits = []
    from_offset = 0
    max_retries = 3

    while True:
        payload = {
            "from": from_offset,
            "size": page_size
        }
        
        success = False
        last_err = ""
        status_code = 0

        for attempt in range(1, max_retries + 1):
            try:
                post_kwargs = {
                    'headers': headers,
                    'cookies': cookie_dict,
                    'json': payload,
                    'timeout': 20
                }
                if is_curl_cffi:
                    post_kwargs['impersonate'] = "chrome120"

                if hasattr(cffi_requests, 'Session'):
                    sess = cffi_requests.Session()
                    resp = sess.post(
                        'https://api.gobiz.co.id/v1/merchants/search',
                        **post_kwargs
                    )
                else:
                    resp = cffi_requests.post(
                        'https://api.gobiz.co.id/v1/merchants/search',
                        **post_kwargs
                    )
                
                status_code = resp.status_code
                if status_code == 200:
                    data = resp.json()
                    hits = []
                    if isinstance(data, dict):
                        if 'hits' in data and isinstance(data['hits'], list):
                            hits = data['hits']
                        elif 'data' in data and isinstance(data['data'], list):
                            hits = data['data']
                    
                    all_hits.extend(hits)
                    success = True
                    
                    # Cek apakah masih ada data di halaman berikutnya
                    total_count = data.get('total', len(hits)) if isinstance(data, dict) else len(hits)
                    if len(hits) < page_size or len(all_hits) >= total_count:
                        return True, 200, all_hits, ""
                    else:
                        from_offset += page_size
                        break
                        
                elif status_code == 401:
                    return False, 401, [], "Session expired / Unauthorized (401)"
                elif status_code == 403:
                    return False, 403, [], "Forbidden / WAF Blocked (403)"
                else:
                    last_err = f"HTTP {status_code}: {resp.text[:200]}"
                    time.sleep(1)
            except Exception as e:
                last_err = str(e)
                time.sleep(1)

        if not success:
            if all_hits:
                return True, 200, all_hits, f"Partial fetch with error: {last_err}"
            return False, status_code or 500, [], last_err


def transform_gobiz_hits(hits, owner_name, portal_name, brand_name):
    """Mengubah hits dari GoBiz API search menjadi format standar listing outlet 17 kolom."""
    outlets_data = []
    for item in hits:
        src = item.get('_source', item) if isinstance(item, dict) else {}
        nama = src.get('outlet_name') or src.get('merchant_name', 'Unknown')
        store_id = src.get('id', '')
        
        status = "Unknown"
        apps = src.get('applications', {}) if isinstance(src.get('applications'), dict) else {}
        if 'goresto' in apps and isinstance(apps['goresto'], dict):
            status = apps['goresto'].get('status', status)
            
        alamat = src.get('outlet_address', '')
        
        # Ekstraksi Group ID
        tags = src.get('tags', {}) if isinstance(src.get('tags'), dict) else {}
        external_ids = src.get('external_ids', {}) if isinstance(src.get('external_ids'), dict) else {}
        group_id = ""
        for container in [tags, external_ids]:
            if isinstance(container, dict):
                entity_list = container.get('entity', [])
                brand_list = container.get('brand', [])
                if isinstance(entity_list, list) and entity_list:
                    group_id = str(entity_list[0])
                    break
                elif isinstance(brand_list, list) and brand_list:
                    group_id = str(brand_list[0])
                    break
        if not group_id:
            group_id = str(src.get('partner_id', '') or '')
        
        rest_id = extract_gofood_rest_id(src, store_id)
        gofood_link = f"http://gofood.co.id/surabaya/restaurant/{rest_id}" if rest_id else ""
        
        bank_no = ""
        bank_name = ""
        acc_name = ""
        if 'bank_account' in src and isinstance(src['bank_account'], dict):
            bank_no = str(src['bank_account'].get('account_no', '') or '')
            bank_name = src['bank_account'].get('bank_name', '')
            acc_name = src['bank_account'].get('account_name', '')
                
        outlets_data.append({
            'Nama Pemilik': owner_name,
            'Nama Brand': brand_name or owner_name,
            'Aplikator': 'GoFood',
            'Nama Portal': portal_name,
            'Group ID': group_id,
            'Nama Listing': nama,
            'Link': gofood_link,
            'Store ID': store_id,
            'Status Listing': status,
            'Alamat': alamat,
            'Nama Bank': bank_name,
            'Nama Pemilik Rekening': acc_name,
            'Nomor Rekening': str(bank_no) if bank_no else ""
        })
    return outlets_data


def save_portal_results(outlets_data, owner_name, portal_name, brand_name, email, source_type):
    """Menyimpan data portal ke cache JSON dan output Excel per-owner."""
    if not outlets_data:
        print(f"   ⚠️ Tidak ada data outlet untuk disimpan pada portal '{portal_name}'.")
        return None, None
        
    safe_cache_name = get_safe_cache_filename(portal_name)
    safe_owner = re.sub(r'[^a-zA-Z0-9_.-]', '_', owner_name.strip())
    
    # 1. Simpan ke Cache JSON
    cache_json_file = CACHE_DIR / safe_cache_name
    cache_payload = {
        'portal': portal_name,
        'brand': brand_name or owner_name,
        'owner': owner_name,
        'source_type': str(source_type).upper(),
        'email': email,
        'timestamp': datetime.datetime.now().isoformat(),
        'outlets': outlets_data
    }
    try:
        with open(cache_json_file, 'w', encoding='utf-8') as jf:
            json.dump(cache_payload, jf, indent=2, ensure_ascii=False)
        print(f"   💾 Cache portal disimpan di cache/: {cache_json_file.name}")
    except Exception as e:
        print(f"   ⚠️ Gagal menyimpan cache JSON: {e}")
    
    # 2. Simpan ke Output Excel
    import pandas as pd
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
    owner_df = pd.DataFrame(outlets_data)
    if 'Store ID' in owner_df.columns:
        owner_df.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)
        
    owner_file = OUTPUT_DIR / f"{timestamp_str} {safe_owner}.xlsx"
    save_formatted_excel(owner_df, str(owner_file))
    print(f"   💾 File Owner '{owner_name}' tersimpan di output/: {owner_file.name} (Total: {len(owner_df)} outlet)")
    
    if APP_SCRIPT_URL:
        upload_to_drive(str(owner_file))
        
    return cache_json_file, owner_file


def run_manual_mode():
    """Mode manual: Membuka browser non-headless agar user login sendiri, lalu scrape via curl_cffi saat di dashboard."""
    print("\n" + "=" * 60)
    print("  🌐 GOFOOD SCRAPER — MANUAL LOGIN MODE")
    print("=" * 60)
    print("  1. Skrip akan membuka jendela browser Playwright.")
    print("  2. Silakan login (Email / Password / OTP) secara manual di browser.")
    print("  3. Setelah sampai di halaman Dashboard (/dashboard atau /home),")
    print("     kembali ke terminal dan tekan [ENTER] untuk melanjutkan ekstraksi data.")
    print("=" * 60)
    
    owner_input = input("\nMasukkan Nama Pemilik / Brand (Default: Manual): ").strip()
    owner_name = owner_input if owner_input else "Manual"
    portal_name = f"Manual - {owner_name}"
    
    with sync_playwright() as p:
        print("\n[*] Membuka jendela browser Chrome...")
        browser = p.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-infobars',
                '--no-sandbox',
            ]
        )
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1366, 'height': 768}
        )
        page = context.new_page()
        page.goto("https://portal.gofoodmerchant.co.id/auth/login/email", wait_until="domcontentloaded")
        
        print("\n" + "─" * 60)
        print("👉 Silakan lakukan login di jendela browser yang terbuka.")
        print("👉 Setelah masuk ke Dashboard (/dashboard), kembali ke terminal ini.")
        print("─" * 60)
        
        input("\n⌨️  Tekan [ENTER] di terminal ini jika sudah berada di halaman Dashboard... ")
        
        print("\n[*] Memeriksa sesi login dan mengambil token autentikasi...")
        
        access_token = None
        cookies = context.cookies()
        for cookie in cookies:
            if cookie.get('name') == 'access_token' and cookie.get('value'):
                access_token = cookie['value']
                break
                
        if not access_token:
            try:
                access_token = page.evaluate("() => localStorage.getItem('access_token') || sessionStorage.getItem('access_token') || ''")
            except Exception:
                pass
                
        if not access_token:
            print("⚠️ Token belum terdeteksi. Mencoba membuka /dashboard untuk verifikasi...")
            try:
                page.goto("https://portal.gofoodmerchant.co.id/dashboard", wait_until="domcontentloaded")
                time.sleep(2)
                for cookie in context.cookies():
                    if cookie.get('name') == 'access_token' and cookie.get('value'):
                        access_token = cookie['value']
                        break
            except Exception as e:
                print(f"⚠️ Error navigasi: {e}")

        if not access_token:
            print("❌ Gagal mendeteksi token login. Pastikan Anda sudah benar-benar login di browser.")
            browser.close()
            return
            
        print("🎉 Token login berhasil didapatkan!")
        
        # Simpan sesi untuk dipakai lagi di masa mendatang
        session_id = f"manual_{re.sub(r'[^a-zA-Z0-9_.-]', '_', owner_name.lower())}"
        save_gofood_session(session_id, cookies, access_token)
        
        # Tutup browser segera untuk menghemat RAM & CPU
        browser.close()
        
    # Ekstraksi cepat menggunakan curl_cffi dengan Chrome TLS impersonation
    print(f"\n⚡ [FAST EXTRACTION] Mengambil data outlet untuk '{owner_name}' via curl_cffi...")
    ok, status, hits, err = fetch_gobiz_merchants_fast(access_token, cookies=cookies)
    
    if ok and hits:
        print(f"   ✅ Berhasil menarik {len(hits)} data outlet!")
        outlets = transform_gobiz_hits(hits, owner_name, portal_name, owner_name)
        save_portal_results(outlets, owner_name, portal_name, owner_name, "manual_login", "MANUAL")
    else:
        print(f"   ❌ Gagal menarik data outlet: {err}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="GoFood Scraper")
    parser.add_argument("--type", type=str, choices=["agency", "vb", "vercel", "manual"], help="Pilih tipe sumber: 'agency', 'vb', 'vercel', atau 'manual'")
    parser.add_argument("--agency", action="store_true", help="Gunakan sumber kredensial Agency")
    parser.add_argument("--vb", action="store_true", help="Gunakan sumber kredensial VB (Virtual Brand)")
    parser.add_argument("--vercel", action="store_true", help="Gunakan sumber kredensial Vercel Sheet")
    parser.add_argument("--manual", action="store_true", help="Jalankan Manual Mode (Login manual di browser lalu scrape)")
    parser.add_argument("--vb-url", type=str, help="URL Google Sheet / CSV khusus untuk VB")
    parser.add_argument("--url", type=str, help="URL Google Sheet / CSV custom")
    parser.add_argument("--portal", type=str, help="Portal name or number (e.g. 1, 2-3, all)")
    parser.add_argument("--owner", type=str, help="Filter nama owner tertentu untuk ditarik")
    parser.add_argument("--all", action="store_true", help="Process all portals from CSV")
    parser.add_argument("--resume", action="store_true", help="Skip portals that have already been scraped")
    parser.add_argument("--combine-only", "--generate-master", dest="combine_only", action="store_true", help="Hanya gabungkan cache JSON menjadi file Master tanpa melakukan scraping")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    args = parser.parse_args()

    print("="*60)
    print("  🚀 GOFOOD OUTLET SCRAPER (LOGIN VIA OTP & PASSWORD)")
    print("="*60)

    source_type = "agency"
    custom_url = args.url or args.vb_url

    if args.manual or args.type == "manual":
        run_manual_mode()
        return
    elif args.vb or args.type == "vb":
        source_type = "vb"
    elif args.vercel or args.type == "vercel":
        source_type = "vercel"
    elif args.agency or args.type == "agency":
        source_type = "agency"
    elif not args.portal and not args.owner and not args.all and not args.combine_only:
        print("\n" + "=" * 60)
        print("  PILIH SUMBER KREDENSIAL GOFOOD")
        print("=" * 60)
        print("  [1] Agency       (Master Agency Google Sheet)")
        print("  [2] VB           (Virtual Brand - Dokumen VB)")
        print("  [3] Vercel Sheet (Live CSV Vercel Sheet)")
        print("  [4] Manual Mode  (Login manual di browser)")
        print("=" * 60)
        pilihan = input("Pilih [1/2/3/4] (Default: 1 - Agency): ").strip()
        if pilihan == "2" or pilihan.lower() == "vb":
            source_type = "vb"
        elif pilihan == "3" or pilihan.lower() == "vercel":
            source_type = "vercel"
        elif pilihan == "4" or pilihan.lower() == "manual":
            run_manual_mode()
            return
        else:
            source_type = "agency"

    if args.combine_only:
        combine_master(source_type=source_type)
        return
    
    portals = get_credentials_from_sheet(source_type=source_type, custom_url=custom_url)
    target_portals = []
    is_all_selected = False
    
    if args.resume:
        for p in portals:
            cache_p = p.get('cache_file', '')
            if not cache_p or not os.path.exists(cache_p):
                target_portals.append(p)
        print(f"[*] Mode Resume: {len(target_portals)} portal tersisa untuk ditarik (dari total {len(portals)} portal).")
    elif args.owner:
        owner_clean = args.owner.strip().lower()
        for p in portals:
            p_owner = str(p.get('owner', '')).strip().lower()
            if p_owner == owner_clean or owner_clean in p_owner:
                target_portals.append(p)
        print(f"[*] Filter Owner '{args.owner}': Ditemukan {len(target_portals)} akun portal GoFood.")
    elif args.all:
        target_portals = portals
        is_all_selected = True
    elif args.portal:
        choice = args.portal.strip().lower()
        if choice in ('0', 'all'):
            target_portals = portals
            is_all_selected = True
        elif choice == 'm':
            combine_master(source_type=source_type)
            return
        else:
            for p in portals:
                if p['portal'].lower() == choice or p['email'].lower() == choice:
                    target_portals.append(p)
            if not target_portals:
                selected_indices = parse_selection(choice, len(portals))
                for idx in selected_indices:
                    target_portals.append(portals[idx])
    elif portals:
        if source_type in ("agency", "vercel"):
            # Kelompokkan berdasarkan Owner untuk tampilan Agency & Vercel
            owners = {}
            for p in portals:
                o_name = p.get('owner') or p.get('brand', 'Unknown')
                if o_name not in owners:
                    owners[o_name] = []
                owners[o_name].append(p)
                
            owner_names = list(owners.keys())
            
            print(f"\nDaftar Pemilik (Owner) GoFood [{source_type.upper()}]:")
            print(f"  [0] Pilih Semua ({len(owner_names)} Pemilik, {len(portals)} Akun Login Unik)")
            for idx, o_name in enumerate(owner_names):
                plist = owners[o_name]
                emails_summary = ", ".join([p['email'] for p in plist])
                print(f"  [{idx+1}] {o_name} ({len(plist)} akun login: {emails_summary})")
            print("  [m] Generate Master dari Cache JSON")
            print("  [n] Input Email Baru secara manual")
            
            choice = input("\nPilih pemilik / owner (contoh: 1, 2-3, all, m, n): ").strip().lower()
            if choice == 'm':
                combine_master(source_type=source_type)
                return
            elif choice == 'n':
                email = input("\nMasukkan Email Akun GoFood: ").strip()
                password = input("Masukkan Password (kosongkan jika via OTP): ").strip()
                if email:
                    target_portals.append({
                        'owner': 'Manual',
                        'brand': 'Manual',
                        'email': email,
                        'emails': [email],
                        'password': password,
                        'portal': "Manual Input"
                    })
            elif choice in ('0', 'all'):
                target_portals = portals
                is_all_selected = True
            else:
                selected_indices = parse_selection(choice, len(owner_names))
                for idx in selected_indices:
                    target_portals.extend(owners[owner_names[idx]])
        else:
            # Tampilan per portal untuk VB
            print(f"\nDaftar Portal GoFood [VB]:")
            print(f"  [0] Pilih Semua ({len(portals)} Portal)")
            for idx, p in enumerate(portals):
                login_method = "Password" if p.get("password") else "OTP"
                print(f"  [{idx+1}] {p['portal']} ({p['email']}) - Owner: {p.get('owner', '-')} [{login_method}]")
            print("  [m] Generate Master dari Cache JSON")
            print("  [n] Input Email Baru secara manual")
            
            choice = input("\nPilih portal (contoh: 1, 2-3, all, m, n): ").strip().lower()
            if choice == 'm':
                combine_master(source_type=source_type)
                return
            elif choice == 'n':
                email = input("\nMasukkan Email Akun GoFood: ").strip()
                password = input("Masukkan Password (kosongkan jika via OTP): ").strip()
                if email:
                    target_portals.append({
                        'owner': 'VB',
                        'brand': 'Manual',
                        'email': email,
                        'emails': [email],
                        'password': password,
                        'portal': "Manual Input"
                    })
            elif choice in ('0', 'all'):
                target_portals = portals
                is_all_selected = True
            else:
                selected_indices = parse_selection(choice, len(portals))
                for idx in selected_indices:
                    target_portals.append(portals[idx])
                
    if not target_portals:
        print("Tidak ada portal yang dipilih untuk diproses.")
        return

    with sync_playwright() as p:
        print("\n[*] Membuka browser Playwright...")
        browser = p.chromium.launch(
            headless=args.headless, 
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-infobars',
                '--no-sandbox',
            ]
        )
        
        banned_emails = set()
        completed_portals = set()
        queue = list(target_portals)
        postponed_portals = []
        is_retry_round = False
        
        # Kolektor outlet yang dikelompokkan per Owner
        owner_outlets_collected = {}

        while queue:
            target = queue.pop(0)
            portal_name_str = target.get('portal', 'Manual Input')
            owner_name_str = target.get('owner') or ('VB' if source_type == 'vb' else target.get('brand', 'Unknown'))
            
            if portal_name_str in completed_portals:
                continue

            emails_to_try = target.get('emails', [target.get('email')])
            out_file = target.get('output')

            # Cek apakah semua email saat ini tercatat di banned_emails
            active_emails = [e for e in emails_to_try if e not in banned_emails]
            if not active_emails:
                print(f"\n⏭️ [SKIP SEMENTARA] Portal '{portal_name_str}' dilewati karena seluruh emailnya ({', '.join(emails_to_try)}) sedang ter-limit. Akan dicoba ulang di akhir.")
                if target not in postponed_portals:
                    postponed_portals.append(target)
                
                # Cek jika antrean utama habis tapi ada yang tertunda
                if not queue and postponed_portals and not is_retry_round:
                    print("\n" + "=" * 60)
                    print(f"  🔄 {len(postponed_portals)} PORTAL TERTUNDA AKIBAT LIMIT EMAIL")
                    print("=" * 60)
                    for p in postponed_portals:
                        print(f"  - {p['portal']} ({', '.join(p.get('emails', []))})")
                    print("\n[*] Menunggu jeda 10 detik sebelum memulai putaran ulang portal yang tertunda...")
                    time.sleep(10)
                    queue = list(postponed_portals)
                    postponed_portals = []
                    banned_emails.clear()
                    is_retry_round = True
                continue

            print(f"\n{'='*60}")
            print(f"  🏢 Owner   : {owner_name_str}")
            print(f"  📍 Portal  : {portal_name_str}")
            print(f"  📧 Emails  : {', '.join(emails_to_try)}")
            print(f"{'='*60}")

            # ----------------------------------------------------
            # 1. FAST PATH: Uji apakah sesi aktif ada di disk dan valid via curl_cffi
            # ----------------------------------------------------
            fast_path_success = False
            for email in emails_to_try:
                cached_data = load_gofood_session(email)
                if cached_data and cached_data.get('access_token'):
                    print(f"   ⚡ [FAST PATH] Menguji token sesi aktif untuk {email} via curl_cffi...")
                    ok, status, hits, err = fetch_gobiz_merchants_fast(
                        cached_data['access_token'],
                        cookies=cached_data.get('cookies')
                    )
                    if ok and hits:
                        print(f"   🎉 [FAST PATH BERHASIL] Berhasil menarik {len(hits)} data outlet dalam < 1 detik (tanpa membuka browser)!")
                        brand_str = target.get('brand', portal_name_str.split(' - ')[0] if ' - ' in portal_name_str else portal_name_str)
                        outlets_data = transform_gobiz_hits(hits, owner_name_str, portal_name_str, brand_str)
                        
                        if outlets_data:
                            safe_cache_name = get_safe_cache_filename(portal_name_str)
                            cache_json_file = CACHE_DIR / safe_cache_name
                            cache_payload = {
                                'portal': portal_name_str,
                                'brand': brand_str,
                                'owner': owner_name_str,
                                'source_type': source_type.upper(),
                                'email': email,
                                'timestamp': datetime.datetime.now().isoformat(),
                                'outlets': outlets_data
                            }
                            try:
                                with open(cache_json_file, 'w', encoding='utf-8') as jf:
                                    json.dump(cache_payload, jf, indent=2, ensure_ascii=False)
                                print(f"   💾 Cache portal disimpan di cache/: {cache_json_file.name} ({len(outlets_data)} outlet)")
                            except Exception as e:
                                print(f"   ⚠️ Gagal menyimpan cache JSON: {e}")
                                
                            if owner_name_str not in owner_outlets_collected:
                                owner_outlets_collected[owner_name_str] = []
                            owner_outlets_collected[owner_name_str].extend(outlets_data)
                            
                        completed_portals.add(portal_name_str)
                        fast_path_success = True
                        break
                    else:
                        print(f"   ⚠️ Sesi {email} tidak valid / kedaluwarsa ({err}). Melanjutkan ke browser auth...")

            if fast_path_success:
                continue

            # ----------------------------------------------------
            # 2. BROWSER AUTH FALLBACK: Buka Playwright untuk autentikasi jika sesi kedaluwarsa
            # ----------------------------------------------------
            print(f"   🌐 [BROWSER AUTH] Membuka browser untuk autentikasi portal '{portal_name_str}'...")
            access_token = None
            session_loaded_successfully = False
            logged_in_email = None
            portal_encountered_ban = False
            cookies_collected = []

            context = browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1366, 'height': 768}
            )
            # Blokir image, media, font, dan analytics pihak ketiga agar browser jauh lebih ringan dan cepat
            context.route("**/*", lambda route: (
                route.abort() if route.request.resource_type in ["image", "media", "font"] or any(t in route.request.url for t in ["google-analytics", "analytics", "doubleclick", "clarity.ms", "hotjar", "crisp.chat", "freshchat"]) else route.continue_()
            ))

            for email_index, current_email in enumerate(emails_to_try, 1):
                if current_email in banned_emails:
                    print(f"   ⏭️ Melewati email '{current_email}' karena sedang ter-limit.")
                    continue

                print(f"\n   ➡️ [Email: {current_email}] Membuka halaman login email... (Percobaan {email_index}/{len(emails_to_try)})")
                try:
                    page = context.new_page()
                    page.goto("https://portal.gofoodmerchant.co.id/auth/login/email", wait_until="domcontentloaded")
                    time.sleep(2)

                    # Handle halaman awal
                    if page.locator(":text('Pilih Akun')").count() > 0:
                        page.locator(":text('Pilih Akun')").first.click()
                        time.sleep(1)

                    # Cek input email
                    email_input = page.locator("input[type='email'], input[name='email'], input[placeholder*='email' i]").first
                    if email_input.count() == 0:
                        email_input = page.locator("input").first

                    email_input.wait_for(state="visible", timeout=10000)
                    email_input.click()
                    time.sleep(0.5)
                    
                    # Masukkan email lengkap secara instan dan pasti
                    email_input.fill(current_email)
                    time.sleep(0.5)
                    
                    if email_input.input_value() != current_email:
                        email_input.fill("")
                        email_input.fill(current_email)
                        time.sleep(0.5)
                        
                    print(f"   📧 Email berhasil diinputkan: {email_input.input_value()}")
                    time.sleep(1)

                    portal_password = target.get('password', '').strip()
                    login_with_password = bool(portal_password)

                    submit_btn = page.locator("button:has-text('Lanjut'), button:has-text('Next'), button:has-text('Masuk'), button[type='submit']").first
                    if submit_btn.count() > 0:
                        try:
                            page.wait_for_function("btn => !btn.disabled", submit_btn.element_handle(), timeout=3000)
                        except Exception:
                            pass
                        print(f"   👉 Mengklik tombol: '{submit_btn.inner_text().strip()}'")
                        submit_btn.click()
                        time.sleep(2)

                    is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                    if is_blocked:
                        print(f"\n   🚫 [TERBLOKIR / LIMIT] Email '{current_email}' terdeteksi dibatasi/diblokir: '{block_msg}'")
                        banned_emails.add(current_email)
                        portal_encountered_ban = True
                        try:
                            page.close()
                        except Exception:
                            pass
                        continue

                    is_vb = (
                        str(target.get('source_type', '')).upper() == 'VB' or 
                        str(target.get('owner', '')).upper() == 'VB' or 
                        str(source_type).lower() == 'vb'
                    )

                    if login_with_password:
                        print(f"   🔑 [Password Login] Menggunakan kata sandi untuk {current_email}...")
                        
                        pass_input = page.locator("input[type='password'], input[name='password']").first
                        if pass_input.count() == 0:
                            pass_option_btn = page.locator("button:has-text('Masuk dengan kata sandi'), button:has-text('Gunakan kata sandi'), button:has-text('Masuk dengan Password')").first
                            if pass_option_btn.count() > 0 and pass_option_btn.is_visible():
                                pass_option_btn.click()
                                time.sleep(1)
                            pass_input = page.locator("input[type='password'], input[name='password']").first

                        if pass_input.count() > 0:
                            try:
                                pass_input.wait_for(state="visible", timeout=5000)
                            except Exception:
                                pass
                            pass_input.fill(portal_password)
                            time.sleep(0.5)
                            print("   ✅ Kata sandi berhasil dimasukkan.")

                            masuk_btn = page.locator("button:has-text('Masuk'), button:has-text('Log in'), button:has-text('Login'), button[type='submit']").first
                            if masuk_btn.count() > 0:
                                print(f"   👉 Mengklik tombol submit: '{masuk_btn.inner_text().strip()}'")
                                masuk_btn.click()
                                time.sleep(2)

                            for _ in range(15):
                                # Cek keberadaan token di cookies
                                found_token = None
                                cookies_collected = context.cookies()
                                for cookie in cookies_collected:
                                    if cookie.get('name') == 'access_token' and cookie.get('value'):
                                        found_token = cookie['value']
                                        break

                                # Cek localStorage jika belum ada di cookies
                                if not found_token:
                                    try:
                                        found_token = page.evaluate("() => localStorage.getItem('access_token') || sessionStorage.getItem('access_token') || ''")
                                    except Exception:
                                        pass

                                current_url = page.url
                                is_success_url = any(u in current_url for u in ["/dashboard", "/home", "/outlets", "/portal", "/merchants"]) or (
                                    "auth/login" not in current_url and "portal.gofoodmerchant.co.id" in current_url and current_url.strip("/") != "https://portal.gofoodmerchant.co.id/auth/login/email"
                                )

                                if found_token or is_success_url:
                                    print(f"\n🎉 LOGIN SUKSES untuk {portal_name_str} ({current_email})!")
                                    access_token = found_token or ""
                                    if not access_token:
                                        for cookie in context.cookies():
                                            if cookie.get('name') == 'access_token' and cookie.get('value'):
                                                access_token = cookie['value']
                                                break
                                    if not access_token:
                                        try:
                                            access_token = page.evaluate("() => localStorage.getItem('access_token') || sessionStorage.getItem('access_token') || ''")
                                        except Exception:
                                            pass
                                    if not access_token:
                                        try:
                                            page.goto("https://portal.gofoodmerchant.co.id/dashboard", wait_until="domcontentloaded")
                                            time.sleep(2)
                                            for cookie in context.cookies():
                                                if cookie.get('name') == 'access_token' and cookie.get('value'):
                                                    access_token = cookie['value']
                                                    break
                                        except Exception:
                                            pass

                                    save_gofood_session(current_email, context.cookies(), access_token)
                                    session_loaded_successfully = True
                                    logged_in_email = current_email
                                    break

                                # Deteksi error di layar jika password salah
                                is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                                if is_blocked:
                                    print(f"   🚫 [LOGIN GAGAL] Pesan di layar: '{block_msg}'")
                                    break

                                time.sleep(1)

                    if not session_loaded_successfully:
                        # Akun VB tidak memerlukan dan tidak menggunakan OTP
                        if is_vb:
                            print(f"   ⚠️ [VB] Login kata sandi untuk '{current_email}' belum berhasil / sesi tidak terdeteksi. VB tidak memerlukan OTP.")
                            try:
                                page.close()
                            except Exception:
                                pass
                            continue

                        otp_option_btn = page.locator("button:has-text('email'), button:has-text('OTP'), :text('Kirim kode via email'), :text('Kirim OTP')").first
                        if otp_option_btn.count() > 0:
                            otp_option_btn.click()
                            time.sleep(1.5)
                            is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                            if is_blocked:
                                print(f"\n   🚫 [TERBLOKIR / LIMIT] Permintaan OTP untuk '{current_email}' dibatasi: '{block_msg}'")
                                banned_emails.add(current_email)
                                portal_encountered_ban = True
                                try:
                                    page.close()
                                except Exception:
                                    pass
                                continue

                        print("   ⏳ Menunggu OTP masuk ke Gmail...")
                        otp_code = tunggu_otp_terbaru(OTP_ENDPOINT_URL, action="getOtpEmail", label_email=GMAIL_OTP_LABEL, timeout_detik=90, page=page)

                        if otp_code:
                            print(f"   🔑 Memasukkan kode OTP: {otp_code}")
                            isi_kode_otp(page, otp_code)

                            time.sleep(2)
                            for attempt in range(20):
                                cookies_collected = context.cookies()
                                found_token = None
                                for cookie in cookies_collected:
                                    if cookie.get('name') == 'access_token' and cookie.get('value'):
                                        found_token = cookie['value']
                                        break
                                        
                                current_url = page.url
                                if found_token or "/dashboard" in current_url or "/home" in current_url:
                                    print(f"\n🎉 LOGIN SUKSES untuk {portal_name_str} ({current_email})!")
                                    access_token = found_token or ""
                                    if not access_token:
                                        try:
                                            access_token = page.evaluate("() => localStorage.getItem('access_token') || sessionStorage.getItem('access_token') || ''")
                                        except Exception:
                                            pass
                                            
                                    save_gofood_session(current_email, cookies_collected, access_token)
                                    session_loaded_successfully = True
                                    logged_in_email = current_email
                                    break
                                    
                                if attempt in (3, 6, 10):
                                    is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                                    if is_blocked:
                                        print(f"   🚫 [TERBLOKIR / OTP SALAH] Pesan di layar: '{block_msg}'")
                                        break
                                    trigger_submit_otp(page)
                                    
                                time.sleep(1)
                        else:
                            is_blocked, block_msg = deteksi_pesan_blokir_atau_error(page)
                            if is_blocked:
                                print(f"   🚫 [TERBLOKIR / LIMIT] Email '{current_email}' terblokir: '{block_msg}'")
                            else:
                                print(f"   ❌ Gagal mendapatkan OTP untuk {current_email}.")
                            banned_emails.add(current_email)
                            portal_encountered_ban = True

                    if session_loaded_successfully:
                        break
                except Exception as e:
                    print(f"   ⚠️ Terjadi kesalahan saat proses login {current_email}: {e}")

            if session_loaded_successfully and access_token:
                completed_portals.add(portal_name_str)
                
                # Ekstraksi cepat menggunakan curl_cffi dengan Chrome TLS impersonation
                print("   ⚡ [FAST EXTRACTION] Menarik data outlet dari GoBiz API via curl_cffi...")
                ok, status, hits, err = fetch_gobiz_merchants_fast(access_token, cookies=cookies_collected)
                
                if not ok or not hits:
                    print(f"   ⚠️ curl_cffi gagal ({err}), mencoba fallback via browser evaluation...")
                    try:
                        if page.is_closed():
                            page = context.new_page()
                            page.goto("https://portal.gofoodmerchant.co.id/dashboard", wait_until="domcontentloaded")
                            time.sleep(1)

                        token_h = f"Bearer {access_token}" if not access_token.startswith("Bearer ") else access_token
                        payload_str = json.dumps({"from": 0, "size": 1000})
                        api_response = page.evaluate("""async ({token, payload}) => {
                            try {
                                const res = await fetch('https://api.gobiz.co.id/v1/merchants/search', {
                                    method: 'POST',
                                    headers: {
                                        'Accept': 'application/json, text/plain, */*',
                                        'Authentication-Type': 'go-id',
                                        'Authorization': token,
                                        'Content-Type': 'application/json'
                                    },
                                    body: payload
                                });
                                return await res.json();
                            } catch (e) {
                                return { error: e.message };
                            }
                        }""", {"token": token_h, "payload": payload_str})
                        
                        if api_response and 'hits' in api_response:
                            hits = api_response['hits']
                        elif api_response and 'data' in api_response:
                            hits = api_response['data']
                    except Exception as e:
                        print(f"   ⚠️ Fallback browser juga mengalami error: {e}")
                        
                if hits:
                    print(f"   ✅ Berhasil menarik {len(hits)} data outlet!")
                    brand_str = target.get('brand', portal_name_str.split(' - ')[0] if ' - ' in portal_name_str else portal_name_str)
                    outlets_data = transform_gobiz_hits(hits, owner_name_str, portal_name_str, brand_str)
                    
                    if outlets_data:
                        safe_cache_name = get_safe_cache_filename(portal_name_str)
                        cache_json_file = CACHE_DIR / safe_cache_name
                        cache_payload = {
                            'portal': portal_name_str,
                            'brand': brand_str,
                            'owner': owner_name_str,
                            'source_type': source_type.upper(),
                            'email': logged_in_email or target.get('email', ''),
                            'timestamp': datetime.datetime.now().isoformat(),
                            'outlets': outlets_data
                        }
                        try:
                            with open(cache_json_file, 'w', encoding='utf-8') as jf:
                                json.dump(cache_payload, jf, indent=2, ensure_ascii=False)
                            print(f"   💾 Cache portal disimpan di cache/: {cache_json_file.name} ({len(outlets_data)} outlet)")
                        except Exception as e:
                            print(f"   ⚠️ Gagal menyimpan cache JSON: {e}")
                        
                        if owner_name_str not in owner_outlets_collected:
                            owner_outlets_collected[owner_name_str] = []
                        owner_outlets_collected[owner_name_str].extend(outlets_data)
                else:
                    print("   ⚠️ Tidak ada data outlet yang ditemukan.")
            else:
                print(f"❌ Gagal login ke portal {portal_name_str}.")
                if target not in postponed_portals:
                    postponed_portals.append(target)

            try:
                context.close()
            except Exception:
                pass

        print("\n✅ Semua portal yang dipilih telah selesai diproses.")
        browser.close()
        
        # Simpan file per-Owner untuk semua outlet yang berhasil dikumpulkan
        if owner_outlets_collected:
            import pandas as pd
            timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")
            for owner_name, o_outlets in owner_outlets_collected.items():
                if o_outlets:
                    owner_df = pd.DataFrame(o_outlets)
                    # Deduplikasi berdasarkan Store ID jika owner memiliki beberapa kredensial dengan outlet yang sama
                    if 'Store ID' in owner_df.columns:
                        owner_df.drop_duplicates(subset=["Store ID"], keep="first", inplace=True)
                    
                    owner_clean = "".join(c for c in str(owner_name) if c.isalnum() or c in " ._-").strip()
                    owner_file = OUTPUT_DIR / f"{timestamp_str} {owner_clean}.xlsx"
                    save_formatted_excel(owner_df, str(owner_file))
                    print(f"\n   💾 File Owner '{owner_name}' berhasil dibuat di output/: {owner_file.name} (Total: {len(owner_df)} outlet)")
                    
                    if APP_SCRIPT_URL:
                        upload_to_drive(str(owner_file))
        
    # Selalu perbarui file master dari cache yang terkumpul
    combine_master(source_type=source_type)


if __name__ == "__main__":
    main()

